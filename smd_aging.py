"""
smd_aging.py - Balance Agee Clients / Fournisseurs
SMD Consulting - Superviseur IA Comptable & Fiscal UEMOA
Version 1.0

Usage dans app.py :
    from smd_aging import page_balance_agee
    elif page == "Balance Agee":
        page_balance_agee()

Format du fichier attendu (Excel/CSV) :
    Colonnes requises : Tiers (nom), Date (echeance ou facture), Montant (solde du)
    Colonnes optionnelles : Compte, Reference, Libelle
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import date, datetime
from io import BytesIO


# ─────────────────────────────────────────────
#  CONSTANTES
# ─────────────────────────────────────────────

TRANCHES = [
    {"label": "0 - 30 jours",  "min": 0,   "max": 30,  "couleur": "#1E8449", "bg": "#d4edda"},
    {"label": "31 - 60 jours", "min": 31,  "max": 60,  "couleur": "#F39C12", "bg": "#fff3cd"},
    {"label": "61 - 90 jours", "min": 61,  "max": 90,  "couleur": "#E67E22", "bg": "#ffd699"},
    {"label": "+ 90 jours",    "min": 91,  "max": 9999,"couleur": "#C0392B", "bg": "#f5c6cb"},
]

COMPTES_CLIENTS     = ["411", "4111", "4118"]
COMPTES_FOURNISSEURS = ["401", "4011", "4018"]


# ─────────────────────────────────────────────
#  PARSING DU FICHIER
# ─────────────────────────────────────────────

def _detecter_colonnes(df: pd.DataFrame) -> dict:
    """Detecte automatiquement les colonnes pertinentes."""
    mapping = {}
    cols_lower = {str(c).lower().strip(): c for c in df.columns}

    # Tiers / nom
    for k in ["tiers", "client", "fournisseur", "nom", "raison sociale", "libelle tiers", "name"]:
        if k in cols_lower:
            mapping["tiers"] = cols_lower[k]
            break

    # Date
    for k in ["date echeance", "echeance", "date", "date facture", "due date", "date limite"]:
        if k in cols_lower:
            mapping["date"] = cols_lower[k]
            break

    # Montant
    for k in ["montant", "solde", "solde du", "montant du", "balance", "amount", "debit", "credit"]:
        if k in cols_lower:
            mapping["montant"] = cols_lower[k]
            break

    # Reference
    for k in ["reference", "ref", "numero", "facture", "piece"]:
        if k in cols_lower:
            mapping["reference"] = cols_lower[k]
            break

    # Compte
    for k in ["compte", "n compte", "numero compte"]:
        if k in cols_lower:
            mapping["compte"] = cols_lower[k]
            break

    return mapping


def _parser_fichier(fichier, date_ref: date, mapping: dict) -> pd.DataFrame:
    """Parse le fichier uploade et calcule l'anciennete des creances."""
    try:
        if fichier.name.endswith(".xlsx") or fichier.name.endswith(".xls"):
            df_raw = pd.read_excel(fichier)
        else:
            try:
                df_raw = pd.read_csv(fichier, encoding="utf-8", sep=None, engine="python")
            except Exception:
                fichier.seek(0)
                df_raw = pd.read_csv(fichier, encoding="latin-1", sep=None, engine="python")
    except Exception as e:
        return pd.DataFrame(), str(e)

    df = pd.DataFrame()

    # Tiers
    if "tiers" in mapping:
        df["Tiers"] = df_raw[mapping["tiers"]].astype(str).str.strip()
    else:
        df["Tiers"] = "Inconnu"

    # Reference
    if "reference" in mapping:
        df["Reference"] = df_raw[mapping["reference"]].astype(str)
    else:
        df["Reference"] = ""

    # Compte
    if "compte" in mapping:
        df["Compte"] = df_raw[mapping["compte"]].astype(str)
    else:
        df["Compte"] = ""

    # Date
    if "date" in mapping:
        df["Date"] = pd.to_datetime(df_raw[mapping["date"]], dayfirst=True, errors="coerce")
    else:
        df["Date"] = pd.NaT

    # Montant
    if "montant" in mapping:
        df["Montant"] = pd.to_numeric(
            df_raw[mapping["montant"]].astype(str).str.replace(" ", "").str.replace(",", "."),
            errors="coerce"
        ).fillna(0)
    else:
        df["Montant"] = 0.0

    # Filtrer montants nuls et negatifs (avances, avoirs)
    df = df[df["Montant"] > 0].copy()

    # Calcul anciennete
    date_ref_ts = pd.Timestamp(date_ref)
    df["Jours"] = (date_ref_ts - df["Date"]).dt.days.fillna(0).astype(int)
    df["Jours"] = df["Jours"].clip(lower=0)

    # Tranche
    def _tranche(j):
        for t in TRANCHES:
            if t["min"] <= j <= t["max"]:
                return t["label"]
        return TRANCHES[-1]["label"]

    df["Tranche"] = df["Jours"].apply(_tranche)

    return df, None


# ─────────────────────────────────────────────
#  CALCUL DE LA BALANCE AGEE
# ─────────────────────────────────────────────

def _calculer_balance_agee(df: pd.DataFrame) -> pd.DataFrame:
    """Calcule le pivot de la balance agee par tiers et tranche."""
    if df.empty:
        return pd.DataFrame()

    pivot = df.pivot_table(
        index="Tiers",
        columns="Tranche",
        values="Montant",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    # S'assurer que toutes les tranches sont presentes
    for t in TRANCHES:
        if t["label"] not in pivot.columns:
            pivot[t["label"]] = 0.0

    # Total par tiers
    cols_tranches = [t["label"] for t in TRANCHES]
    pivot["TOTAL"] = pivot[cols_tranches].sum(axis=1)

    # Trier par total decroissant
    pivot = pivot.sort_values("TOTAL", ascending=False).reset_index(drop=True)

    # Reordonner les colonnes
    pivot = pivot[["Tiers"] + cols_tranches + ["TOTAL"]]

    return pivot


# ─────────────────────────────────────────────
#  GRAPHIQUES
# ─────────────────────────────────────────────

def _chart_repartition(pivot: pd.DataFrame) -> go.Figure:
    """Graphique en barres empilees par tranche."""
    cols = [t["label"] for t in TRANCHES]
    couleurs = [t["couleur"] for t in TRANCHES]

    top10 = pivot.nlargest(10, "TOTAL")

    fig = go.Figure()
    for col, couleur in zip(cols, couleurs):
        fig.add_trace(go.Bar(
            name=col,
            x=top10["Tiers"],
            y=top10[col],
            marker_color=couleur,
        ))

    fig.update_layout(
        barmode="stack",
        title="Top 10 tiers — Repartition par anciennete",
        xaxis_title="Tiers",
        yaxis_title="Montant (FCFA)",
        legend_title="Tranche",
        height=400,
        plot_bgcolor="white",
        paper_bgcolor="white",
    )
    return fig


def _chart_camembert(totaux: dict) -> go.Figure:
    """Graphique circulaire de la repartition globale."""
    labels = list(totaux.keys())
    values = list(totaux.values())
    couleurs = [t["couleur"] for t in TRANCHES]

    fig = go.Figure(go.Pie(
        labels=labels,
        values=values,
        marker_colors=couleurs,
        hole=0.4,
        textinfo="label+percent",
    ))
    fig.update_layout(
        title="Repartition globale des creances",
        height=350,
        paper_bgcolor="white",
    )
    return fig


# ─────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────

def _export_balance_agee_excel(pivot: pd.DataFrame, df_detail: pd.DataFrame,
                                type_tiers: str, date_ref: date, totaux: dict) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Feuille 1 : Balance agee ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Balance Agee"

    BLEU  = "1A5276"
    BLANC = "FFFFFF"
    cols_tranches = [t["label"] for t in TRANCHES]
    couleurs_xl = ["1E8449", "F39C12", "E67E22", "C0392B"]

    # Titre
    nb_cols = len(cols_tranches) + 2
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    ws1.cell(1, 1).value = "BALANCE AGEE " + type_tiers.upper() + " — Au " + date_ref.strftime("%d/%m/%Y")
    ws1.cell(1, 1).font = Font(bold=True, color=BLANC, size=13)
    ws1.cell(1, 1).fill = PatternFill("solid", fgColor=BLEU)
    ws1.cell(1, 1).alignment = Alignment(horizontal="center")

    # En-tetes
    headers = ["Tiers"] + cols_tranches + ["TOTAL"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(2, col, value=h)
        c.font = Font(bold=True, color=BLANC)
        c.fill = PatternFill("solid", fgColor=BLEU if col == 1 or col == nb_cols else couleurs_xl[col - 2])
        c.alignment = Alignment(horizontal="center")

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Donnees
    for i, (_, row) in enumerate(pivot.iterrows()):
        r = 3 + i
        bg = "F2F2F2" if i % 2 == 0 else "FFFFFF"
        vals = [row["Tiers"]] + [row[c] for c in cols_tranches] + [row["TOTAL"]]
        for col, v in enumerate(vals, 1):
            c = ws1.cell(r, col, value=v)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = thin
            if col > 1 and isinstance(v, (int, float)):
                c.number_format = "#,##0"

    # Ligne de total
    r_tot = 3 + len(pivot)
    ws1.cell(r_tot, 1, value="TOTAL GENERAL").font = Font(bold=True)
    ws1.cell(r_tot, 1).fill = PatternFill("solid", fgColor=BLEU)
    ws1.cell(r_tot, 1).font = Font(bold=True, color=BLANC)
    for col, t_label in enumerate(cols_tranches, 2):
        c = ws1.cell(r_tot, col, value=totaux.get(t_label, 0))
        c.font = Font(bold=True, color=BLANC)
        c.fill = PatternFill("solid", fgColor=couleurs_xl[col - 2])
        c.number_format = "#,##0"
    total_global = sum(totaux.values())
    c_tot = ws1.cell(r_tot, nb_cols, value=total_global)
    c_tot.font = Font(bold=True, color=BLANC)
    c_tot.fill = PatternFill("solid", fgColor=BLEU)
    c_tot.number_format = "#,##0"

    ws1.column_dimensions["A"].width = 35
    for i in range(2, nb_cols + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 18

    # ── Feuille 2 : Detail ────────────────────────────────────────
    ws2 = wb.create_sheet("Detail")
    ws2.merge_cells("A1:G1")
    ws2.cell(1, 1, value="DETAIL DES ECHEANCES — " + type_tiers.upper()).font = Font(bold=True, color=BLANC)
    ws2.cell(1, 1).fill = PatternFill("solid", fgColor=BLEU)
    ws2.cell(1, 1).alignment = Alignment(horizontal="center")

    h2 = ["Tiers", "Reference", "Compte", "Date", "Montant (FCFA)", "Jours", "Tranche"]
    for col, h in enumerate(h2, 1):
        c = ws2.cell(2, col, value=h)
        c.font = Font(bold=True, color=BLANC)
        c.fill = PatternFill("solid", fgColor=BLEU)

    couleurs_tr = {t["label"]: t["bg"].replace("#", "") for t in TRANCHES}

    for i, (_, row) in enumerate(df_detail.iterrows()):
        r = 3 + i
        bg = couleurs_tr.get(row.get("Tranche", ""), "FFFFFF")
        vals = [
            row.get("Tiers", ""),
            row.get("Reference", ""),
            row.get("Compte", ""),
            row["Date"].strftime("%d/%m/%Y") if pd.notna(row.get("Date")) else "",
            row.get("Montant", 0),
            int(row.get("Jours", 0)),
            row.get("Tranche", ""),
        ]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(r, col, value=v)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = thin
            if col == 5:
                c.number_format = "#,##0"

    for col, w in zip("ABCDEFG", [30, 18, 12, 13, 18, 8, 16]):
        ws2.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
#  PAGE STREAMLIT
# ─────────────────────────────────────────────

def _fmt(n: float) -> str:
    return f"{n:,.0f} FCFA".replace(",", " ")


def page_balance_agee():
    st.title("Balance Agee Clients / Fournisseurs")
    st.markdown("Analysez l'anciennete de vos creances clients et dettes fournisseurs par tranches de retard.")
    st.divider()

    # ── Parametres ────────────────────────────────────────────────
    col1, col2, col3 = st.columns(3)
    with col1:
        type_tiers = st.selectbox("Type de tiers", ["Clients (411)", "Fournisseurs (401)"])
    with col2:
        date_ref = st.date_input("Date de reference", value=date.today())
    with col3:
        devise = st.selectbox("Devise", ["FCFA", "EUR", "USD"])

    # ── Upload ────────────────────────────────────────────────────
    st.subheader("Importer les echeances")
    st.caption("Fichier Excel ou CSV avec colonnes : Tiers, Date (echeance), Montant (solde du). Reference et Compte sont optionnels.")

    with st.expander("Exemple de format attendu"):
        st.dataframe(pd.DataFrame({
            "Tiers":     ["SOCIETE ALPHA", "BETA SARL", "GAMMA GROUP"],
            "Reference": ["F-001",         "F-002",     "F-003"],
            "Date":      ["15/02/2026",    "01/01/2026","15/11/2025"],
            "Montant":   [500000,          1200000,     850000],
        }), hide_index=True, use_container_width=True)

    fichier = st.file_uploader(
        "Fichier echeances (Excel ou CSV)",
        type=["xlsx", "xls", "csv"],
        key="upload_aging"
    )

    if not fichier:
        st.info("Uploadez votre fichier pour demarrer l'analyse.")
        return

    # ── Mapping des colonnes ──────────────────────────────────────
    try:
        if fichier.name.endswith(".xlsx") or fichier.name.endswith(".xls"):
            df_preview = pd.read_excel(fichier)
        else:
            df_preview = pd.read_csv(fichier, encoding="utf-8", sep=None, engine="python")
        fichier.seek(0)
    except Exception as e:
        st.error("Erreur lecture fichier : " + str(e))
        return

    with st.expander("Apercu du fichier importe"):
        st.dataframe(df_preview.head(5), use_container_width=True, hide_index=True)

    mapping = _detecter_colonnes(df_preview)

    # Permettre a l'utilisateur de corriger le mapping si detection incorrecte
    with st.expander("Correspondance des colonnes (verifier si necessaire)"):
        all_cols = ["-- Non disponible --"] + list(df_preview.columns)
        col_tiers = st.selectbox("Colonne Tiers (nom client/fournisseur)",
                                  all_cols,
                                  index=all_cols.index(mapping["tiers"]) if "tiers" in mapping and mapping["tiers"] in all_cols else 0)
        col_date  = st.selectbox("Colonne Date (echeance ou facture)",
                                  all_cols,
                                  index=all_cols.index(mapping["date"]) if "date" in mapping and mapping["date"] in all_cols else 0)
        col_mnt   = st.selectbox("Colonne Montant (solde du)",
                                  all_cols,
                                  index=all_cols.index(mapping["montant"]) if "montant" in mapping and mapping["montant"] in all_cols else 0)
        col_ref   = st.selectbox("Colonne Reference (optionnel)",
                                  all_cols,
                                  index=all_cols.index(mapping["reference"]) if "reference" in mapping and mapping["reference"] in all_cols else 0)

        if col_tiers != "-- Non disponible --":
            mapping["tiers"] = col_tiers
        if col_date != "-- Non disponible --":
            mapping["date"] = col_date
        if col_mnt != "-- Non disponible --":
            mapping["montant"] = col_mnt
        if col_ref != "-- Non disponible --":
            mapping["reference"] = col_ref

    if st.button("Analyser la balance agee", type="primary", use_container_width=True):
        if "date" not in mapping or "montant" not in mapping:
            st.error("Impossible de detecter les colonnes Date et Montant. Verifiez la correspondance ci-dessus.")
            return

        with st.spinner("Calcul en cours..."):
            result = _parser_fichier(fichier, date_ref, mapping)
            if isinstance(result, tuple):
                df, erreur = result
            else:
                df, erreur = result, None

            if erreur:
                st.error("Erreur parsing : " + str(erreur))
                return

            if df.empty:
                st.warning("Aucune ligne valide trouvee. Verifiez le format du fichier.")
                return

            pivot = _calculer_balance_agee(df)
            cols_tranches = [t["label"] for t in TRANCHES]
            totaux = {t["label"]: float(pivot[t["label"]].sum()) for t in TRANCHES}
            total_global = sum(totaux.values())

        # ── KPIs ──────────────────────────────────────────────────
        st.divider()
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Total " + type_tiers[:7], _fmt(total_global).replace("FCFA", devise))
        k2.metric("0-30 jours (sain)",   _fmt(totaux[TRANCHES[0]["label"]]).replace("FCFA", devise))
        k3.metric("31-60 jours",         _fmt(totaux[TRANCHES[1]["label"]]).replace("FCFA", devise))
        k4.metric("61-90 jours",         _fmt(totaux[TRANCHES[2]["label"]]).replace("FCFA", devise))
        pct_90 = (totaux[TRANCHES[3]["label"]] / total_global * 100) if total_global > 0 else 0
        k5.metric("+90 jours (critique)",
                  _fmt(totaux[TRANCHES[3]["label"]]).replace("FCFA", devise),
                  delta=f"{pct_90:.1f}% du total",
                  delta_color="inverse")

        # Alerte si +90j depasse 30%
        if pct_90 > 30:
            st.error("ALERTE — " + f"{pct_90:.1f}%" + " des creances ont plus de 90 jours ! Risque d'irrecouvrabilite eleve.")
        elif pct_90 > 15:
            st.warning("ATTENTION — " + f"{pct_90:.1f}%" + " des creances depassent 90 jours. Relances prioritaires a lancer.")

        # ── Graphiques ────────────────────────────────────────────
        st.divider()
        gc1, gc2 = st.columns([2, 1])
        with gc1:
            st.plotly_chart(_chart_repartition(pivot), use_container_width=True)
        with gc2:
            st.plotly_chart(_chart_camembert(totaux), use_container_width=True)

        # ── Tableau balance agee ──────────────────────────────────
        st.subheader("Balance agee par tiers")

        def _style_pivot(row):
            pct = (row["+90j"] / row["TOTAL"] * 100) if row["TOTAL"] > 0 else 0
            if pct > 50:
                return ["background-color:#f5c6cb"] * len(row)
            elif pct > 25:
                return ["background-color:#ffd699"] * len(row)
            else:
                return [""] * len(row)

        pivot_display = pivot.copy()
        pivot_display.columns = ["Tiers", "0-30j", "31-60j", "61-90j", "+90j", "TOTAL"]
        for c in ["0-30j", "31-60j", "61-90j", "+90j", "TOTAL"]:
            pivot_display[c] = pivot_display[c].apply(lambda x: f"{x:,.0f}".replace(",", " "))

        st.dataframe(pivot_display, use_container_width=True, hide_index=True, height=350)

        # Top risques
        pivot_risque = pivot[pivot[TRANCHES[3]["label"]] > 0].sort_values(
            TRANCHES[3]["label"], ascending=False
        ).head(5)

        if not pivot_risque.empty:
            st.subheader("Top 5 tiers les plus en retard (+90 jours)")
            for _, row in pivot_risque.iterrows():
                pct = row[TRANCHES[3]["label"]] / row["TOTAL"] * 100 if row["TOTAL"] > 0 else 0
                st.markdown(
                    "**" + str(row["Tiers"]) + "** — " +
                    _fmt(row[TRANCHES[3]["label"]]) + " en retard +90j" +
                    " (" + f"{pct:.0f}%" + " du total tiers)"
                )

        # ── Detail ────────────────────────────────────────────────
        with st.expander("Detail de toutes les echeances"):
            df_aff = df.copy()
            df_aff["Date"] = df_aff["Date"].apply(
                lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) else ""
            )
            df_aff["Montant"] = df_aff["Montant"].apply(
                lambda x: f"{x:,.0f}".replace(",", " ")
            )

            def _style_detail(row):
                bg = next((t["bg"] for t in TRANCHES if t["label"] == row.get("Tranche")), "")
                return ["background-color:" + bg if bg else ""] * len(row)

            st.dataframe(
                df_aff[["Tiers", "Reference", "Date", "Montant", "Jours", "Tranche"]].style.apply(_style_detail, axis=1),
                use_container_width=True, hide_index=True
            )

        # ── Export ────────────────────────────────────────────────
        st.divider()
        st.subheader("Telecharger la balance agee")
        buf = _export_balance_agee_excel(pivot, df, type_tiers, date_ref, totaux)
        st.download_button(
            "Telecharger Excel (Balance Agee + Detail)",
            data=buf,
            file_name="Balance_Agee_" + type_tiers[:7].replace("/", "_") + "_" + date_ref.strftime("%Y%m%d") + ".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
