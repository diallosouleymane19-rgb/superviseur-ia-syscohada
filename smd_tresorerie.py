# -*- coding: utf-8 -*-
"""
Module Tableau de Tresorerie Previsionnelle - SMD Consulting
Suivi encaissements/decaissements sur 30/60/90 jours - SYSCOHADA
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import date, timedelta

CATEGORIES_ENCAISSEMENTS = [
    "Reglements clients","Acomptes clients","Subventions / aides",
    "Remboursements TVA","Cessions d actifs","Emprunts / credits","Autres encaissements",
]

CATEGORIES_DECAISSEMENTS = [
    "Salaires et charges sociales","Fournisseurs matieres","Loyers et charges locatives",
    "Remboursement emprunts","TVA a decaisser","IS / acomptes IS",
    "Investissements","Frais generaux","Autres decaissements",
]

HORIZON_OPTIONS = {"30 jours": 30, "60 jours": 60, "90 jours": 90}

def _construire_tableau(date_debut, nb_jours, lignes, solde_initial):
    dates = [date_debut + timedelta(days=i) for i in range(nb_jours)]
    rows = []
    solde_courant = solde_initial
    for d in dates:
        enc = sum(l["montant"] for l in lignes if l["date"] == d and l["type"] == "E")
        dec = sum(l["montant"] for l in lignes if l["date"] == d and l["type"] == "D")
        solde_courant = solde_courant + enc - dec
        rows.append({"Date": d, "Encaissements": enc, "Decaissements": dec,
                     "Flux net": enc - dec, "Solde": solde_courant})
    return pd.DataFrame(rows)

def _agreger_semaine(df):
    df2 = df.copy()
    df2["Semaine"] = df2["Date"].apply(lambda d: d.strftime("S%W - %b %Y"))
    agg = df2.groupby("Semaine", sort=False).agg(
        Encaissements=("Encaissements","sum"),
        Decaissements=("Decaissements","sum"),
        Flux_net=("Flux net","sum"),
        Solde_fin=("Solde","last"),
    ).reset_index()
    agg.columns = ["Semaine","Encaissements","Decaissements","Flux net","Solde fin semaine"]
    return agg

def _chart_tresorerie(df, devise):
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df["Date"], y=df["Solde"], fill="tozeroy",
        fillcolor="rgba(30,132,73,0.15)",
        line=dict(color="#1E8449", width=2.5),
        name="Solde previsionnel",
        hovertemplate="<b>%{x|%d/%m/%Y}</b><br>Solde : %{y:,.0f} " + devise + "<extra></extra>",
    ))
    fig.add_hline(y=0, line_dash="dash", line_color="#C0392B", line_width=1.5,
                  annotation_text="Seuil zero", annotation_position="right")
    neg = df[df["Solde"] < 0]
    if not neg.empty:
        fig.add_trace(go.Scatter(
            x=neg["Date"], y=neg["Solde"], mode="markers",
            marker=dict(color="#C0392B", size=8, symbol="x"),
            name="Solde negatif",
        ))
    fig.update_layout(
        title="Evolution previsionnelle de la tresorerie",
        xaxis_title="Date", yaxis_title=f"Solde ({devise})",
        height=400, margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", y=-0.15), hovermode="x unified",
    )
    fig.update_xaxes(showgrid=True, gridcolor="#eee")
    fig.update_yaxes(showgrid=True, gridcolor="#eee", tickformat=",.0f")
    return fig

def _chart_flux(df, devise):
    df_sem = _agreger_semaine(df)
    fig = go.Figure()
    fig.add_trace(go.Bar(x=df_sem["Semaine"], y=df_sem["Encaissements"],
        name="Encaissements", marker_color="#1E8449",
        hovertemplate="%{y:,.0f} " + devise + "<extra>Encaissements</extra>"))
    fig.add_trace(go.Bar(x=df_sem["Semaine"], y=-df_sem["Decaissements"],
        name="Decaissements", marker_color="#C0392B",
        hovertemplate="%{y:,.0f} " + devise + "<extra>Decaissements</extra>"))
    fig.update_layout(
        title="Encaissements / Decaissements par semaine",
        barmode="relative", height=340,
        margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", y=-0.2), yaxis_tickformat=",.0f",
    )
    return fig

def _export_excel(df, lignes, solde_initial, devise, date_debut):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule

        df_exp = df.copy()
        df_exp["Date"] = df_exp["Date"].apply(lambda d: d.strftime("%d/%m/%Y"))
        df_exp.to_excel(writer, index=False, sheet_name="Journalier")
        ws = writer.sheets["Journalier"]
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(color="FFFFFF", bold=True)
        ws.conditional_formatting.add(
            f"E2:E{len(df)+1}",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=PatternFill("solid", fgColor="F5C6CB")))
        ws.conditional_formatting.add(
            f"E2:E{len(df)+1}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill("solid", fgColor="D4EDDA")))
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 20

        df_sem = _agreger_semaine(df)
        df_sem.to_excel(writer, index=False, sheet_name="Hebdomadaire")
        ws2 = writer.sheets["Hebdomadaire"]
        for cell in ws2[1]:
            cell.fill = PatternFill("solid", fgColor="1E8449")
            cell.font = Font(color="FFFFFF", bold=True)
        for col in ws2.columns:
            ws2.column_dimensions[get_column_letter(col[0].column)].width = 24

        if lignes:
            df_l = pd.DataFrame(lignes)
            df_l["date"] = df_l["date"].apply(lambda d: d.strftime("%d/%m/%Y"))
            df_l.columns = ["Date","Libelle","Type","Montant","Categorie"]
            df_l.to_excel(writer, index=False, sheet_name="Detail Flux")
            ws3 = writer.sheets["Detail Flux"]
            for cell in ws3[1]:
                cell.fill = PatternFill("solid", fgColor="5B2C6F")
                cell.font = Font(color="FFFFFF", bold=True)
            for col in ws3.columns:
                ws3.column_dimensions[get_column_letter(col[0].column)].width = 28

    buf.seek(0)
    return buf

def _lire_flux_fichier(fichier):
    try:
        nom = fichier.name.lower()
        if nom.endswith(".xlsx") or nom.endswith(".xls"):
            df = pd.read_excel(fichier)
        else:
            df = pd.read_csv(fichier, encoding="utf-8", sep=None, engine="python")
        fichier.seek(0)
    except Exception as e:
        return None, str(e)
    cols = {str(c).lower().strip(): c for c in df.columns}
    lignes, erreurs = [], []
    for i, row in df.iterrows():
        try:
            col_date = next((cols[k] for k in ["date","date flux","date operation"] if k in cols), None)
            col_lib  = next((cols[k] for k in ["libelle","libelle","description"] if k in cols), None)
            col_type = next((cols[k] for k in ["type","type (e/d)","sens"] if k in cols), None)
            col_mnt  = next((cols[k] for k in ["montant","amount","valeur"] if k in cols), None)
            col_cat  = next((cols[k] for k in ["categorie","category"] if k in cols), None)
            if not col_date or not col_mnt or not col_type:
                continue
            d = pd.to_datetime(row[col_date], dayfirst=True, errors="coerce")
            if pd.isna(d):
                continue
            t = str(row[col_type]).strip().upper()
            if t not in ("E","D"):
                t = "E" if any(x in t.lower() for x in ["enc","cred","e"]) else "D"
            m = float(str(row[col_mnt]).replace(" ","").replace(",","."))
            lib = str(row[col_lib]) if col_lib else ""
            cat = str(row[col_cat]) if col_cat else "Importe"
            lignes.append({"date": d.date(), "libelle": lib, "type": t,
                           "montant": abs(m), "categorie": cat})
        except Exception as e:
            erreurs.append(f"Ligne {i+2}: {e}")
    return lignes, erreurs or None

def page_tresorerie_previsionnelle():
    st.title("📊 Tresorerie Previsionnelle")
    st.markdown("*Suivi des encaissements et decaissements — Horizon 30 / 60 / 90 jours*")
    st.divider()

    col_p1, col_p2, col_p3, col_p4 = st.columns(4)
    with col_p1:
        date_debut = st.date_input("Date de debut", value=date.today())
    with col_p2:
        horizon_label = st.selectbox("Horizon", list(HORIZON_OPTIONS.keys()), index=0)
        nb_jours = HORIZON_OPTIONS[horizon_label]
    with col_p3:
        devise = st.selectbox("Devise", ["FCFA","XOF","XAF","EUR","USD"], index=0)
    with col_p4:
        solde_initial = st.number_input("Solde initial (tresorerie actuelle)",
                                         value=0.0, step=100_000.0, format="%.0f")

    date_fin = date_debut + timedelta(days=nb_jours - 1)
    st.caption(f"Periode : **{date_debut.strftime('%d/%m/%Y')}** a **{date_fin.strftime('%d/%m/%Y')}**")
    st.divider()

    mode = st.radio("Mode de saisie", ["Saisie manuelle", "Import fichier Excel/CSV"], horizontal=True)
    lignes = []

    if mode == "Import fichier Excel/CSV":
        with st.expander("Format attendu"):
            st.dataframe(pd.DataFrame({
                "Date":["15/05/2026","20/05/2026","31/05/2026"],
                "Libelle":["Reglement SOTRACOM","Salaires mai","TVA decaissee"],
                "Type (E/D)":["E","D","D"],
                "Montant":[4_850_000, 8_500_000, 1_200_000],
                "Categorie":["Reglements clients","Salaires et charges sociales","TVA a decaisser"],
            }), use_container_width=True, hide_index=True)
            st.caption("E = Encaissement | D = Decaissement")

        fichier_flux = st.file_uploader("Importer les flux", type=["xlsx","xls","csv"])
        if fichier_flux:
            lignes_import, err = _lire_flux_fichier(fichier_flux)
            if err:
                for e in err[:5]:
                    st.warning(e)
            if lignes_import:
                lignes = lignes_import
                st.success(f"{len(lignes)} flux importes.")
                ci1, ci2 = st.columns(2)
                ci1.metric("Encaissements importes",
                           f"{sum(l['montant'] for l in lignes if l['type']=='E'):,.0f} {devise}")
                ci2.metric("Decaissements importes",
                           f"{sum(l['montant'] for l in lignes if l['type']=='D'):,.0f} {devise}")
    else:
        st.subheader("Encaissements")
        nb_enc = int(st.number_input("Nombre de lignes", min_value=1, max_value=20,
                                      value=3, step=1, key="nb_enc"))
        cols_h = st.columns([2,3,2,2])
        for lbl, col in zip(["Date","Libelle","Categorie","Montant"], cols_h):
            col.markdown(f"**{lbl}**")
        for i in range(nb_enc):
            c1,c2,c3,c4 = st.columns([2,3,2,2])
            d   = c1.date_input("", value=date_debut+timedelta(days=i*7),
                                  min_value=date_debut, max_value=date_fin,
                                  key=f"ed{i}", label_visibility="collapsed")
            lib = c2.text_input("", key=f"el{i}", label_visibility="collapsed",
                                 placeholder="Ex: Reglement client X")
            cat = c3.selectbox("", CATEGORIES_ENCAISSEMENTS, key=f"ec{i}", label_visibility="collapsed")
            mnt = c4.number_input("", value=0.0, min_value=0.0, step=10_000.0,
                                   format="%.0f", key=f"em{i}", label_visibility="collapsed")
            if mnt > 0:
                lignes.append({"date":d,"libelle":lib or cat,"type":"E","montant":mnt,"categorie":cat})

        st.divider()
        st.subheader("Decaissements")
        nb_dec = int(st.number_input("Nombre de lignes", min_value=1, max_value=20,
                                      value=4, step=1, key="nb_dec"))
        cols_h2 = st.columns([2,3,2,2])
        for lbl, col in zip(["Date","Libelle","Categorie","Montant"], cols_h2):
            col.markdown(f"**{lbl}**")
        for i in range(nb_dec):
            c1,c2,c3,c4 = st.columns([2,3,2,2])
            d   = c1.date_input("", value=date_debut+timedelta(days=i*7+3),
                                  min_value=date_debut, max_value=date_fin,
                                  key=f"dd{i}", label_visibility="collapsed")
            lib = c2.text_input("", key=f"dl{i}", label_visibility="collapsed",
                                 placeholder="Ex: Salaires mai 2026")
            cat = c3.selectbox("", CATEGORIES_DECAISSEMENTS, key=f"dc{i}", label_visibility="collapsed")
            mnt = c4.number_input("", value=0.0, min_value=0.0, step=10_000.0,
                                   format="%.0f", key=f"dm{i}", label_visibility="collapsed")
            if mnt > 0:
                lignes.append({"date":d,"libelle":lib or cat,"type":"D","montant":mnt,"categorie":cat})

    st.divider()
    if st.button("Generer le tableau previsionnel", type="primary", use_container_width=True):
        df_prev = _construire_tableau(date_debut, nb_jours, lignes, solde_initial)

        total_enc  = df_prev["Encaissements"].sum()
        total_dec  = df_prev["Decaissements"].sum()
        solde_fin  = df_prev["Solde"].iloc[-1]
        solde_min  = df_prev["Solde"].min()
        jours_neg  = int((df_prev["Solde"] < 0).sum())

        st.subheader("Synthese")
        k1,k2,k3,k4,k5 = st.columns(5)
        k1.metric("Solde Initial",        f"{solde_initial:,.0f}", devise)
        k2.metric("Total Encaissements",  f"{total_enc:,.0f}",     devise)
        k3.metric("Total Decaissements",  f"{total_dec:,.0f}",     devise)
        k4.metric("Solde Final",          f"{solde_fin:,.0f}",
                  delta=f"{solde_fin - solde_initial:+,.0f}", delta_color="normal")
        k5.metric("Jours en negatif", jours_neg,
                  delta="Risque" if jours_neg > 0 else "OK",
                  delta_color="inverse" if jours_neg > 0 else "off")

        if jours_neg > 0:
            date_neg = df_prev[df_prev["Solde"] < 0]["Date"].iloc[0]
            st.error(
                f"ALERTE — Tresorerie negative pendant {jours_neg} jour(s). "
                f"Premier deficit le {date_neg.strftime('%d/%m/%Y')} "
                f"(min : {solde_min:,.0f} {devise}). Anticipez un financement."
            )
        elif solde_fin < solde_initial * 0.2 and solde_initial > 0:
            st.warning(f"Tresorerie finale faible ({solde_fin:,.0f} {devise}). Surveillez vos recouvrements.")
        else:
            st.success(f"Tresorerie positive sur toute la periode. Solde final : {solde_fin:,.0f} {devise}.")

        st.divider()
        st.plotly_chart(_chart_tresorerie(df_prev, devise), use_container_width=True)
        st.plotly_chart(_chart_flux(df_prev, devise), use_container_width=True)

        st.divider()
        st.subheader("Vue hebdomadaire")
        df_sem = _agreger_semaine(df_prev)
        st.dataframe(df_sem.style.applymap(
            lambda v: "background-color:#f5c6cb;color:#721c24;font-weight:bold"
            if isinstance(v,(int,float)) and v < 0 else "",
            subset=["Solde fin semaine"]
        ), use_container_width=True, hide_index=True)

        with st.expander(f"Detail journalier ({nb_jours} jours)"):
            df_show = df_prev.copy()
            df_show["Date"] = df_show["Date"].apply(lambda d: d.strftime("%d/%m/%Y"))
            st.dataframe(df_show, use_container_width=True, hide_index=True)

        if lignes:
            st.divider()
            st.subheader("Analyse par categorie")
            df_cat = pd.DataFrame(lignes)
            ca1,ca2 = st.columns(2)
            with ca1:
                st.markdown("**Encaissements**")
                enc_cat = df_cat[df_cat["type"]=="E"].groupby("categorie")["montant"].sum().sort_values(ascending=False)
                if not enc_cat.empty:
                    df_e = enc_cat.reset_index()
                    df_e.columns = ["Categorie","Montant"]
                    df_e["Montant"] = df_e["Montant"].apply(lambda x: f"{x:,.0f} {devise}")
                    st.dataframe(df_e, use_container_width=True, hide_index=True)
            with ca2:
                st.markdown("**Decaissements**")
                dec_cat = df_cat[df_cat["type"]=="D"].groupby("categorie")["montant"].sum().sort_values(ascending=False)
                if not dec_cat.empty:
                    df_d = dec_cat.reset_index()
                    df_d.columns = ["Categorie","Montant"]
                    df_d["Montant"] = df_d["Montant"].apply(lambda x: f"{x:,.0f} {devise}")
                    st.dataframe(df_d, use_container_width=True, hide_index=True)

        st.divider()
        excel_buf = _export_excel(df_prev, lignes, solde_initial, devise, date_debut)
        st.download_button(
            label="Telecharger Tresorerie Excel",
            data=excel_buf,
            file_name=f"Tresorerie_{date_debut.strftime('%Y%m%d')}_{horizon_label.replace(' ','')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
