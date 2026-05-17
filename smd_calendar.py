"""
smd_calendar.py - Calendrier des Obligations Fiscales UEMOA
SMD Consulting - Superviseur IA Comptable & Fiscal
Version 1.0

Usage dans app.py :
    from smd_calendar import page_calendrier_fiscal
    elif page == "Calendrier Fiscal":
        page_calendrier_fiscal()
"""

import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
from io import BytesIO
import calendar


# ─────────────────────────────────────────────
#  REFERENTIEL DES OBLIGATIONS PAR PAYS UEMOA
# ─────────────────────────────────────────────

OBLIGATIONS_UEMOA = {
    "Senegal": {
        "admin": "DGID",
        "tva_jour": 15,
        "inps_jour": 15,
        "is_acomptes": [4, 7, 10],   # mois avril, juillet, octobre
        "is_annuel_mois": 4,
        "is_annuel_jour": 30,
        "autres": [
            {"libelle": "CFE / CEL (Contribution Economique Locale)", "mois": 3, "jour": 31},
            {"libelle": "Declaration employeur annuelle (IPRES)", "mois": 1, "jour": 31},
        ]
    },
    "Cote d'Ivoire": {
        "admin": "DGI",
        "tva_jour": 15,
        "inps_jour": 15,
        "is_acomptes": [3, 6, 9, 12],
        "is_annuel_mois": 4,
        "is_annuel_jour": 30,
        "autres": [
            {"libelle": "Contribution des patentes", "mois": 3, "jour": 31},
            {"libelle": "Declaration ITS employeur", "mois": 1, "jour": 31},
        ]
    },
    "Mali": {
        "admin": "DGI",
        "tva_jour": 20,
        "inps_jour": 20,
        "is_acomptes": [3, 6, 9, 12],
        "is_annuel_mois": 4,
        "is_annuel_jour": 30,
        "autres": [
            {"libelle": "Contribution des patentes", "mois": 2, "jour": 28},
        ]
    },
    "Burkina Faso": {
        "admin": "DGI",
        "tva_jour": 20,
        "inps_jour": 20,  # trimestriel
        "inps_trim": True,
        "is_acomptes": [3, 6, 9, 12],
        "is_annuel_mois": 4,
        "is_annuel_jour": 30,
        "autres": [
            {"libelle": "Taxe patronale apprentissage (TPA)", "mois": 3, "jour": 31},
        ]
    },
    "Benin": {
        "admin": "DGI",
        "tva_jour": 20,
        "inps_jour": 20,
        "is_acomptes": [3, 6, 9, 12],
        "is_annuel_mois": 4,
        "is_annuel_jour": 30,
        "autres": [
            {"libelle": "Taxe professionnelle synthetique (TPS)", "mois": 3, "jour": 31},
        ]
    },
    "Togo": {
        "admin": "OTR",
        "tva_jour": 20,
        "inps_jour": 20,
        "is_acomptes": [3, 6, 9, 12],
        "is_annuel_mois": 4,
        "is_annuel_jour": 30,
        "autres": [
            {"libelle": "Taxe professionnelle unique (TPU)", "mois": 3, "jour": 31},
        ]
    },
    "Niger": {
        "admin": "DGI",
        "tva_jour": 15,
        "inps_jour": 20,  # trimestriel
        "inps_trim": True,
        "is_acomptes": [3, 6, 9, 12],
        "is_annuel_mois": 4,
        "is_annuel_jour": 30,
        "autres": [
            {"libelle": "Patente et licence", "mois": 3, "jour": 31},
        ]
    },
    "Guinee-Bissau": {
        "admin": "DGCI",
        "tva_jour": 20,
        "inps_jour": 20,
        "is_acomptes": [3, 6, 9, 12],
        "is_annuel_mois": 3,
        "is_annuel_jour": 31,
        "autres": [
            {"libelle": "Taxe professionnelle (patente)", "mois": 3, "jour": 31},
        ]
    },
}

# Noms complets pour correspondance avec PAYS_UEMOA de smd_engine
NOM_PAYS_MAP = {
    "Senegal":          "Senegal",
    "Cote d'Ivoire":    "Cote d'Ivoire",
    "Mali":             "Mali",
    "Burkina Faso":     "Burkina Faso",
    "Benin":            "Benin",
    "Togo":             "Togo",
    "Niger":            "Niger",
    "Guinee-Bissau":    "Guinee-Bissau",
}

NOMS_MOIS_FR = [
    "", "Janvier", "Fevrier", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Aout", "Septembre", "Octobre", "Novembre", "Decembre"
]


# ─────────────────────────────────────────────
#  CALCUL DES ECHEANCES
# ─────────────────────────────────────────────

def _last_day(annee: int, mois: int) -> int:
    return calendar.monthrange(annee, mois)[1]


def _echeance_tva(annee: int, mois_decl: int, jour: int) -> date:
    """Date limite TVA pour le mois de declaration (mois suivant le mois fiscal)."""
    mois_suivant = mois_decl + 1 if mois_decl < 12 else 1
    annee_suiv   = annee if mois_decl < 12 else annee + 1
    last = _last_day(annee_suiv, mois_suivant)
    return date(annee_suiv, mois_suivant, min(jour, last))


def _generer_echeances(pays: str, annee: int, mois: int) -> list:
    """Genere la liste des echeances pour un pays, une annee et un mois donnes."""
    cfg = OBLIGATIONS_UEMOA.get(pays, {})
    if not cfg:
        return []

    today     = date.today()
    echeances = []

    # ── TVA mensuelle ──────────────────────────────────────────────
    # Echeance du mois en cours = TVA du mois precedent
    mois_fiscal = mois - 1 if mois > 1 else 12
    annee_fisc  = annee if mois > 1 else annee - 1
    ech_tva = _echeance_tva(annee, mois_fiscal, cfg["tva_jour"])
    retard_tva = (today - ech_tva).days if ech_tva < today else 0
    echeances.append({
        "Obligation":    "TVA mensuelle — " + NOMS_MOIS_FR[mois_fiscal] + " " + str(annee_fisc),
        "Type":          "TVA",
        "Echeance":      ech_tva,
        "Jours restants": (ech_tva - today).days,
        "Retard (jours)": max(0, retard_tva),
        "Administration": cfg["admin"],
        "Formulaire":    "Decl. TVA mensuelle",
    })

    # ── INPS / CNSS mensuel ────────────────────────────────────────
    inps_trim = cfg.get("inps_trim", False)
    if not inps_trim:
        jour_inps = cfg.get("inps_jour", 20)
        last_inps = _last_day(annee, mois)
        ech_inps = date(annee, mois, min(jour_inps, last_inps))
        echeances.append({
            "Obligation":    "INPS/CNSS — salaires " + NOMS_MOIS_FR[mois] + " " + str(annee),
            "Type":          "Social",
            "Echeance":      ech_inps,
            "Jours restants": (ech_inps - today).days,
            "Retard (jours)": max(0, (today - ech_inps).days if ech_inps < today else 0),
            "Administration": cfg["admin"],
            "Formulaire":    "Decl. cotisations sociales",
        })
    else:
        # Trimestriel : echeances Q1=mars, Q2=juin, Q3=sept, Q4=dec
        trimestres = {3: "T1", 6: "T2", 9: "T3", 12: "T4"}
        if mois in trimestres:
            jour_inps = cfg.get("inps_jour", 20)
            ech_inps = date(annee, mois, min(jour_inps, _last_day(annee, mois)))
            echeances.append({
                "Obligation":    "INPS/CNSS trimestriel — " + trimestres[mois] + " " + str(annee),
                "Type":          "Social",
                "Echeance":      ech_inps,
                "Jours restants": (ech_inps - today).days,
                "Retard (jours)": max(0, (today - ech_inps).days if ech_inps < today else 0),
                "Administration": cfg["admin"],
                "Formulaire":    "Decl. cotisations sociales trimestrielle",
            })

    # ── Retenues sur salaires (IRPP/IRPS) ─────────────────────────
    jour_ret = cfg.get("tva_jour", 15)
    ech_ret  = _echeance_tva(annee, mois_fiscal, jour_ret)
    echeances.append({
        "Obligation":    "Retenues salaires (IRPP/IRPS) — " + NOMS_MOIS_FR[mois_fiscal] + " " + str(annee_fisc),
        "Type":          "IR Salaires",
        "Echeance":      ech_ret,
        "Jours restants": (ech_ret - today).days,
        "Retard (jours)": max(0, (today - ech_ret).days if ech_ret < today else 0),
        "Administration": cfg["admin"],
        "Formulaire":    "Decl. retenues a la source",
    })

    # ── Acomptes IS (si mois concerne) ────────────────────────────
    if mois in cfg.get("is_acomptes", []):
        ech_acc = date(annee, mois, min(20, _last_day(annee, mois)))
        echeances.append({
            "Obligation":    "Acompte IS provisionnel — " + NOMS_MOIS_FR[mois] + " " + str(annee),
            "Type":          "IS Acompte",
            "Echeance":      ech_acc,
            "Jours restants": (ech_acc - today).days,
            "Retard (jours)": max(0, (today - ech_acc).days if ech_acc < today else 0),
            "Administration": cfg["admin"],
            "Formulaire":    "Acompte IS (1/4 IS N-1)",
        })

    # ── Declaration IS annuelle ────────────────────────────────────
    is_m = cfg.get("is_annuel_mois", 4)
    is_j = cfg.get("is_annuel_jour", 30)
    if mois == is_m:
        ech_is = date(annee, is_m, min(is_j, _last_day(annee, is_m)))
        echeances.append({
            "Obligation":    "Declaration IS annuelle — exercice " + str(annee - 1),
            "Type":          "IS Annuel",
            "Echeance":      ech_is,
            "Jours restants": (ech_is - today).days,
            "Retard (jours)": max(0, (today - ech_is).days if ech_is < today else 0),
            "Administration": cfg["admin"],
            "Formulaire":    "Liasse fiscale annuelle",
        })

    # ── Autres obligations ─────────────────────────────────────────
    for autre in cfg.get("autres", []):
        if autre["mois"] == mois:
            ech_a = date(annee, mois, min(autre["jour"], _last_day(annee, mois)))
            echeances.append({
                "Obligation":    autre["libelle"],
                "Type":          "Autre",
                "Echeance":      ech_a,
                "Jours restants": (ech_a - today).days,
                "Retard (jours)": max(0, (today - ech_a).days if ech_a < today else 0),
                "Administration": cfg["admin"],
                "Formulaire":    autre["libelle"],
            })

    return sorted(echeances, key=lambda x: x["Echeance"])


# ─────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────

def _export_calendrier_excel(df: pd.DataFrame, pays: str, mois: int, annee: int) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendrier Fiscal"

    BLEU   = "1A5276"
    ROUGE  = "C0392B"
    ORANGE = "E67E22"
    VERT   = "1E8449"
    GRIS   = "F2F2F2"
    BLANC  = "FFFFFF"

    ws.merge_cells("A1:G1")
    ws["A1"] = "CALENDRIER DES OBLIGATIONS FISCALES — " + pays.upper()
    ws["A1"].font = Font(bold=True, color=BLANC, size=13)
    ws["A1"].fill = PatternFill("solid", fgColor=BLEU)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = NOMS_MOIS_FR[mois] + " " + str(annee) + " — SMD Consulting"
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Obligation", "Type", "Echeance", "J. Restants", "Retard (j)", "Administration", "Formulaire"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color=BLANC)
        c.fill = PatternFill("solid", fgColor=BLEU)
        c.alignment = Alignment(horizontal="center")

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for i, (_, row) in enumerate(df.iterrows()):
        r = 5 + i
        jr = row["Jours restants"]
        retard = row["Retard (jours)"]

        if retard > 0:
            bg = ROUGE
            fc = BLANC
        elif jr <= 3:
            bg = ROUGE
            fc = BLANC
        elif jr <= 7:
            bg = ORANGE
            fc = BLANC
        else:
            bg = GRIS if i % 2 == 0 else BLANC
            fc = "000000"

        vals = [
            row["Obligation"], row["Type"],
            row["Echeance"].strftime("%d/%m/%Y"),
            int(jr), int(retard),
            row["Administration"], row["Formulaire"]
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(color=fc)
            c.border = thin

    for col, w in zip("ABCDEFG", [45, 14, 13, 12, 12, 16, 35]):
        ws.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
#  PAGE STREAMLIT
# ─────────────────────────────────────────────

def page_calendrier_fiscal():
    st.title("Calendrier des Obligations Fiscales — UEMOA")
    st.markdown("Suivez toutes vos echeances fiscales et sociales par pays. Alertes automatiques J-7 et J-3.")
    st.divider()

    # ── Selecteurs ────────────────────────────────────────────────
    col1, col2, col3 = st.columns(3)
    with col1:
        pays = st.selectbox("Pays UEMOA", list(OBLIGATIONS_UEMOA.keys()))
    with col2:
        today = date.today()
        mois  = st.selectbox("Mois", list(range(1, 13)),
                             index=today.month - 1,
                             format_func=lambda m: NOMS_MOIS_FR[m])
    with col3:
        annee = st.number_input("Annee", value=today.year, min_value=2020, max_value=2035, step=1)

    # ── Calcul ────────────────────────────────────────────────────
    echeances = _generer_echeances(pays, int(annee), mois)
    if not echeances:
        st.warning("Aucune echeance trouvee pour ce pays.")
        return

    df = pd.DataFrame(echeances)

    # ── KPIs ──────────────────────────────────────────────────────
    en_retard  = df[df["Retard (jours)"] > 0]
    urgentes   = df[(df["Jours restants"] >= 0) & (df["Jours restants"] <= 3)]
    bientot    = df[(df["Jours restants"] > 3) & (df["Jours restants"] <= 7)]
    a_venir    = df[df["Jours restants"] > 7]

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Total obligations",  len(df))
    k2.metric("En retard",          len(en_retard),  delta=f"-{len(en_retard)}" if len(en_retard) else None, delta_color="inverse")
    k3.metric("Urgentes (<=3j)",    len(urgentes),   delta=f"-{len(urgentes)}" if len(urgentes) else None, delta_color="inverse")
    k4.metric("Dans 7 jours",       len(bientot))

    st.divider()

    # ── Alertes critiques ─────────────────────────────────────────
    if len(en_retard) > 0:
        st.error("**RETARD FISCAL** — " + str(len(en_retard)) + " obligation(s) depassee(s) ! Penalites en cours.")
        for _, row in en_retard.iterrows():
            st.markdown(
                "**" + row["Obligation"] + "** — Echeance : " +
                row["Echeance"].strftime("%d/%m/%Y") +
                " (" + str(int(row["Retard (jours)"])) + " jours de retard)"
            )

    if len(urgentes) > 0 and len(en_retard) == 0:
        st.warning("**URGENT** — " + str(len(urgentes)) + " echeance(s) dans moins de 3 jours !")

    # ── Tableau complet ───────────────────────────────────────────
    st.subheader("Toutes les echeances — " + NOMS_MOIS_FR[mois] + " " + str(annee))

    df_aff = df.copy()
    df_aff["Echeance"] = df_aff["Echeance"].apply(lambda d: d.strftime("%d/%m/%Y"))
    df_aff["Statut"] = df.apply(
        lambda r: "RETARD" if r["Retard (jours)"] > 0
        else ("URGENT" if r["Jours restants"] <= 3
              else ("BIENTOT" if r["Jours restants"] <= 7 else "OK")),
        axis=1
    )

    cols_aff = ["Statut", "Obligation", "Type", "Echeance", "Jours restants", "Administration", "Formulaire"]
    df_display = df_aff[cols_aff].copy()

    def _couleur_ligne(row):
        s = row["Statut"]
        if s == "RETARD":
            return ["background-color:#f5c6cb"] * len(row)
        elif s == "URGENT":
            return ["background-color:#f5c6cb"] * len(row)
        elif s == "BIENTOT":
            return ["background-color:#ffd699"] * len(row)
        else:
            return ["background-color:#d4edda"] * len(row)

    st.dataframe(
        df_display.style.apply(_couleur_ligne, axis=1),
        use_container_width=True,
        hide_index=True,
        height=400
    )

    # ── Legende ───────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(":red_circle: **RETARD** — Penalites encourues")
    c2.markdown(":orange_circle: **URGENT** — Moins de 3 jours")
    c3.markdown(":yellow_circle: **BIENTOT** — Dans 7 jours")
    c4.markdown(":green_circle: **OK** — Plus de 7 jours")

    st.divider()

    # ── Vue sur 3 mois ────────────────────────────────────────────
    with st.expander("Vue sur 3 mois (mois en cours + 2 suivants)"):
        lignes_3m = []
        for i in range(3):
            m = (mois - 1 + i) % 12 + 1
            a = int(annee) + ((mois - 1 + i) // 12)
            for e in _generer_echeances(pays, a, m):
                e["Mois"] = NOMS_MOIS_FR[m] + " " + str(a)
                lignes_3m.append(e)
        if lignes_3m:
            df_3m = pd.DataFrame(lignes_3m)
            df_3m["Echeance"] = df_3m["Echeance"].apply(lambda d: d.strftime("%d/%m/%Y"))
            st.dataframe(
                df_3m[["Mois", "Obligation", "Type", "Echeance", "Jours restants", "Administration"]],
                use_container_width=True, hide_index=True
            )

    # ── Export ────────────────────────────────────────────────────
    st.subheader("Telecharger le calendrier")
    buf_xl = _export_calendrier_excel(df, pays, mois, int(annee))
    st.download_button(
        "Telecharger Excel",
        data=buf_xl,
        file_name="Calendrier_Fiscal_" + pays + "_" + NOMS_MOIS_FR[mois] + str(annee) + ".xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
