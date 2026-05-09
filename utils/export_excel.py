# -*- coding: utf-8 -*-
"""
Module Export Excel SYSCOHADA - SMD Consulting
Génération d'états financiers professionnels en Excel
Bilan | Compte de Résultat | TAFIRE | Notes | Liasse Fiscale
"""

import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# =============================================================================
# COULEURS SMD CONSULTING
# =============================================================================
BLEU_SMD      = "1F77B4"  # Bleu principal
BLEU_CLAIR    = "D6E8F7"  # Bleu très clair
BLEU_MOYEN    = "9BC8E8"  # Bleu moyen
GRIS_CLAIR    = "F5F5F5"  # Gris très clair
GRIS_MOYEN    = "D9D9D9"  # Gris moyen
VERT_SMD      = "217346"  # Vert Excel
VERT_CLAIR    = "E2EFDA"  # Vert clair
ROUGE_SMD     = "C0392B"  # Rouge alerte
ORANGE_SMD    = "E67E22"  # Orange
BLANC         = "FFFFFF"
NOIR          = "000000"
JAUNE_TITRE   = "FFF2CC"  # Jaune titre


# =============================================================================
# STYLES RÉUTILISABLES
# =============================================================================

def style_titre_principal(ws, row, col, texte, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=texte)
    cell.font = Font(name="Calibri", bold=True, size=14, color=BLANC)
    cell.fill = PatternFill("solid", fgColor=BLEU_SMD)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                      end_row=row, end_column=merge_end_col)
    return cell


def style_sous_titre(ws, row, col, texte, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=texte)
    cell.font = Font(name="Calibri", bold=True, size=11, color=BLANC)
    cell.fill = PatternFill("solid", fgColor=BLEU_MOYEN)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                      end_row=row, end_column=merge_end_col)
    return cell


def style_section(ws, row, col, texte, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=texte)
    cell.font = Font(name="Calibri", bold=True, size=10, color=BLEU_SMD)
    cell.fill = PatternFill("solid", fgColor=BLEU_CLAIR)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                      end_row=row, end_column=merge_end_col)
    return cell


def style_total(ws, row, col, texte, valeur=None, merge_label_end=None):
    cell_label = ws.cell(row=row, column=col, value=texte)
    cell_label.font = Font(name="Calibri", bold=True, size=10, color=BLANC)
    cell_label.fill = PatternFill("solid", fgColor=BLEU_SMD)
    cell_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if merge_label_end:
        ws.merge_cells(start_row=row, start_column=col,
                      end_row=row, end_column=merge_label_end)
    if valeur is not None:
        cv = ws.cell(row=row, column=merge_label_end + 1 if merge_label_end else col + 1,
                    value=valeur)
        cv.font = Font(name="Calibri", bold=True, size=10, color=BLANC)
        cv.fill = PatternFill("solid", fgColor=BLEU_SMD)
        cv.alignment = Alignment(horizontal="right", vertical="center")
        cv.number_format = '#,##0'
    return cell_label


def style_ligne(ws, row, col_label, texte, valeur, col_valeur,
                alternance=False, indent=2):
    bg = GRIS_CLAIR if alternance else BLANC
    cell = ws.cell(row=row, column=col_label, value=texte)
    cell.font = Font(name="Calibri", size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)

    cv = ws.cell(row=row, column=col_valeur, value=valeur)
    cv.font = Font(name="Calibri", size=9)
    cv.fill = PatternFill("solid", fgColor=bg)
    cv.alignment = Alignment(horizontal="right", vertical="center")
    cv.number_format = '#,##0'

    return cell, cv


def ajouter_bordures(ws, min_row, max_row, min_col, max_col):
    thin = Side(style='thin', color=GRIS_MOYEN)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )


def entete_document(ws, titre, nom_entreprise, pays, exercice, devise, nb_cols=6):
    """Crée l'en-tête standard SMD Consulting"""
    # Logo/Titre cabinet
    ws.row_dimensions[1].height = 30
    style_titre_principal(ws, 1, 1, "SMD CONSULTING — SUPERVISEUR IA COMPTABLE SYSCOHADA", nb_cols)

    # Sous-titre document
    ws.row_dimensions[2].height = 22
    style_sous_titre(ws, 2, 1, f"📊 {titre.upper()}", nb_cols)

    # Infos entreprise
    ws.row_dimensions[3].height = 18
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=nb_cols)
    c = ws.cell(row=3, column=1,
                value=f"Entreprise : {nom_entreprise}   |   Pays : {pays}   |   Exercice : {exercice}   |   Devise : {devise}   |   Généré le : {datetime.now().strftime('%d/%m/%Y')}")
    c.font = Font(name="Calibri", size=9, italic=True, color="555555")
    c.fill = PatternFill("solid", fgColor=JAUNE_TITRE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Ligne vide
    ws.row_dimensions[4].height = 8
    return 5  # Retourne la prochaine ligne disponible


def pied_page(ws, row, nb_cols=6):
    """Ajoute le pied de page"""
    ws.row_dimensions[row].height = 15
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
    c = ws.cell(row=row, column=1,
                value=f"© {datetime.now().year} SMD Consulting — Document confidentiel — Normes SYSCOHADA/OHADA — Généré par Superviseur IA Comptable")
    c.font = Font(name="Calibri", size=8, italic=True, color="888888")
    c.alignment = Alignment(horizontal="center")


# =============================================================================
# EXTRACTION DONNÉES DEPUIS LA BALANCE
# =============================================================================

def extraire_montants_balance(df):
    """
    Extrait les montants par classe de comptes depuis la balance.
    Détecte automatiquement les colonnes Compte, Débit, Crédit.
    """
    result = {}

    # Détection colonnes
    col_compte = None
    col_debit = None
    col_credit = None

    for col in df.columns:
        col_lower = str(col).lower()
        if any(x in col_lower for x in ['compte', 'num', 'code']):
            col_compte = col
        elif any(x in col_lower for x in ['debit', 'débit', 'deb']):
            col_debit = col
        elif any(x in col_lower for x in ['credit', 'crédit', 'cred']):
            col_credit = col

    if not col_compte or not col_debit or not col_credit:
        # Essai par position
        cols = df.columns.tolist()
        if len(cols) >= 3:
            col_compte = cols[0]
            col_debit = cols[-2]
            col_credit = cols[-1]
        else:
            return {}

    # Conversion numérique
    df = df.copy()
    df[col_debit] = pd.to_numeric(
        df[col_debit].astype(str).str.replace(',', '.').str.replace(' ', ''), errors='coerce'
    ).fillna(0)
    df[col_credit] = pd.to_numeric(
        df[col_credit].astype(str).str.replace(',', '.').str.replace(' ', ''), errors='coerce'
    ).fillna(0)
    df['_solde'] = df[col_debit] - df[col_credit]
    df['_compte_str'] = df[col_compte].astype(str).str.strip()

    def somme_comptes(prefixes):
        mask = df['_compte_str'].str.startswith(tuple(prefixes))
        return df.loc[mask, '_solde'].sum()

    def somme_comptes_credit(prefixes):
        mask = df['_compte_str'].str.startswith(tuple(prefixes))
        return df.loc[mask, col_credit].sum()

    def somme_comptes_debit(prefixes):
        mask = df['_compte_str'].str.startswith(tuple(prefixes))
        return df.loc[mask, col_debit].sum()

    # ── ACTIF ──────────────────────────────────────────────────────────────
    result['immo_incorp']      = abs(somme_comptes(['21']))
    result['immo_corp']        = abs(somme_comptes(['22', '23', '24']))
    result['immo_fin']         = abs(somme_comptes(['25', '26', '27']))
    result['total_actif_immo'] = result['immo_incorp'] + result['immo_corp'] + result['immo_fin']

    result['stocks']           = abs(somme_comptes(['3']))
    result['creances_clients'] = abs(somme_comptes(['41', '42']))
    result['autres_creances']  = abs(somme_comptes(['43', '44', '45', '46', '47', '48']))
    result['total_actif_circ'] = result['stocks'] + result['creances_clients'] + result['autres_creances']

    result['banques']          = abs(somme_comptes(['51', '52', '53', '54']))
    result['total_trso_actif'] = result['banques']

    result['total_actif']      = (result['total_actif_immo'] +
                                  result['total_actif_circ'] +
                                  result['total_trso_actif'])

    # ── PASSIF ─────────────────────────────────────────────────────────────
    result['capital']          = abs(somme_comptes(['101', '102', '103', '104', '105']))
    result['reserves']         = abs(somme_comptes(['106', '107', '108']))
    result['resultat_net']     = somme_comptes(['109'])
    result['subventions']      = abs(somme_comptes(['12', '13']))
    result['total_cap_propres']= (result['capital'] + result['reserves'] +
                                  result['resultat_net'] + result['subventions'])

    result['emprunts']         = abs(somme_comptes(['14', '15', '16']))
    result['total_dettes_fin'] = result['emprunts']

    result['fournisseurs']     = abs(somme_comptes(['40']))
    result['dettes_fisc_soc']  = abs(somme_comptes(['43', '44']))
    result['autres_dettes']    = abs(somme_comptes(['45', '46', '47', '48']))
    result['total_passif_circ']= (result['fournisseurs'] +
                                  result['dettes_fisc_soc'] +
                                  result['autres_dettes'])

    result['decouvert']        = abs(somme_comptes(['55', '56']))
    result['total_trso_passif']= result['decouvert']

    result['total_passif']     = (result['total_cap_propres'] +
                                  result['total_dettes_fin'] +
                                  result['total_passif_circ'] +
                                  result['total_trso_passif'])

    # ── COMPTE DE RÉSULTAT ─────────────────────────────────────────────────
    result['ventes_march']     = somme_comptes_credit(['701'])
    result['achats_march']     = somme_comptes_debit(['601', '603'])
    result['marge_commerciale']= result['ventes_march'] - result['achats_march']

    result['prod_vendue']      = somme_comptes_credit(['702', '703', '704', '705', '706', '707', '708'])
    result['ca_total']         = result['ventes_march'] + result['prod_vendue']

    result['achats_matieres']  = somme_comptes_debit(['602', '604', '605', '606'])
    result['services_ext']     = somme_comptes_debit(['61', '62'])
    result['impots_taxes']     = somme_comptes_debit(['641', '642', '643', '644'])
    result['valeur_ajoutee']   = (result['ca_total'] -
                                  result['achats_matieres'] -
                                  result['services_ext'] -
                                  result['impots_taxes'])

    result['charges_pers']     = somme_comptes_debit(['66'])
    result['ebe']              = result['valeur_ajoutee'] - result['charges_pers']

    result['dotations_amort']  = somme_comptes_debit(['671', '672', '673'])
    result['resultat_exploit'] = result['ebe'] - result['dotations_amort']

    result['produits_fin']     = somme_comptes_credit(['75', '77'])
    result['charges_fin']      = somme_comptes_debit(['63'])
    result['resultat_fin']     = result['produits_fin'] - result['charges_fin']

    result['rao']              = result['resultat_exploit'] + result['resultat_fin']

    result['produits_hao']     = somme_comptes_credit(['78', '79', '82', '83'])
    result['charges_hao']      = somme_comptes_debit(['68', '69', '80', '81'])
    result['resultat_hao']     = result['produits_hao'] - result['charges_hao']

    result['resultat_avant_is']= result['rao'] + result['resultat_hao']

    # ── RATIOS ─────────────────────────────────────────────────────────────
    result['frng'] = result['total_cap_propres'] + result['total_dettes_fin'] - result['total_actif_immo']
    result['bfr']  = result['total_actif_circ'] - result['total_passif_circ']
    result['tn']   = result['total_trso_actif'] - result['total_trso_passif']

    return result


# =============================================================================
# FEUILLE 1 — BILAN SYSCOHADA
# =============================================================================

def creer_feuille_bilan(wb, data, nom_entreprise, pays, exercice, devise, taux_is):
    ws = wb.create_sheet("📋 Bilan SYSCOHADA")
    ws.sheet_properties.tabColor = BLEU_SMD

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18

    row = entete_document(ws, "BILAN SYSCOHADA", nom_entreprise, pays, exercice, devise, 5)

    # ── EN-TÊTES COLONNES ──────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    for col, texte in [(1, "ACTIF"), (2, f"Montant ({devise})"),
                       (4, "PASSIF"), (5, f"Montant ({devise})")]:
        c = ws.cell(row=row, column=col, value=texte)
        c.font = Font(name="Calibri", bold=True, size=10, color=BLANC)
        c.fill = PatternFill("solid", fgColor=VERT_SMD)
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    start_data_row = row

    # ─────────────────────────────────────────────────────────────────────
    # ACTIF (colonnes A et B)
    # ─────────────────────────────────────────────────────────────────────
    actif_rows = []

    style_section(ws, row, 1, "1. ACTIF IMMOBILISÉ")
    actif_rows.append(row); row += 1

    lignes_actif_immo = [
        ("Immobilisations incorporelles (21x)", data.get('immo_incorp', 0)),
        ("Immobilisations corporelles (22x-24x)", data.get('immo_corp', 0)),
        ("Immobilisations financières (25x-27x)", data.get('immo_fin', 0)),
    ]
    for i, (lbl, val) in enumerate(lignes_actif_immo):
        style_ligne(ws, row, 1, lbl, val, 2, alternance=i % 2 == 0)
        actif_rows.append(row); row += 1

    style_total(ws, row, 1, "TOTAL ACTIF IMMOBILISÉ",
                data.get('total_actif_immo', 0), 1)
    ws.cell(row=row, column=2, value=data.get('total_actif_immo', 0)).number_format = '#,##0'
    actif_rows.append(row); row += 1

    ws.row_dimensions[row].height = 6; row += 1

    style_section(ws, row, 1, "2. ACTIF CIRCULANT")
    actif_rows.append(row); row += 1

    lignes_actif_circ = [
        ("Stocks (3xx)", data.get('stocks', 0)),
        ("Créances clients (41x-42x)", data.get('creances_clients', 0)),
        ("Autres créances (43x-48x)", data.get('autres_creances', 0)),
    ]
    for i, (lbl, val) in enumerate(lignes_actif_circ):
        style_ligne(ws, row, 1, lbl, val, 2, alternance=i % 2 == 0)
        actif_rows.append(row); row += 1

    style_total(ws, row, 1, "TOTAL ACTIF CIRCULANT",
                data.get('total_actif_circ', 0), 1)
    ws.cell(row=row, column=2, value=data.get('total_actif_circ', 0)).number_format = '#,##0'
    actif_rows.append(row); row += 1

    ws.row_dimensions[row].height = 6; row += 1

    style_section(ws, row, 1, "3. TRÉSORERIE ACTIF")
    actif_rows.append(row); row += 1

    style_ligne(ws, row, 1, "Banques et caisses (51x-54x)",
                data.get('banques', 0), 2)
    actif_rows.append(row); row += 1

    style_total(ws, row, 1, "TOTAL TRÉSORERIE ACTIF",
                data.get('total_trso_actif', 0), 1)
    ws.cell(row=row, column=2, value=data.get('total_trso_actif', 0)).number_format = '#,##0'
    actif_rows.append(row); row += 1

    ws.row_dimensions[row].height = 6; row += 1

    # TOTAL ACTIF
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    c1 = ws.cell(row=row, column=1, value="✅ TOTAL ACTIF")
    c1.font = Font(name="Calibri", bold=True, size=12, color=BLANC)
    c1.fill = PatternFill("solid", fgColor=VERT_SMD)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(row=row, column=2, value=data.get('total_actif', 0))
    c2.font = Font(name="Calibri", bold=True, size=12, color=BLANC)
    c2.fill = PatternFill("solid", fgColor=VERT_SMD)
    c2.alignment = Alignment(horizontal="right", vertical="center")
    c2.number_format = '#,##0'
    total_actif_row = row; row += 1

    # ─────────────────────────────────────────────────────────────────────
    # PASSIF (colonnes D et E) — on revient à start_data_row
    # ─────────────────────────────────────────────────────────────────────
    row_p = start_data_row

    style_section(ws, row_p, 4, "1. CAPITAUX PROPRES")
    row_p += 1

    lignes_cap = [
        ("Capital social (101x)", data.get('capital', 0)),
        ("Réserves et report (106x-108x)", data.get('reserves', 0)),
        ("Résultat net (109)", data.get('resultat_net', 0)),
        ("Subventions (12x-13x)", data.get('subventions', 0)),
    ]
    for i, (lbl, val) in enumerate(lignes_cap):
        style_ligne(ws, row_p, 4, lbl, val, 5, alternance=i % 2 == 0)
        row_p += 1

    style_total(ws, row_p, 4, "TOTAL CAPITAUX PROPRES",
                data.get('total_cap_propres', 0), 4)
    ws.cell(row=row_p, column=5, value=data.get('total_cap_propres', 0)).number_format = '#,##0'
    row_p += 1

    ws.row_dimensions[row_p].height = 6; row_p += 1

    style_section(ws, row_p, 4, "2. DETTES FINANCIÈRES")
    row_p += 1

    style_ligne(ws, row_p, 4, "Emprunts et dettes (14x-16x)",
                data.get('emprunts', 0), 5)
    row_p += 1

    style_total(ws, row_p, 4, "TOTAL DETTES FINANCIÈRES",
                data.get('total_dettes_fin', 0), 4)
    ws.cell(row=row_p, column=5, value=data.get('total_dettes_fin', 0)).number_format = '#,##0'
    row_p += 1

    ws.row_dimensions[row_p].height = 6; row_p += 1

    style_section(ws, row_p, 4, "3. PASSIF CIRCULANT")
    row_p += 1

    lignes_pass_circ = [
        ("Fournisseurs (40x)", data.get('fournisseurs', 0)),
        ("Dettes fiscales et sociales (43x-44x)", data.get('dettes_fisc_soc', 0)),
        ("Autres dettes (45x-48x)", data.get('autres_dettes', 0)),
    ]
    for i, (lbl, val) in enumerate(lignes_pass_circ):
        style_ligne(ws, row_p, 4, lbl, val, 5, alternance=i % 2 == 0)
        row_p += 1

    style_total(ws, row_p, 4, "TOTAL PASSIF CIRCULANT",
                data.get('total_passif_circ', 0), 4)
    ws.cell(row=row_p, column=5, value=data.get('total_passif_circ', 0)).number_format = '#,##0'
    row_p += 1

    ws.row_dimensions[row_p].height = 6; row_p += 1

    style_section(ws, row_p, 4, "4. TRÉSORERIE PASSIF")
    row_p += 1

    style_ligne(ws, row_p, 4, "Découverts bancaires (55x-56x)",
                data.get('decouvert', 0), 5)
    row_p += 1

    style_total(ws, row_p, 4, "TOTAL TRÉSORERIE PASSIF",
                data.get('total_trso_passif', 0), 4)
    ws.cell(row=row_p, column=5, value=data.get('total_trso_passif', 0)).number_format = '#,##0'
    row_p += 1

    ws.row_dimensions[row_p].height = 6; row_p += 1

    # TOTAL PASSIF
    ws.row_dimensions[row_p].height = 22
    c3 = ws.cell(row=row_p, column=4, value="✅ TOTAL PASSIF")
    c3.font = Font(name="Calibri", bold=True, size=12, color=BLANC)
    c3.fill = PatternFill("solid", fgColor=VERT_SMD)
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c4 = ws.cell(row=row_p, column=5, value=data.get('total_passif', 0))
    c4.font = Font(name="Calibri", bold=True, size=12, color=BLANC)
    c4.fill = PatternFill("solid", fgColor=VERT_SMD)
    c4.alignment = Alignment(horizontal="right", vertical="center")
    c4.number_format = '#,##0'

    # Équilibre
    row_eq = max(row, row_p) + 2
    ecart = data.get('total_actif', 0) - data.get('total_passif', 0)
    ws.merge_cells(start_row=row_eq, start_column=1, end_row=row_eq, end_column=5)
    c_eq = ws.cell(row=row_eq, column=1,
                   value=f"⚖️ Vérification équilibre : Actif - Passif = {ecart:,.0f} {devise}  {'✅ BILAN ÉQUILIBRÉ' if abs(ecart) < 1 else '⚠️ ÉCART À ANALYSER'}")
    c_eq.font = Font(name="Calibri", bold=True, size=10,
                     color=BLANC if abs(ecart) < 1 else "000000")
    c_eq.fill = PatternFill("solid", fgColor=VERT_SMD if abs(ecart) < 1 else "FFC107")
    c_eq.alignment = Alignment(horizontal="center", vertical="center")

    pied_page(ws, row_eq + 2, 5)


# =============================================================================
# FEUILLE 2 — COMPTE DE RÉSULTAT
# =============================================================================

def creer_feuille_cr(wb, data, nom_entreprise, pays, exercice, devise, taux_is):
    ws = wb.create_sheet("📈 Compte de Résultat")
    ws.sheet_properties.tabColor = "217346"

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    row = entete_document(ws, "COMPTE DE RÉSULTAT SYSCOHADA", nom_entreprise, pays, exercice, devise, 3)

    # En-têtes
    ws.row_dimensions[row].height = 20
    for col, txt in [(1, "LIBELLÉ"), (2, f"Montant ({devise})"), (3, "% CA")]:
        c = ws.cell(row=row, column=col, value=txt)
        c.font = Font(name="Calibri", bold=True, size=10, color=BLANC)
        c.fill = PatternFill("solid", fgColor=VERT_SMD)
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ca = data.get('ca_total', 1) or 1

    def pct(val):
        return round(val / ca * 100, 1) if ca else 0

    def ligne_cr(label, valeur, is_total=False, is_section=False, alt=False):
        nonlocal row
        if is_section:
            style_section(ws, row, 1, label, 2)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        elif is_total:
            bg = BLEU_SMD
            for col in [1, 2, 3]:
                c = ws.cell(row=row, column=col)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Calibri", bold=True, size=10, color=BLANC)
                c.alignment = Alignment(horizontal="right" if col > 1 else "left",
                                       vertical="center", indent=1 if col == 1 else 0)
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = valeur
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3).value = f"{pct(valeur):.1f}%"
        else:
            bg = GRIS_CLAIR if alt else BLANC
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = Font(name="Calibri", size=9)
            c1.fill = PatternFill("solid", fgColor=bg)
            c1.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            c2 = ws.cell(row=row, column=2, value=valeur)
            c2.font = Font(name="Calibri", size=9)
            c2.fill = PatternFill("solid", fgColor=bg)
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c2.number_format = '#,##0'
            c3 = ws.cell(row=row, column=3, value=f"{pct(valeur):.1f}%")
            c3.font = Font(name="Calibri", size=9)
            c3.fill = PatternFill("solid", fgColor=bg)
            c3.alignment = Alignment(horizontal="right", vertical="center")
        row += 1

    # ACTIVITÉ D'EXPLOITATION
    ligne_cr("I. ACTIVITÉ D'EXPLOITATION", None, is_section=True)
    ligne_cr("Ventes de marchandises (701)", data.get('ventes_march', 0), alt=True)
    ligne_cr("Achats de marchandises (601-603)", -data.get('achats_march', 0))
    ligne_cr("MARGE COMMERCIALE", data.get('marge_commerciale', 0), is_total=True)
    ligne_cr("Production vendue (702-708)", data.get('prod_vendue', 0), alt=True)
    ligne_cr("CHIFFRE D'AFFAIRES TOTAL", data.get('ca_total', 0), is_total=True)
    ligne_cr("Achats de matières (602-606)", -data.get('achats_matieres', 0))
    ligne_cr("Services extérieurs (61x-62x)", -data.get('services_ext', 0), alt=True)
    ligne_cr("Impôts et taxes (64x)", -data.get('impots_taxes', 0))
    ligne_cr("VALEUR AJOUTÉE (VA)", data.get('valeur_ajoutee', 0), is_total=True)
    ligne_cr("Charges de personnel (66x)", -data.get('charges_pers', 0), alt=True)
    ligne_cr("EXCÉDENT BRUT D'EXPLOITATION (EBE)", data.get('ebe', 0), is_total=True)
    ligne_cr("Dotations aux amortissements (671-673)", -data.get('dotations_amort', 0))
    ligne_cr("RÉSULTAT D'EXPLOITATION", data.get('resultat_exploit', 0), is_total=True)

    # ACTIVITÉ FINANCIÈRE
    ligne_cr("II. ACTIVITÉ FINANCIÈRE", None, is_section=True)
    ligne_cr("Produits financiers (75x-77x)", data.get('produits_fin', 0), alt=True)
    ligne_cr("Charges financières (63x)", -data.get('charges_fin', 0))
    ligne_cr("RÉSULTAT FINANCIER", data.get('resultat_fin', 0), is_total=True)

    # RAO
    ligne_cr("RÉSULTAT DES ACTIVITÉS ORDINAIRES (RAO)", data.get('rao', 0), is_total=True)

    # HAO
    ligne_cr("III. HORS ACTIVITÉS ORDINAIRES (HAO)", None, is_section=True)
    ligne_cr("Produits HAO (78x-83x)", data.get('produits_hao', 0), alt=True)
    ligne_cr("Charges HAO (68x-81x)", -data.get('charges_hao', 0))
    ligne_cr("RÉSULTAT HAO", data.get('resultat_hao', 0), is_total=True)

    # RÉSULTAT NET
    ligne_cr("IV. RÉSULTAT NET", None, is_section=True)
    resultat_av_is = data.get('resultat_avant_is', 0)
    is_du = max(0, resultat_av_is * taux_is / 100)
    resultat_net = resultat_av_is - is_du

    ligne_cr("Résultat avant IS", resultat_av_is, alt=True)
    ligne_cr(f"Impôt sur les sociétés IS ({taux_is}%)", -is_du)
    ligne_cr("RÉSULTAT NET DE L'EXERCICE", resultat_net, is_total=True)

    # Mise en évidence résultat final
    row_rn = row - 1
    for col in [1, 2, 3]:
        c = ws.cell(row=row_rn, column=col)
        c.fill = PatternFill("solid", fgColor=VERT_SMD if resultat_net >= 0 else ROUGE_SMD)
        c.font = Font(name="Calibri", bold=True, size=12, color=BLANC)

    pied_page(ws, row + 2, 3)


# =============================================================================
# FEUILLE 3 — TAFIRE
# =============================================================================

def creer_feuille_tafire(wb, data, nom_entreprise, pays, exercice, devise, taux_is):
    ws = wb.create_sheet("💰 TAFIRE")
    ws.sheet_properties.tabColor = ORANGE_SMD

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 22

    row = entete_document(ws, "TAFIRE — TABLEAU FINANCIER DES RESSOURCES ET EMPLOIS",
                         nom_entreprise, pays, exercice, devise, 2)

    resultat_avant_is = data.get('resultat_avant_is', 0)
    is_du = max(0, resultat_avant_is * taux_is / 100)
    resultat_net = resultat_avant_is - is_du
    cafg = resultat_net + data.get('dotations_amort', 0)

    def ligne_tafire(label, valeur, is_total=False, is_section=False, alt=False):
        nonlocal row
        ws.row_dimensions[row].height = 18
        if is_section:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            style_sous_titre(ws, row, 1, label)
        elif is_total:
            for col in [1, 2]:
                c = ws.cell(row=row, column=col)
                c.fill = PatternFill("solid", fgColor=BLEU_SMD)
                c.font = Font(name="Calibri", bold=True, size=10, color=BLANC)
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", indent=1)
            ws.cell(row=row, column=2).value = valeur
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        else:
            bg = GRIS_CLAIR if alt else BLANC
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = Font(name="Calibri", size=9)
            c1.fill = PatternFill("solid", fgColor=bg)
            c1.alignment = Alignment(horizontal="left", indent=2)
            c2 = ws.cell(row=row, column=2, value=valeur)
            c2.font = Font(name="Calibri", size=9)
            c2.fill = PatternFill("solid", fgColor=bg)
            c2.alignment = Alignment(horizontal="right")
            c2.number_format = '#,##0'
        row += 1

    # RESSOURCES
    ligne_tafire("I. RESSOURCES", None, is_section=True)
    ligne_tafire("A. Capacité d'Autofinancement Globale (CAFG)", None, is_section=False, alt=True)
    ligne_tafire("  Résultat net de l'exercice", resultat_net, alt=True)
    ligne_tafire("  + Dotations aux amortissements (671-673)", data.get('dotations_amort', 0))
    ligne_tafire("CAFG", cafg, is_total=True)
    ligne_tafire("B. Cessions d'immobilisations (731)", data.get('produits_hao', 0), alt=True)
    ligne_tafire("C. Augmentation de capital (101-105)", 0)
    ligne_tafire("D. Nouveaux emprunts (14x)", 0, alt=True)
    total_ressources = cafg + data.get('produits_hao', 0)
    ligne_tafire("✅ TOTAL RESSOURCES", total_ressources, is_total=True)

    ws.row_dimensions[row].height = 8; row += 1

    # EMPLOIS
    ligne_tafire("II. EMPLOIS", None, is_section=True)
    ligne_tafire("A. Dividendes distribués (452)", 0, alt=True)
    ligne_tafire("B. Acquisitions d'immobilisations (2xx)", data.get('total_actif_immo', 0))
    ligne_tafire("C. Remboursements d'emprunts (14x)", 0, alt=True)
    ligne_tafire("D. Variation du BFR", data.get('bfr', 0))
    total_emplois = data.get('total_actif_immo', 0) + data.get('bfr', 0)
    ligne_tafire("✅ TOTAL EMPLOIS", total_emplois, is_total=True)

    ws.row_dimensions[row].height = 8; row += 1

    # VARIATION TRÉSORERIE
    ligne_tafire("III. VARIATION DE TRÉSORERIE", None, is_section=True)
    var_trso = total_ressources - total_emplois
    ligne_tafire("Total Ressources - Total Emplois", var_trso, alt=True)
    ligne_tafire("Trésorerie Nette (TN = FRNG - BFR)", data.get('tn', 0), is_total=True)

    ws.row_dimensions[row].height = 8; row += 1

    # RATIOS DE SYNTHÈSE
    ligne_tafire("IV. RATIOS DE SYNTHÈSE", None, is_section=True)
    ligne_tafire("Fonds de Roulement Net Global (FRNG)", data.get('frng', 0), alt=True)
    ligne_tafire("Besoin en Fonds de Roulement (BFR)", data.get('bfr', 0))
    ligne_tafire("Trésorerie Nette (TN)", data.get('tn', 0), alt=True)

    pied_page(ws, row + 2, 2)


# =============================================================================
# FEUILLE 4 — RATIOS ET INDICATEURS
# =============================================================================

def creer_feuille_ratios(wb, data, nom_entreprise, pays, exercice, devise, taux_is):
    ws = wb.create_sheet("📊 Ratios & KPIs")
    ws.sheet_properties.tabColor = "7030A0"

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

    row = entete_document(ws, "RATIOS FINANCIERS & INDICATEURS CLÉS",
                         nom_entreprise, pays, exercice, devise, 4)

    ca = data.get('ca_total', 1) or 1
    resultat_avant_is = data.get('resultat_avant_is', 0)
    is_du = max(0, resultat_avant_is * taux_is / 100)
    resultat_net = resultat_avant_is - is_du

    ratios = [
        ("RENTABILITÉ", None, None, None),
        ("Chiffre d'Affaires Total", data.get('ca_total', 0), f"{devise}", "Base de calcul"),
        ("Marge Commerciale", data.get('marge_commerciale', 0),
         f"{data.get('marge_commerciale',0)/ca*100:.1f}% CA", "Marge sur ventes"),
        ("Valeur Ajoutée (VA)", data.get('valeur_ajoutee', 0),
         f"{data.get('valeur_ajoutee',0)/ca*100:.1f}% CA", "Richesse créée"),
        ("EBE", data.get('ebe', 0),
         f"{data.get('ebe',0)/ca*100:.1f}% CA", "Profitabilité opérationnelle"),
        ("Résultat d'Exploitation", data.get('resultat_exploit', 0),
         f"{data.get('resultat_exploit',0)/ca*100:.1f}% CA", "Performance exploitation"),
        ("Résultat Net", resultat_net,
         f"{resultat_net/ca*100:.1f}% CA", "Profitabilité nette"),

        ("STRUCTURE FINANCIÈRE", None, None, None),
        ("FRNG", data.get('frng', 0), f"{devise}", "Équilibre long terme > 0 = sain"),
        ("BFR", data.get('bfr', 0), f"{devise}", "Besoin financement court terme"),
        ("Trésorerie Nette", data.get('tn', 0), f"{devise}", "TN = FRNG - BFR"),
        ("Autonomie Financière",
         f"{data.get('total_cap_propres',0)/max(data.get('total_passif',1),1)*100:.1f}%",
         "", "Cap. Propres / Total Passif > 30%"),
        ("Endettement",
         f"{data.get('emprunts',0)/max(data.get('total_cap_propres',1),1)*100:.1f}%",
         "", "Dettes / Cap. Propres < 100%"),

        ("FISCALITÉ", None, None, None),
        (f"IS Estimé ({taux_is}%)", is_du, f"{devise}", "Impôt sur les sociétés provisionnel"),
        ("TVA collectée estimée", data.get('ca_total', 0) * 0.18,
         f"{devise}", "Base TVA × 18% (estimation)"),
    ]

    alt = False
    for ratio in ratios:
        ws.row_dimensions[row].height = 20
        if ratio[1] is None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            style_sous_titre(ws, row, 1, f"  {ratio[0]}")
        else:
            bg = GRIS_CLAIR if alt else BLANC
            for col, val in [(1, ratio[0]), (2, ratio[1]), (3, ratio[2]), (4, ratio[3])]:
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Calibri", size=9,
                             bold=(col == 1))
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(
                    horizontal="right" if col in [2, 3] else "left",
                    vertical="center",
                    indent=2 if col == 1 else 0
                )
                if col == 2 and isinstance(val, (int, float)):
                    c.number_format = '#,##0'
            alt = not alt
        row += 1

    pied_page(ws, row + 2, 4)


# =============================================================================
# FONCTION PRINCIPALE — EXPORT COMPLET
# =============================================================================

def export_etats_financiers_excel(df_balance, code_pays="SN",
                                   nom_entreprise="", exercice="2025"):
    """
    Génère un fichier Excel complet avec tous les états financiers SYSCOHADA.

    Args:
        df_balance: DataFrame de la balance comptable
        code_pays: Code pays UEMOA (SN, CI, ML, etc.)
        nom_entreprise: Nom de l'entreprise
        exercice: Exercice fiscal

    Returns:
        BytesIO: Fichier Excel prêt au téléchargement
    """
    try:
        from data.plan_comptable_syscohada import get_info_pays
        pays_info = get_info_pays(code_pays)
        pays = pays_info.get('nom', 'UEMOA')
        devise = pays_info.get('devise', 'FCFA (XOF)')
        taux_is = pays_info.get('taux_is', 30)
    except Exception:
        pays = "UEMOA"
        devise = "FCFA (XOF)"
        taux_is = 30

    # Extraction des données depuis la balance
    data = extraire_montants_balance(df_balance)

    # Création du workbook
    wb = Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut

    # Feuille de garde
    ws_garde = wb.create_sheet("🏠 Synthèse")
    ws_garde.sheet_properties.tabColor = BLEU_SMD
    ws_garde.column_dimensions['A'].width = 60
    ws_garde.column_dimensions['B'].width = 25

    ws_garde.row_dimensions[1].height = 40
    ws_garde.merge_cells('A1:B1')
    c = ws_garde.cell(row=1, column=1,
                      value="SMD CONSULTING — SUPERVISEUR IA COMPTABLE SYSCOHADA")
    c.font = Font(name="Calibri", bold=True, size=16, color=BLANC)
    c.fill = PatternFill("solid", fgColor=BLEU_SMD)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws_garde.row_dimensions[2].height = 25
    ws_garde.merge_cells('A2:B2')
    c2 = ws_garde.cell(row=2, column=1, value="ÉTATS FINANCIERS SYSCOHADA / OHADA")
    c2.font = Font(name="Calibri", bold=True, size=13, color=BLANC)
    c2.fill = PatternFill("solid", fgColor=VERT_SMD)
    c2.alignment = Alignment(horizontal="center", vertical="center")

    infos_garde = [
        ("Entreprise", nom_entreprise),
        ("Pays", pays),
        ("Exercice", exercice),
        ("Devise", devise),
        ("Date de génération", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Généré par", "SMD Consulting — Superviseur IA Comptable SYSCOHADA"),
        ("Normes", "SYSCOHADA Révisé / Acte Uniforme OHADA"),
    ]

    for i, (lbl, val) in enumerate(infos_garde):
        r = i + 4
        ws_garde.row_dimensions[r].height = 20
        c_lbl = ws_garde.cell(row=r, column=1, value=lbl)
        c_lbl.font = Font(name="Calibri", bold=True, size=10)
        c_lbl.fill = PatternFill("solid", fgColor=BLEU_CLAIR if i % 2 == 0 else BLANC)
        c_lbl.alignment = Alignment(horizontal="left", indent=2)
        c_val = ws_garde.cell(row=r, column=2, value=val)
        c_val.font = Font(name="Calibri", size=10)
        c_val.fill = PatternFill("solid", fgColor=BLEU_CLAIR if i % 2 == 0 else BLANC)
        c_val.alignment = Alignment(horizontal="left", indent=1)

    # Sommaire onglets
    r_som = len(infos_garde) + 6
    ws_garde.row_dimensions[r_som].height = 20
    ws_garde.merge_cells(start_row=r_som, start_column=1, end_row=r_som, end_column=2)
    c_som = ws_garde.cell(row=r_som, column=1, value="📑 CONTENU DU FICHIER")
    c_som.font = Font(name="Calibri", bold=True, size=11, color=BLANC)
    c_som.fill = PatternFill("solid", fgColor=BLEU_SMD)
    c_som.alignment = Alignment(horizontal="center")

    onglets = [
        "📋 Bilan SYSCOHADA",
        "📈 Compte de Résultat",
        "💰 TAFIRE",
        "📊 Ratios & KPIs",
    ]
    for i, onglet in enumerate(onglets):
        r = r_som + i + 1
        ws_garde.row_dimensions[r].height = 18
        ws_garde.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c_ong = ws_garde.cell(row=r, column=1, value=f"  {i+1}. {onglet}")
        c_ong.font = Font(name="Calibri", size=10)
        c_ong.fill = PatternFill("solid", fgColor=GRIS_CLAIR if i % 2 == 0 else BLANC)
        c_ong.alignment = Alignment(horizontal="left", indent=2)

    # Création des feuilles
    creer_feuille_bilan(wb, data, nom_entreprise, pays, exercice, devise, taux_is)
    creer_feuille_cr(wb, data, nom_entreprise, pays, exercice, devise, taux_is)
    creer_feuille_tafire(wb, data, nom_entreprise, pays, exercice, devise, taux_is)
    creer_feuille_ratios(wb, data, nom_entreprise, pays, exercice, devise, taux_is)

    # Sauvegarde
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer