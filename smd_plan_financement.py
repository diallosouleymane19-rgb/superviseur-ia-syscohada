# -*- coding: utf-8 -*-
"""
Module Plan de Financement SYSCOHADA — SMD Consulting
Saisie manuelle + Import balance | Analyse IA Mistral | Export Excel/Word
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import datetime

# ─── Catégories standard SYSCOHADA ──────────────────────────────────────────

RESSOURCES = [
    "Capacité d'autofinancement (CAF)",
    "Cessions d'immobilisations",
    "Augmentation de capital",
    "Subventions d'investissement reçues",
    "Emprunts à long terme",
    "Emprunts à moyen terme",
    "Autres ressources stables",
]

EMPLOIS = [
    "Acquisitions d'immobilisations incorporelles",
    "Acquisitions d'immobilisations corporelles",
    "Acquisitions d'immobilisations financières",
    "Remboursements d'emprunts LT",
    "Remboursements d'emprunts MT",
    "Distribution de dividendes",
    "Variation du besoin en fonds de roulement (BFR)",
    "Autres emplois stables",
]


# ─── Helpers ────────────────────────────────────────────────────────────────

def _chart_plan(df_r, df_e, annees, devise):
    total_r = [df_r[a].sum() for a in annees]
    total_e = [df_e[a].sum() for a in annees]
    solde   = [r - e for r, e in zip(total_r, total_e)]

    fig = go.Figure()
    fig.add_trace(go.Bar(name="Ressources", x=annees, y=total_r,
                         marker_color="#1E8449",
                         hovertemplate="%{y:,.0f} " + devise + "<extra>Ressources</extra>"))
    fig.add_trace(go.Bar(name="Emplois", x=annees, y=total_e,
                         marker_color="#C0392B",
                         hovertemplate="%{y:,.0f} " + devise + "<extra>Emplois</extra>"))
    fig.add_trace(go.Scatter(name="Solde", x=annees, y=solde,
                             mode="lines+markers",
                             line=dict(color="#2C3E50", width=2.5, dash="dot"),
                             marker=dict(size=9),
                             hovertemplate="%{y:,.0f} " + devise + "<extra>Solde</extra>"))
    fig.add_hline(y=0, line_dash="dash", line_color="#888", line_width=1)
    fig.update_layout(
        title="Plan de financement — Ressources vs Emplois",
        barmode="group", height=420,
        margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", y=-0.15),
        yaxis_tickformat=",.0f",
    )
    return fig


def _export_excel_plan(df_r, df_e, annees, entreprise, devise):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        def style_ws(ws, color_header):
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=color_header)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 28

        # Feuille Ressources
        df_r_exp = df_r.copy()
        df_r_exp.loc["TOTAL RESSOURCES"] = df_r_exp[annees].sum()
        df_r_exp.to_excel(writer, sheet_name="Ressources")
        style_ws(writer.sheets["Ressources"], "1E8449")

        # Feuille Emplois
        df_e_exp = df_e.copy()
        df_e_exp.loc["TOTAL EMPLOIS"] = df_e_exp[annees].sum()
        df_e_exp.to_excel(writer, sheet_name="Emplois")
        style_ws(writer.sheets["Emplois"], "C0392B")

        # Feuille Synthèse
        synth_data = {
            "Indicateur": ["Total Ressources", "Total Emplois", "Solde net", "Solde cumulé"]
        }
        tot_r = df_r[annees].sum()
        tot_e = df_e[annees].sum()
        solde = tot_r - tot_e
        cumul = solde.cumsum()
        for a in annees:
            synth_data[a] = [tot_r[a], tot_e[a], solde[a], cumul[a]]
        df_s = pd.DataFrame(synth_data).set_index("Indicateur")
        df_s.to_excel(writer, sheet_name="Synthèse")
        style_ws(writer.sheets["Synthèse"], "1F4E79")

    buf.seek(0)
    return buf


def _analyser_plan_ia(df_r, df_e, annees, devise, entreprise):
    from utils.ai import appel_mistral
    tot_r  = df_r[annees].sum()
    tot_e  = df_e[annees].sum()
    solde  = tot_r - tot_e
    cumul  = solde.cumsum()

    lignes = []
    for a in annees:
        lignes.append(
            f"  {a} : Ressources={tot_r[a]:,.0f} | Emplois={tot_e[a]:,.0f} | "
            f"Solde={solde[a]:,.0f} | Solde cumulé={cumul[a]:,.0f} {devise}"
        )

    prompt = f"""Tu es un expert-comptable SYSCOHADA analysant un plan de financement.

Entreprise : {entreprise or 'Non précisée'}
Période : {annees[0]} à {annees[-1]}
Devise : {devise}

Données du plan :
{chr(10).join(lignes)}

Ressources principales :
{df_r[annees].sum(axis=1).sort_values(ascending=False).head(5).to_string()}

Emplois principaux :
{df_e[annees].sum(axis=1).sort_values(ascending=False).head(5).to_string()}

Analyse ce plan de financement selon les normes SYSCOHADA et réponds en structurant ta réponse ainsi :

## 1. Équilibre financier
Analyse de l'équilibre ressources/emplois sur la période.

## 2. Points forts
Les éléments positifs du plan.

## 3. Points de vigilance
Les risques et déséquilibres détectés.

## 4. Recommandations SYSCOHADA
Actions concrètes pour améliorer le plan (refinancement, réduction BFR, etc.).

## 5. Conclusion
Appréciation globale de la viabilité financière.
"""
    return appel_mistral(prompt)


def _lire_balance_pour_plan(fichier):
    """Extrait CAF et BFR depuis une balance SYSCOHADA."""
    try:
        if fichier.name.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(fichier)
        else:
            df = pd.read_csv(fichier, encoding="utf-8", sep=None, engine="python")
        fichier.seek(0)

        cols = {str(c).lower().strip(): c for c in df.columns}
        col_num  = next((cols[k] for k in ["compte","numero","num","n°"] if k in cols), None)
        col_sold = next((cols[k] for k in ["solde","solde net","montant","balance"] if k in cols), None)

        if not col_num or not col_sold:
            return {}

        df[col_num]  = df[col_num].astype(str).str.strip()
        df[col_sold] = pd.to_numeric(df[col_sold].astype(str).str.replace(" ","").str.replace(",","."), errors="coerce").fillna(0)

        def somme(prefixes):
            mask = df[col_num].str.startswith(tuple(prefixes))
            return float(df.loc[mask, col_sold].sum())

        # Résultat net (cl. 12/13)
        resultat = somme(["12","13"])
        # Dotations amortissements (cl. 681)
        dotations = somme(["681","6811","6812"])
        caf = resultat + dotations

        # BFR = Stocks + Créances - Dettes CT
        stocks    = somme(["3"])
        creances  = somme(["41","409"])
        dettes_ct = somme(["40","401","42","43","44"])
        bfr       = stocks + creances - dettes_ct

        return {"CAF estimée": max(caf, 0), "Variation BFR estimée": abs(bfr)}
    except Exception:
        return {}


# ─── Page principale ─────────────────────────────────────────────────────────

def page_plan_financement():
    st.title("📐 Plan de Financement SYSCOHADA")
    st.markdown("*Équilibre ressources / emplois — Horizon 1 à 5 ans*")
    st.divider()

    # ── Paramètres généraux ──────────────────────────────────────────────
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        annee_base = st.number_input("Année de départ", value=datetime.now().year,
                                     min_value=2000, max_value=2050, step=1)
    with col2:
        nb_annees = st.selectbox("Horizon", [1, 2, 3, 4, 5], index=2)
    with col3:
        devise = st.selectbox("Devise", ["FCFA","XOF","XAF","EUR","USD"])
    with col4:
        entreprise = st.text_input("Entreprise", placeholder="Nom de l'entreprise")

    annees = [str(int(annee_base) + i) for i in range(nb_annees)]
    st.caption(f"Période : **{annees[0]}** → **{annees[-1]}**")
    st.divider()

    # ── Mode de saisie ───────────────────────────────────────────────────
    mode = st.radio("Mode de saisie", ["✏️ Saisie manuelle", "📂 Import balance SYSCOHADA"],
                    horizontal=True)

    prefill_r = {}
    prefill_e = {}

    if mode == "📂 Import balance SYSCOHADA":
        with st.expander("ℹ️ Format attendu"):
            st.caption("Fichier CSV/XLSX avec colonnes : Compte | Solde. "
                       "La CAF et le BFR seront estimés automatiquement.")
        fichier_bal = st.file_uploader("Importer la balance", type=["xlsx","xls","csv"],
                                        key="bal_plan")
        if fichier_bal:
            vals = _lire_balance_pour_plan(fichier_bal)
            if vals:
                prefill_r["Capacité d'autofinancement (CAF)"] = vals.get("CAF estimée", 0)
                prefill_e["Variation du besoin en fonds de roulement (BFR)"] = vals.get("Variation BFR estimée", 0)
                st.success(f"✅ CAF estimée : {prefill_r.get(\"Capacité d'autofinancement (CAF)\", 0):,.0f} {devise} | "
                           f"BFR estimé : {prefill_e.get('Variation du besoin en fonds de roulement (BFR)', 0):,.0f} {devise}")
            else:
                st.warning("⚠️ Impossible d'extraire automatiquement — vérifiez les colonnes (Compte, Solde).")

    st.divider()

    # ── Tableau de saisie Ressources ─────────────────────────────────────
    st.subheader("📥 Ressources")
    r_data = {"Ressource": RESSOURCES}
    for a in annees:
        r_data[a] = [prefill_r.get(lib, 0.0) for lib in RESSOURCES]
    df_r = pd.DataFrame(r_data).set_index("Ressource")
    df_r_edit = st.data_editor(
        df_r, use_container_width=True, num_rows="fixed",
        column_config={a: st.column_config.NumberColumn(a, format="%.0f", min_value=0)
                       for a in annees},
        key="edit_ressources"
    )

    # Ligne autres ressources libre
    st.caption("💡 Modifiez directement les cellules du tableau.")

    st.divider()

    # ── Tableau de saisie Emplois ─────────────────────────────────────────
    st.subheader("📤 Emplois")
    e_data = {"Emploi": EMPLOIS}
    for a in annees:
        e_data[a] = [prefill_e.get(lib, 0.0) for lib in EMPLOIS]
    df_e = pd.DataFrame(e_data).set_index("Emploi")
    df_e_edit = st.data_editor(
        df_e, use_container_width=True, num_rows="fixed",
        column_config={a: st.column_config.NumberColumn(a, format="%.0f", min_value=0)
                       for a in annees},
        key="edit_emplois"
    )

    st.divider()

    if st.button("📐 Générer le Plan de Financement", type="primary", use_container_width=True):
        # ── Calculs ───────────────────────────────────────────────────────
        tot_r   = df_r_edit[annees].sum()
        tot_e   = df_e_edit[annees].sum()
        solde   = tot_r - tot_e
        cumul   = solde.cumsum()

        # ── KPIs ──────────────────────────────────────────────────────────
        st.subheader("📊 Synthèse")
        cols = st.columns(len(annees) + 1)
        cols[0].metric("Période", f"{annees[0]}–{annees[-1]}")
        for i, a in enumerate(annees):
            delta_color = "normal" if solde[a] >= 0 else "inverse"
            cols[i+1].metric(
                a,
                f"{solde[a]:+,.0f}",
                f"Cumulé : {cumul[a]:+,.0f}",
                delta_color=delta_color
            )

        # Alertes
        annees_neg = [a for a in annees if solde[a] < 0]
        if annees_neg:
            st.error(f"⚠️ Solde négatif détecté en : **{', '.join(annees_neg)}** — plan à rééquilibrer.")
        elif cumul[annees[-1]] < 0:
            st.warning("⚠️ Solde cumulé négatif en fin de période — risque de déséquilibre financier.")
        else:
            st.success("✅ Plan équilibré sur toute la période.")

        st.divider()

        # ── Tableau récapitulatif ─────────────────────────────────────────
        synth = pd.DataFrame({
            "": ["Total Ressources", "Total Emplois", "Solde net", "Solde cumulé"],
        })
        for a in annees:
            synth[a] = [f"{tot_r[a]:,.0f}", f"{tot_e[a]:,.0f}",
                        f"{solde[a]:+,.0f}", f"{cumul[a]:+,.0f}"]
        st.dataframe(synth.set_index(""), use_container_width=True)

        # ── Graphique ─────────────────────────────────────────────────────
        st.plotly_chart(_chart_plan(df_r_edit, df_e_edit, annees, devise),
                        use_container_width=True)

        # ── Détail ressources / emplois ───────────────────────────────────
        col_r, col_e = st.columns(2)
        with col_r:
            st.markdown("**Détail Ressources**")
            df_r_show = df_r_edit[annees].copy()
            df_r_show.loc["TOTAL"] = df_r_show.sum()
            st.dataframe(df_r_show.style.format("{:,.0f}"), use_container_width=True)
        with col_e:
            st.markdown("**Détail Emplois**")
            df_e_show = df_e_edit[annees].copy()
            df_e_show.loc["TOTAL"] = df_e_show.sum()
            st.dataframe(df_e_show.style.format("{:,.0f}"), use_container_width=True)

        # ── Analyse IA ────────────────────────────────────────────────────
        st.divider()
        if st.button("🤖 Analyse IA du plan", use_container_width=True):
            with st.spinner("Analyse IA SYSCOHADA en cours..."):
                analyse = _analyser_plan_ia(df_r_edit, df_e_edit, annees, devise, entreprise)
                st.session_state["plan_analyse_ia"] = analyse

        if st.session_state.get("plan_analyse_ia"):
            st.subheader("🤖 Analyse IA SYSCOHADA")
            st.markdown(st.session_state["plan_analyse_ia"])

        # ── Exports ───────────────────────────────────────────────────────
        st.divider()
        col_x1, col_x2 = st.columns(2)
        with col_x1:
            excel_buf = _export_excel_plan(df_r_edit, df_e_edit, annees, entreprise, devise)
            st.download_button(
                "📊 Télécharger Excel",
                excel_buf,
                f"Plan_Financement_{entreprise or 'SMD'}_{annees[0]}_{annees[-1]}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_x2:
            # Export HTML simple
            html = f"""<html><head><meta charset='UTF-8'>
            <title>Plan de Financement {entreprise}</title>
            <style>body{{font-family:Arial;margin:40px}}
            h1{{color:#1E8449}} table{{border-collapse:collapse;width:100%}}
            th{{background:#1E8449;color:white;padding:8px}} td{{border:1px solid #ddd;padding:6px}}
            </style></head><body>
            <h1>Plan de Financement SYSCOHADA</h1>
            <p><b>Entreprise :</b> {entreprise or 'N/A'} | <b>Période :</b> {annees[0]}–{annees[-1]} | <b>Devise :</b> {devise}</p>
            <h2>Ressources</h2>{df_r_edit.style.format('{{:,.0f}}').to_html()}
            <h2>Emplois</h2>{df_e_edit.style.format('{{:,.0f}}').to_html()}
            </body></html>"""
            import base64
            b64 = base64.b64encode(html.encode()).decode()
            st.markdown(
                f'<a href="data:text/html;base64,{b64}" download="Plan_Financement.html">'
                f'📄 Télécharger HTML</a>', unsafe_allow_html=True
            )
