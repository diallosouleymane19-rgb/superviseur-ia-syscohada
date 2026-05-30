# -*- coding: utf-8 -*-
"""
Module Aide TVA CA3 / CA12
Calcul, vérification et aide à la déclaration TVA France
SMD Consulting - DGFiP / PCG France
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
from datetime import datetime


# ─────────────────────────────────────────────
# CONSTANTES TVA FRANCE
# ─────────────────────────────────────────────

TAUX_TVA = {
    "20% — Taux normal": 0.20,
    "10% — Taux intermédiaire": 0.10,
    "5,5% — Taux réduit": 0.055,
    "2,1% — Taux super-réduit": 0.021,
    "0% — Exonéré / Export": 0.0,
}

RUBRIQUES_CA3 = {
    "Ventes soumises à 20%": {"taux": 0.20, "ligne_ca3": "A1", "compte_pcg": "70x"},
    "Ventes soumises à 10%": {"taux": 0.10, "ligne_ca3": "A2", "compte_pcg": "70x"},
    "Ventes soumises à 5,5%": {"taux": 0.055, "ligne_ca3": "A3", "compte_pcg": "70x"},
    "Ventes soumises à 2,1%": {"taux": 0.021, "ligne_ca3": "A4", "compte_pcg": "70x"},
    "Acquisitions intracommunautaires (20%)": {"taux": 0.20, "ligne_ca3": "B1", "compte_pcg": "401/445"},
    "Autoliquidation achats (20%)": {"taux": 0.20, "ligne_ca3": "B2", "compte_pcg": "401/445"},
    "Ventes exonérées / exports": {"taux": 0.0, "ligne_ca3": "E1", "compte_pcg": "70x"},
}

COMPTES_TVA_PCG = {
    "44566 — TVA déductible sur ABS": "TVA déductible sur biens et services",
    "44562 — TVA déductible sur immos": "TVA déductible sur immobilisations",
    "44571 — TVA collectée": "TVA collectée",
    "44551 — TVA à décaisser": "Solde TVA à payer",
    "44567 — Crédit de TVA": "Crédit de TVA reportable",
    "44563 — TVA intracommunautaire": "TVA intracommunautaire déductible",
}


# ─────────────────────────────────────────────
# CALCULS
# ─────────────────────────────────────────────

def _calculer_tva(data: dict) -> dict:
    """Calcule TVA collectée, déductible, solde à partir des données saisies."""
    tva_collectee_detail = {}
    tva_collectee_total = 0.0

    for rubrique, info in RUBRIQUES_CA3.items():
        ht = data.get(rubrique, 0.0) or 0.0
        tva = ht * info["taux"]
        if ht != 0:
            tva_collectee_detail[rubrique] = {"base_ht": ht, "taux": info["taux"], "tva": tva}
        tva_collectee_total += tva

    # TVA déductible saisie directement
    tva_ded_abs = data.get("TVA déductible sur achats (ABS)", 0.0) or 0.0
    tva_ded_immo = data.get("TVA déductible sur immobilisations", 0.0) or 0.0
    tva_ded_intra = data.get("TVA intracommunautaire déductible", 0.0) or 0.0
    tva_deductible_total = tva_ded_abs + tva_ded_immo + tva_ded_intra

    credit_reporte = data.get("Crédit de TVA période précédente", 0.0) or 0.0

    solde = tva_collectee_total - tva_deductible_total - credit_reporte

    return {
        "tva_collectee_detail": tva_collectee_detail,
        "tva_collectee": tva_collectee_total,
        "tva_ded_abs": tva_ded_abs,
        "tva_ded_immo": tva_ded_immo,
        "tva_ded_intra": tva_ded_intra,
        "tva_deductible": tva_deductible_total,
        "credit_reporte": credit_reporte,
        "solde": solde,
        "a_payer": max(solde, 0),
        "credit_genere": max(-solde, 0),
    }


def _extraire_tva_depuis_balance(df: pd.DataFrame) -> dict:
    """Extrait les montants TVA depuis une balance PCG."""
    df = df.copy()

    def _col(df, names):
        for n in names:
            if n in df.columns:
                return n
        return None

    col_num = _col(df, ['CompteNum', 'Compte', 'compte', 'NumCompte'])
    col_deb = _col(df, ['Debit', 'debit', 'SoldeDebit', 'Mouvement_Debit'])
    col_cred = _col(df, ['Credit', 'credit', 'SoldeCredit', 'Mouvement_Credit'])

    if not col_num:
        return {}

    df['_num'] = df[col_num].astype(str).str.strip()
    df['_deb'] = pd.to_numeric(df[col_deb].astype(str).str.replace(',', '.').str.replace(' ', ''), errors='coerce').fillna(0) if col_deb else 0
    df['_cred'] = pd.to_numeric(df[col_cred].astype(str).str.replace(',', '.').str.replace(' ', ''), errors='coerce').fillna(0) if col_cred else 0
    df['_solde'] = df['_deb'] - df['_cred']

    def _somme(prefixes):
        mask = df['_num'].str.startswith(tuple(prefixes), na=False)
        return abs(df.loc[mask, '_solde'].sum())

    result = {
        "TVA collectée (44571)": _somme(['44571']),
        "TVA déductible ABS (44566)": _somme(['44566']),
        "TVA déductible immos (44562)": _somme(['44562']),
        "TVA intracommunautaire (44563)": _somme(['44563']),
        "Crédit TVA (44567)": _somme(['44567']),
        "TVA à décaisser (44551)": _somme(['44551']),
        "Chiffre d'affaires HT (70x)": _somme(['70']),
        "Achats HT (60x)": _somme(['60', '61', '62']),
    }
    return result


def _verifier_coherence(res: dict) -> list:
    """Retourne des alertes de cohérence."""
    alertes = []
    col = res.get("tva_collectee", 0)
    ded = res.get("tva_deductible", 0)
    solde = res.get("solde", 0)

    if col == 0 and ded > 0:
        alertes.append(("warning", "TVA collectée nulle mais déductible > 0 — vérifiez les comptes 44571"))
    if ded > col * 1.5 and col > 0:
        alertes.append(("warning", f"TVA déductible ({ded:,.0f} €) très supérieure à la TVA collectée ({col:,.0f} €) — crédit structurel, vérifiez"))
    if solde > 50000:
        alertes.append(("info", f"TVA à décaisser importante ({solde:,.0f} €) — pensez à la provision en comptabilité"))
    if res.get("credit_genere", 0) > 0:
        alertes.append(("success", f"Crédit de TVA de {res['credit_genere']:,.0f} € — remboursement possible si > 760 € (régime réel normal)"))
    return alertes


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────

def _export_excel_tva(res: dict, data: dict, periode: str, entreprise: str) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Feuille Déclaration
        rows_decl = []
        for rubrique, detail in res["tva_collectee_detail"].items():
            rows_decl.append({
                "Rubrique": rubrique,
                "Base HT (€)": detail["base_ht"],
                "Taux": f"{detail['taux']*100:.1f}%",
                "TVA collectée (€)": detail["tva"],
            })
        rows_decl.append({"Rubrique": "TOTAL TVA COLLECTÉE", "Base HT (€)": "", "Taux": "", "TVA collectée (€)": res["tva_collectee"]})
        rows_decl.append({"Rubrique": "TVA déductible ABS", "Base HT (€)": "", "Taux": "", "TVA collectée (€)": -res["tva_ded_abs"]})
        rows_decl.append({"Rubrique": "TVA déductible immos", "Base HT (€)": "", "Taux": "", "TVA collectée (€)": -res["tva_ded_immo"]})
        rows_decl.append({"Rubrique": "Crédit TVA reporté", "Base HT (€)": "", "Taux": "", "TVA collectée (€)": -res["credit_reporte"]})
        rows_decl.append({"Rubrique": "SOLDE TVA", "Base HT (€)": "", "Taux": "", "TVA collectée (€)": res["solde"]})
        pd.DataFrame(rows_decl).to_excel(writer, sheet_name="Déclaration TVA", index=False)
        ws = writer.sheets["Déclaration TVA"]
        from openpyxl.styles import PatternFill, Font, Alignment
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(color="FFFFFF", bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 35

        # Feuille Comptes PCG
        rows_pcg = [{"Compte PCG": k, "Nature": v} for k, v in COMPTES_TVA_PCG.items()]
        pd.DataFrame(rows_pcg).to_excel(writer, sheet_name="Comptes PCG", index=False)

        # Feuille Info
        pd.DataFrame([
            {"Champ": "Entreprise", "Valeur": entreprise},
            {"Champ": "Période", "Valeur": periode},
            {"Champ": "Généré le", "Valeur": datetime.now().strftime("%d/%m/%Y %H:%M")},
        ]).to_excel(writer, sheet_name="Info", index=False)

    return buf.getvalue()


# ─────────────────────────────────────────────
# GRAPHIQUE
# ─────────────────────────────────────────────

def _chart_tva(res: dict) -> go.Figure:
    labels = ["TVA Collectée", "TVA Déductible"]
    values = [res["tva_collectee"], res["tva_deductible"]]
    colors = ["#e74c3c", "#27ae60"]

    fig = go.Figure(go.Bar(
        x=labels,
        y=values,
        marker_color=colors,
        text=[f"{v:,.0f} €" for v in values],
        textposition="outside",
    ))
    solde = res["solde"]
    fig.add_hline(
        y=0, line_dash="dash", line_color="grey",
        annotation_text=f"Solde : {solde:+,.0f} €",
        annotation_position="bottom right"
    )
    fig.update_layout(
        title="TVA Collectée vs Déductible",
        yaxis_title="Montant (€)",
        height=350
    )
    return fig


# ─────────────────────────────────────────────
# PAGE PRINCIPALE
# ─────────────────────────────────────────────

def page_tva():
    st.title("🧾 Aide Déclaration TVA — CA3 / CA12")
    st.markdown("**Calcul et vérification TVA France** — Régimes réel normal (CA3) et simplifié (CA12)")
    st.caption("✨ Pour Cabinets, DAF et Dirigeants - DGFiP / PCG France")

    # ── Paramètres ──
    col1, col2, col3 = st.columns(3)
    with col1:
        entreprise = st.text_input("🏢 Entreprise", value="Entreprise")
    with col2:
        regime = st.selectbox("📋 Régime TVA", ["CA3 — Réel normal (mensuel/trimestriel)", "CA12 — Réel simplifié (annuel)"])
    with col3:
        periode = st.text_input("📅 Période", value=f"{datetime.now().strftime('%m/%Y')}")

    st.divider()

    # ── Tabs ──
    tab1, tab2, tab3 = st.tabs(["📝 Saisie manuelle", "📁 Import balance", "📖 Guide taux & comptes"])

    # ══════════════════════════════════════════
    # TAB 1 — Saisie manuelle
    # ══════════════════════════════════════════
    with tab1:
        st.markdown("### 💰 Bases HT soumises à TVA (ventes / opérations imposables)")

        data = {}
        col_a, col_b = st.columns(2)

        with col_a:
            st.markdown("**Ventes / Opérations taxables**")
            data["Ventes soumises à 20%"] = st.number_input("Ventes 20% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Ventes soumises à 10%"] = st.number_input("Ventes 10% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Ventes soumises à 5,5%"] = st.number_input("Ventes 5,5% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Ventes soumises à 2,1%"] = st.number_input("Ventes 2,1% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Ventes exonérées / exports"] = st.number_input("Ventes exonérées / exports (€ HT)", min_value=0.0, step=100.0, format="%.2f")

        with col_b:
            st.markdown("**Opérations particulières**")
            data["Acquisitions intracommunautaires (20%)"] = st.number_input("Acquisitions intracom. 20% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Autoliquidation achats (20%)"] = st.number_input("Autoliquidation achats 20% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")

        st.divider()
        st.markdown("### 🔻 TVA Déductible")
        col_d1, col_d2, col_d3 = st.columns(3)
        with col_d1:
            data["TVA déductible sur achats (ABS)"] = st.number_input("TVA ded. ABS — 44566 (€)", min_value=0.0, step=10.0, format="%.2f")
        with col_d2:
            data["TVA déductible sur immobilisations"] = st.number_input("TVA ded. Immos — 44562 (€)", min_value=0.0, step=10.0, format="%.2f")
        with col_d3:
            data["TVA intracommunautaire déductible"] = st.number_input("TVA intracom. ded. — 44563 (€)", min_value=0.0, step=10.0, format="%.2f")

        data["Crédit de TVA période précédente"] = st.number_input("Crédit TVA période précédente — 44567 (€)", min_value=0.0, step=10.0, format="%.2f")

        st.divider()
        if st.button("🧮 Calculer la déclaration TVA", type="primary", use_container_width=True):
            res = _calculer_tva(data)
            _afficher_resultats(res, data, periode, entreprise, regime)

    # ══════════════════════════════════════════
    # TAB 2 — Import balance
    # ══════════════════════════════════════════
    with tab2:
        st.markdown("### 📁 Import balance comptable")
        st.info("La balance est analysée pour extraire automatiquement les comptes TVA (445xx) et CA (70x).")

        uploaded = st.file_uploader("Balance CSV / Excel", type=["csv", "xlsx"], key="tva_balance")
        if uploaded:
            try:
                from utils.intelligent_parser import parser_balance_intelligent
                with st.spinner("Analyse de la balance..."):
                    df, info = parser_balance_intelligent(uploaded)
                st.success(f"✅ {len(df):,} comptes chargés — {info.get('format_detecte', 'format détecté')}")

                extrait = _extraire_tva_depuis_balance(df)
                if extrait:
                    st.markdown("### 📊 Comptes TVA extraits")
                    df_ext = pd.DataFrame([
                        {"Compte / Rubrique": k, "Montant (€)": f"{v:,.2f}"}
                        for k, v in extrait.items() if v != 0
                    ])
                    st.dataframe(df_ext, use_container_width=True, hide_index=True)

                    st.markdown("---")
                    st.markdown("#### 🔍 Vérification de cohérence")
                    col_tvac = extrait.get("TVA collectée (44571)", 0)
                    col_tvad = extrait.get("TVA déductible ABS (44566)", 0) + extrait.get("TVA déductible immos (44562)", 0)
                    solde_balance = col_tvac - col_tvad

                    col1, col2, col3 = st.columns(3)
                    col1.metric("TVA Collectée (44571)", f"{col_tvac:,.0f} €")
                    col2.metric("TVA Déductible", f"{col_tvad:,.0f} €")
                    col3.metric("Solde", f"{solde_balance:,.0f} €",
                                delta="À payer" if solde_balance > 0 else "Crédit",
                                delta_color="inverse" if solde_balance > 0 else "normal")

                    # Pré-remplir la saisie manuelle
                    st.info("💡 Utilisez l'onglet **Saisie manuelle** pour affiner et générer la déclaration complète.")
                else:
                    st.warning("Aucun compte TVA (445xx) détecté dans la balance.")

            except Exception as e:
                st.error(f"Erreur : {e}")
                import traceback
                with st.expander("Détails"):
                    st.code(traceback.format_exc())

    # ══════════════════════════════════════════
    # TAB 3 — Guide
    # ══════════════════════════════════════════
    with tab3:
        st.markdown("### 📖 Taux de TVA applicables en France")
        df_taux = pd.DataFrame([
            {"Taux": "20%", "Catégorie": "Taux normal", "Exemples": "Biens et services courants, honoraires, conseil"},
            {"Taux": "10%", "Catégorie": "Taux intermédiaire", "Exemples": "Restauration, transport, travaux logement, médicaments remboursables"},
            {"Taux": "5,5%", "Catégorie": "Taux réduit", "Exemples": "Alimentation, livres, abonnements énergie, équipement PMR"},
            {"Taux": "2,1%", "Catégorie": "Taux super-réduit", "Exemples": "Presse, médicaments remboursés SS, spectacles vivants (100 premières représentations)"},
            {"Taux": "0%", "Catégorie": "Exonéré", "Exemples": "Exports hors UE, intracommunautaire, activités médicales, enseignement, assurance"},
        ])
        st.dataframe(df_taux, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("### 🏦 Comptes PCG — TVA")
        df_pcg = pd.DataFrame([
            {"Compte": k, "Nature": v} for k, v in COMPTES_TVA_PCG.items()
        ])
        st.dataframe(df_pcg, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("### 📋 CA3 vs CA12 — Comparatif régimes")
        df_regime = pd.DataFrame([
            {"Critère": "Fréquence", "CA3 (Réel normal)": "Mensuelle ou trimestrielle", "CA12 (Réel simplifié)": "Annuelle (mai N+1)"},
            {"Critère": "CA seuil (BIC/BNC)", "CA3 (Réel normal)": "> 840 000 € (négoce) / 254 000 € (services)", "CA12 (Réel simplifié)": "< 840 000 € (négoce) / 254 000 € (services)"},
            {"Critère": "Acomptes", "CA3 (Réel normal)": "Aucun — déclaration mensuelle", "CA12 (Réel simplifié)": "2 acomptes (55% en juil., 40% en déc.)"},
            {"Critère": "Crédit TVA", "CA3 (Réel normal)": "Remboursable dès 760 €", "CA12 (Réel simplifié)": "Sur demande ou imputation"},
            {"Critère": "Comptabilité", "CA3 (Réel normal)": "Obligatoirement complète", "CA12 (Réel simplifié)": "Simplifiée possible"},
        ])
        st.dataframe(df_regime, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("### ⚡ Autoliquidation — Cas principaux")
        df_auto = pd.DataFrame([
            {"Opération": "Acquisitions intracommunautaires (AIC)", "Mécanisme": "Autoliquidation — acheteur déclare et déduit simultanément"},
            {"Opération": "Sous-traitance BTP", "Mécanisme": "Donneur d'ordre autoliquide la TVA du sous-traitant"},
            {"Opération": "Services étrangers (art. 283-1)", "Mécanisme": "Preneur français autoliquide si prestataire non établi en France"},
            {"Opération": "Livraisons intracommunautaires (LIC)", "Mécanisme": "Exonéré côté vendeur — TVA du pays acheteur"},
        ])
        st.dataframe(df_auto, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────
# AFFICHAGE RÉSULTATS (partagé saisie/import)
# ─────────────────────────────────────────────

def _afficher_resultats(res: dict, data: dict, periode: str, entreprise: str, regime: str):
    st.divider()
    st.markdown(f"## 📋 Déclaration TVA — {entreprise} — {periode}")
    st.caption(regime)

    # Métriques
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("💰 TVA Collectée", f"{res['tva_collectee']:,.0f} €")
    col2.metric("🔻 TVA Déductible", f"{res['tva_deductible']:,.0f} €")
    if res['a_payer'] > 0:
        col3.metric("💸 TVA à Décaisser", f"{res['a_payer']:,.0f} €", delta="À payer", delta_color="inverse")
        col4.metric("✅ Crédit TVA", "0 €")
    else:
        col3.metric("💸 TVA à Décaisser", "0 €")
        col4.metric("✅ Crédit TVA", f"{res['credit_genere']:,.0f} €", delta="Crédit", delta_color="normal")

    st.divider()

    # Détail collectée
    if res["tva_collectee_detail"]:
        st.markdown("### 🔺 Détail TVA Collectée")
        rows = []
        for rubrique, detail in res["tva_collectee_detail"].items():
            rows.append({
                "Rubrique": rubrique,
                "Base HT (€)": f"{detail['base_ht']:,.2f}",
                "Taux": f"{detail['taux']*100:.1f}%",
                "TVA (€)": f"{detail['tva']:,.2f}",
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # Détail déductible
    st.markdown("### 🔻 Détail TVA Déductible")
    rows_ded = []
    if res["tva_ded_abs"] > 0:
        rows_ded.append({"Compte": "44566", "Nature": "TVA déductible sur ABS", "Montant (€)": f"{res['tva_ded_abs']:,.2f}"})
    if res["tva_ded_immo"] > 0:
        rows_ded.append({"Compte": "44562", "Nature": "TVA déductible sur immos", "Montant (€)": f"{res['tva_ded_immo']:,.2f}"})
    if res["tva_ded_intra"] > 0:
        rows_ded.append({"Compte": "44563", "Nature": "TVA intracom. déductible", "Montant (€)": f"{res['tva_ded_intra']:,.2f}"})
    if res["credit_reporte"] > 0:
        rows_ded.append({"Compte": "44567", "Nature": "Crédit TVA période précédente", "Montant (€)": f"{res['credit_reporte']:,.2f}"})
    if rows_ded:
        st.dataframe(pd.DataFrame(rows_ded), use_container_width=True, hide_index=True)
    else:
        st.info("Aucune TVA déductible saisie.")

    # Synthèse
    st.markdown("### 🏁 Synthèse")
    synth_data = {
        "TVA Collectée": res["tva_collectee"],
        "TVA Déductible": res["tva_deductible"],
        "Solde (+ = à payer)": res["solde"],
    }
    cols_s = st.columns(3)
    for i, (k, v) in enumerate(synth_data.items()):
        cols_s[i].metric(k, f"{v:,.2f} €")

    # Graphique
    st.plotly_chart(_chart_tva(res), use_container_width=True)

    # Alertes cohérence
    alertes = _verifier_coherence(res)
    if alertes:
        st.divider()
        st.markdown("### ⚠️ Contrôles de cohérence")
        for typ, msg in alertes:
            if typ == "warning":
                st.warning(msg)
            elif typ == "success":
                st.success(msg)
            else:
                st.info(msg)

    # Écriture PCG à passer
    st.divider()
    st.markdown("### 📝 Écriture comptable à passer (PCG)")
    if res["a_payer"] > 0:
        st.code(f"""
Débit  44551 — TVA à décaisser      {res['a_payer']:>12,.2f} €
  Crédit 512 — Banque                        {res['a_payer']:>12,.2f} €
  → Règlement TVA {periode}
""", language="text")
    else:
        st.code(f"""
Débit  44567 — Crédit de TVA        {res['credit_genere']:>12,.2f} €
  Crédit 44551 — TVA à décaisser             {res['credit_genere']:>12,.2f} €
  → Report crédit TVA {periode}
""", language="text")

    # Export Excel
    st.divider()
    try:
        excel = _export_excel_tva(res, data, periode, entreprise)
        st.download_button(
            label="📥 Télécharger Excel Déclaration TVA",
            data=excel,
            file_name=f"TVA_{entreprise}_{periode.replace('/', '-')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    except Exception as e:
        st.warning(f"Export Excel non disponible : {e}")
