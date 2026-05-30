# -*- coding: utf-8 -*-
"""
Module Comparatif N vs N-1
Bilan + Compte de Résultat côte à côte avec écarts €/%
SMD Consulting - PCG France
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
from datetime import datetime


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _ecart(val_n: float, val_n1: float):
    """Retourne (ecart_abs, ecart_pct)"""
    ecart_abs = val_n - val_n1
    if val_n1 != 0:
        ecart_pct = (ecart_abs / abs(val_n1)) * 100
    else:
        ecart_pct = 100.0 if val_n != 0 else 0.0
    return ecart_abs, ecart_pct


def _fmt(v: float) -> str:
    if v is None:
        return "—"
    return f"{v:,.0f} €"


def _fmt_pct(v: float) -> str:
    if v is None:
        return "—"
    sign = "+" if v >= 0 else ""
    return f"{sign}{v:.1f}%"


def _fmt_ecart(v: float) -> str:
    sign = "+" if v >= 0 else ""
    return f"{sign}{v:,.0f} €"


# ─────────────────────────────────────────────
# TABLE COMPARATIVE GÉNÉRIQUE
# ─────────────────────────────────────────────

def _build_comparatif_df(dict_n: dict, dict_n1: dict, label_n: str, label_n1: str) -> pd.DataFrame:
    """Construit un DataFrame comparatif à partir de deux dictionnaires montants."""
    rows = []
    all_keys = list(dict_n.keys())
    for k in all_keys:
        vn = dict_n.get(k, 0) or 0
        vn1 = dict_n1.get(k, 0) or 0
        ea, ep = _ecart(vn, vn1)
        rows.append({
            'Rubrique': k,
            label_n1: vn1,
            label_n: vn,
            'Écart (€)': ea,
            'Écart (%)': ep,
        })
    return pd.DataFrame(rows)


def _style_comparatif(df: pd.DataFrame, col_n: str, col_n1: str):
    """Applique couleurs : vert écart positif, rouge négatif pour totaux."""
    def color_ecart(val):
        if isinstance(val, (int, float)):
            if val > 0:
                return 'color: #28a745; font-weight: bold'
            elif val < 0:
                return 'color: #dc3545; font-weight: bold'
        return ''

    def bold_totaux(row):
        if row['Rubrique'].startswith('TOTAL') or row['Rubrique'].startswith('Résultat'):
            return ['font-weight: bold; background-color: #f0f2f6'] * len(row)
        return [''] * len(row)

    styled = (
        df.style
        .apply(bold_totaux, axis=1)
        .applymap(color_ecart, subset=['Écart (€)', 'Écart (%)'])
        .format({col_n1: '{:,.0f}', col_n: '{:,.0f}', 'Écart (€)': '{:+,.0f}', 'Écart (%)': '{:+.1f}'})
    )
    return styled


# ─────────────────────────────────────────────
# GRAPHIQUES
# ─────────────────────────────────────────────

def _chart_cdr_comparatif(sig_n: dict, sig_n1: dict, label_n: str, label_n1: str):
    """Bar chart groupé SIG N vs N-1."""
    indicateurs = [
        "Chiffre d'affaires",
        "Marge commerciale",
        "Valeur ajoutée (VA)",
        "Excedent Brut d'Exploitation (EBE)",
        "Resultat d'exploitation",
        "Resultat net"
    ]
    labels, vals_n1, vals_n = [], [], []
    for ind in indicateurs:
        vn = sig_n.get(ind, 0) or 0
        vn1 = sig_n1.get(ind, 0) or 0
        if vn != 0 or vn1 != 0:
            labels.append(ind.replace("Excedent Brut d'Exploitation", "EBE"))
            vals_n1.append(vn1)
            vals_n.append(vn)

    fig = go.Figure(data=[
        go.Bar(name=label_n1, x=labels, y=vals_n1, marker_color='#aab7d4'),
        go.Bar(name=label_n, x=labels, y=vals_n, marker_color='#1f77b4'),
    ])
    fig.update_layout(
        barmode='group',
        title="SIG — Comparaison N vs N-1",
        yaxis_title="Montant (€)",
        legend=dict(orientation='h', yanchor='bottom', y=1.02),
        height=400
    )
    return fig


def _chart_bilan_comparatif(bilan_n: dict, bilan_n1: dict, label_n: str, label_n1: str):
    """Bar chart structure bilan N vs N-1."""
    postes = [
        ('actif', 'TOTAL IMMOBILISATIONS'),
        ('actif', 'TOTAL ACTIF CIRCULANT'),
        ('actif', 'TOTAL TRESORERIE'),
        ('passif', 'Capitaux propres'),
        ('passif', 'Dettes financieres (16)'),
        ('passif', 'TOTAL DETTES'),
    ]
    labels, vals_n1, vals_n = [], [], []
    for section, cle in postes:
        vn = bilan_n.get(section, {}).get(cle, 0) or 0
        vn1 = bilan_n1.get(section, {}).get(cle, 0) or 0
        if vn != 0 or vn1 != 0:
            labels.append(cle.replace('TOTAL ', '').title())
            vals_n1.append(abs(vn1))
            vals_n.append(abs(vn))

    fig = go.Figure(data=[
        go.Bar(name=label_n1, x=labels, y=vals_n1, marker_color='#aab7d4'),
        go.Bar(name=label_n, x=labels, y=vals_n, marker_color='#2ca02c'),
    ])
    fig.update_layout(
        barmode='group',
        title="Bilan — Comparaison N vs N-1",
        yaxis_title="Montant (€)",
        legend=dict(orientation='h', yanchor='bottom', y=1.02),
        height=400
    )
    return fig


# ─────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────

def _export_excel_comparatif(
    df_cdr: pd.DataFrame,
    df_actif: pd.DataFrame,
    df_passif: pd.DataFrame,
    entreprise: str,
    label_n: str,
    label_n1: str
) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Feuille Compte de Résultat
        df_cdr.to_excel(writer, sheet_name='Compte de Résultat', index=False)
        ws_cdr = writer.sheets['Compte de Résultat']
        _style_excel_sheet(ws_cdr)

        # Feuille Bilan Actif
        df_actif.to_excel(writer, sheet_name='Bilan Actif', index=False)
        _style_excel_sheet(writer.sheets['Bilan Actif'])

        # Feuille Bilan Passif
        df_passif.to_excel(writer, sheet_name='Bilan Passif', index=False)
        _style_excel_sheet(writer.sheets['Bilan Passif'])

        # Feuille Synthèse
        synth_rows = [
            {'Indicateur': 'Rapport généré', 'Valeur': datetime.now().strftime('%d/%m/%Y %H:%M')},
            {'Indicateur': 'Entreprise', 'Valeur': entreprise},
            {'Indicateur': 'Exercice N', 'Valeur': label_n},
            {'Indicateur': 'Exercice N-1', 'Valeur': label_n1},
        ]
        pd.DataFrame(synth_rows).to_excel(writer, sheet_name='Synthèse', index=False)

    return buf.getvalue()


def _style_excel_sheet(ws):
    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 35)


# ─────────────────────────────────────────────
# PAGE PRINCIPALE
# ─────────────────────────────────────────────

def page_comparatif():
    st.title("📊 Comparatif N vs N-1")
    st.markdown("**Comparaison Bilan + Compte de Résultat** entre deux exercices avec analyse des écarts")
    st.caption("✨ Pour Cabinets, DAF et Dirigeants - PCG France")

    from utils.intelligent_parser import parser_balance_intelligent
    from utils.compte_resultat import calculer_compte_resultat
    from utils.bilan import calculer_bilan

    # ── Paramètres ──
    col1, col2, col3 = st.columns(3)
    with col1:
        entreprise = st.text_input("🏢 Entreprise", value="Entreprise")
    with col2:
        label_n = st.text_input("📅 Exercice N", value=str(datetime.now().year))
    with col3:
        label_n1 = st.text_input("📅 Exercice N-1", value=str(datetime.now().year - 1))

    col_type, _ = st.columns([1, 2])
    with col_type:
        type_entreprise = st.selectbox("🏭 Type", ["Mixte", "Commerciale", "Industrielle", "Services"])

    st.divider()

    # ── Upload balances ──
    col_n, col_n1 = st.columns(2)
    with col_n:
        st.markdown(f"### 📁 Balance N ({label_n})")
        file_n = st.file_uploader(
            "Balance N",
            type=["csv", "xlsx", "txt"],
            key="cmp_n",
            label_visibility="collapsed"
        )
    with col_n1:
        st.markdown(f"### 📁 Balance N-1 ({label_n1})")
        file_n1 = st.file_uploader(
            "Balance N-1",
            type=["csv", "xlsx", "txt"],
            key="cmp_n1",
            label_visibility="collapsed"
        )

    if not file_n or not file_n1:
        st.info("👆 Déposez les deux balances pour lancer le comparatif.")
        st.markdown("""
        **Format accepté :** Balance CSV/Excel avec colonnes `CompteNum`, `Debit`, `Credit`
        (ou `SoldeDebit` / `SoldeCredit`).
        """)
        return

    # ── Parsing ──
    try:
        with st.spinner("🤖 Analyse balance N..."):
            df_n, info_n = parser_balance_intelligent(file_n)
        with st.spinner("🤖 Analyse balance N-1..."):
            df_n1, info_n1 = parser_balance_intelligent(file_n1)
    except Exception as e:
        st.error(f"❌ Erreur parsing balance : {e}")
        return

    st.success(
        f"✅ N ({label_n}) : **{len(df_n):,} comptes** | "
        f"N-1 ({label_n1}) : **{len(df_n1):,} comptes**"
    )
    st.divider()

    # ── Calculs ──
    try:
        cdr_n = calculer_compte_resultat(df_n, type_entreprise)
        cdr_n1 = calculer_compte_resultat(df_n1, type_entreprise)
        bilan_n = calculer_bilan(df_n)
        bilan_n1 = calculer_bilan(df_n1)
    except Exception as e:
        st.error(f"❌ Erreur de calcul : {e}")
        import traceback
        with st.expander("Détails"):
            st.code(traceback.format_exc())
        return

    if 'erreur' in cdr_n:
        st.error(f"CdR N : {cdr_n['erreur']}")
        return
    if 'erreur' in cdr_n1:
        st.error(f"CdR N-1 : {cdr_n1['erreur']}")
        return

    # ═══════════════════════════════════════════
    # SECTION 1 — COMPTE DE RÉSULTAT
    # ═══════════════════════════════════════════
    st.markdown("## 📈 Compte de Résultat")

    sig_n = cdr_n['sig']
    sig_n1 = cdr_n1['sig']

    # Métriques clés
    indicateurs_cles = [
        ("💰 CA", "Chiffre d'affaires"),
        ("⚙️ VA", "Valeur ajoutée (VA)"),
        ("📈 EBE", "Excedent Brut d'Exploitation (EBE)"),
        ("🎯 Résultat Net", "Resultat net"),
    ]
    cols = st.columns(4)
    for i, (label_m, cle) in enumerate(indicateurs_cles):
        vn = sig_n.get(cle, 0) or 0
        vn1 = sig_n1.get(cle, 0) or 0
        ea, ep = _ecart(vn, vn1)
        with cols[i]:
            st.metric(
                label=label_m,
                value=_fmt(vn),
                delta=f"{_fmt_ecart(ea)} ({_fmt_pct(ep)})",
                delta_color="normal" if ea >= 0 else "inverse"
            )

    st.divider()

    # Tableau SIG
    st.markdown("### 📋 Soldes Intermédiaires de Gestion")
    df_sig = _build_comparatif_df(sig_n, sig_n1, label_n, label_n1)
    st.dataframe(
        df_sig.style
        .format({label_n1: '{:,.0f}', label_n: '{:,.0f}', 'Écart (€)': '{:+,.0f}', 'Écart (%)': '{:+.1f}'})
        .applymap(
            lambda v: 'color: #28a745; font-weight:bold' if isinstance(v, (int, float)) and v > 0
            else ('color: #dc3545; font-weight:bold' if isinstance(v, (int, float)) and v < 0 else ''),
            subset=['Écart (€)', 'Écart (%)']
        ),
        use_container_width=True, hide_index=True
    )

    # Graphique SIG
    st.plotly_chart(_chart_cdr_comparatif(sig_n, sig_n1, label_n, label_n1), use_container_width=True)

    # Tableaux charges / produits
    col_p, col_c = st.columns(2)
    with col_p:
        st.markdown("### 💰 Produits N vs N-1")
        df_prod = _build_comparatif_df(cdr_n['produits'], cdr_n1['produits'], label_n, label_n1)
        df_prod = df_prod[df_prod[label_n].abs() + df_prod[label_n1].abs() > 0]
        st.dataframe(df_prod.style.format(
            {label_n1: '{:,.0f}', label_n: '{:,.0f}', 'Écart (€)': '{:+,.0f}', 'Écart (%)': '{:+.1f}'}
        ), use_container_width=True, hide_index=True)
    with col_c:
        st.markdown("### 💸 Charges N vs N-1")
        df_chg = _build_comparatif_df(cdr_n['charges'], cdr_n1['charges'], label_n, label_n1)
        df_chg = df_chg[df_chg[label_n].abs() + df_chg[label_n1].abs() > 0]
        st.dataframe(df_chg.style.format(
            {label_n1: '{:,.0f}', label_n: '{:,.0f}', 'Écart (€)': '{:+,.0f}', 'Écart (%)': '{:+.1f}'}
        ), use_container_width=True, hide_index=True)

    # ═══════════════════════════════════════════
    # SECTION 2 — BILAN
    # ═══════════════════════════════════════════
    st.divider()
    st.markdown("## 📊 Bilan Comptable")

    # Métriques bilan
    bilan_cles = [
        ("🏗️ Total Actif", 'actif', 'TOTAL ACTIF'),
        ("💼 Capitaux Propres", 'passif', 'Capitaux propres'),
        ("🏦 FDR", 'ratios', 'Fonds de roulement net global (FRNG)'),
        ("⚡ BFR", 'ratios', 'Besoin en fonds de roulement (BFR)'),
    ]
    cols2 = st.columns(4)
    for i, (label_m, section, cle) in enumerate(bilan_cles):
        vn = (bilan_n.get(section) or {}).get(cle, 0) or 0
        vn1 = (bilan_n1.get(section) or {}).get(cle, 0) or 0
        ea, ep = _ecart(vn, vn1)
        with cols2[i]:
            st.metric(
                label=label_m,
                value=_fmt(vn),
                delta=f"{_fmt_ecart(ea)} ({_fmt_pct(ep)})",
                delta_color="normal" if ea >= 0 else "inverse"
            )

    st.divider()

    # Tableaux ACTIF / PASSIF
    col_a, col_p2 = st.columns(2)
    with col_a:
        st.markdown("### 📦 ACTIF N vs N-1")
        df_actif = _build_comparatif_df(bilan_n['actif'], bilan_n1['actif'], label_n, label_n1)
        df_actif = df_actif[df_actif[label_n].abs() + df_actif[label_n1].abs() > 0]
        st.dataframe(df_actif.style.format(
            {label_n1: '{:,.0f}', label_n: '{:,.0f}', 'Écart (€)': '{:+,.0f}', 'Écart (%)': '{:+.1f}'}
        ).applymap(
            lambda v: 'color: #28a745; font-weight:bold' if isinstance(v, (int, float)) and v > 0
            else ('color: #dc3545; font-weight:bold' if isinstance(v, (int, float)) and v < 0 else ''),
            subset=['Écart (€)', 'Écart (%)']
        ), use_container_width=True, hide_index=True)
    with col_p2:
        st.markdown("### 🏛️ PASSIF N vs N-1")
        df_passif = _build_comparatif_df(bilan_n['passif'], bilan_n1['passif'], label_n, label_n1)
        df_passif = df_passif[df_passif[label_n].abs() + df_passif[label_n1].abs() > 0]
        st.dataframe(df_passif.style.format(
            {label_n1: '{:,.0f}', label_n: '{:,.0f}', 'Écart (€)': '{:+,.0f}', 'Écart (%)': '{:+.1f}'}
        ).applymap(
            lambda v: 'color: #28a745; font-weight:bold' if isinstance(v, (int, float)) and v > 0
            else ('color: #dc3545; font-weight:bold' if isinstance(v, (int, float)) and v < 0 else ''),
            subset=['Écart (€)', 'Écart (%)']
        ), use_container_width=True, hide_index=True)

    # Graphique bilan
    st.plotly_chart(_chart_bilan_comparatif(bilan_n, bilan_n1, label_n, label_n1), use_container_width=True)

    # ═══════════════════════════════════════════
    # SECTION 3 — ANALYSE AUTOMATIQUE
    # ═══════════════════════════════════════════
    st.divider()
    st.markdown("## 💡 Analyse des Écarts")

    alertes = []
    # CA
    ca_n = sig_n.get("Chiffre d'affaires", 0) or 0
    ca_n1 = sig_n1.get("Chiffre d'affaires", 0) or 0
    if ca_n1 > 0:
        ea_ca = (ca_n - ca_n1) / abs(ca_n1) * 100
        if ea_ca < -10:
            alertes.append(('error', f"🔴 CA en baisse de {abs(ea_ca):.1f}% — analyse des causes requise"))
        elif ea_ca > 15:
            alertes.append(('success', f"✅ CA en hausse de {ea_ca:.1f}% — performance commerciale solide"))
        else:
            alertes.append(('info', f"ℹ️ CA stable ({ea_ca:+.1f}%)"))

    # Résultat net
    rn_n = sig_n.get("Resultat net", 0) or 0
    rn_n1 = sig_n1.get("Resultat net", 0) or 0
    if rn_n < 0 and rn_n1 >= 0:
        alertes.append(('error', "🔴 Résultat net : passage en déficit — vigilance requise"))
    elif rn_n > 0 and rn_n1 < 0:
        alertes.append(('success', "✅ Résultat net : retour à l'excédent"))
    elif rn_n1 != 0:
        ea_rn, _ = _ecart(rn_n, rn_n1)
        if abs(ea_rn) / abs(rn_n1) > 0.20:
            alertes.append(('warning', f"⚠️ Résultat net : variation significative de {ea_rn:+,.0f} €"))

    # BFR
    bfr_n = (bilan_n.get('ratios') or {}).get('Besoin en fonds de roulement (BFR)', 0) or 0
    bfr_n1 = (bilan_n1.get('ratios') or {}).get('Besoin en fonds de roulement (BFR)', 0) or 0
    if bfr_n > bfr_n1 * 1.20 and bfr_n1 > 0:
        alertes.append(('warning', f"⚠️ BFR en hausse de {((bfr_n/bfr_n1)-1)*100:.1f}% — surveiller le cycle d'exploitation"))

    # FDR
    fdr_n = (bilan_n.get('ratios') or {}).get('Fonds de roulement net global (FRNG)', 0) or 0
    if fdr_n < 0:
        alertes.append(('error', "🔴 Fonds de Roulement négatif — risque de déséquilibre financier"))

    for typ, msg in alertes:
        if typ == 'success':
            st.success(msg)
        elif typ == 'warning':
            st.warning(msg)
        elif typ == 'error':
            st.error(msg)
        else:
            st.info(msg)

    if not alertes:
        st.info("ℹ️ Aucune alerte significative détectée.")

    # ═══════════════════════════════════════════
    # EXPORT EXCEL
    # ═══════════════════════════════════════════
    st.divider()
    df_actif_exp = _build_comparatif_df(bilan_n['actif'], bilan_n1['actif'], label_n, label_n1)
    df_passif_exp = _build_comparatif_df(bilan_n['passif'], bilan_n1['passif'], label_n, label_n1)

    try:
        excel_bytes = _export_excel_comparatif(
            df_sig, df_actif_exp, df_passif_exp,
            entreprise, label_n, label_n1
        )
        st.download_button(
            label="📥 Télécharger Excel Comparatif",
            data=excel_bytes,
            file_name=f"Comparatif_{entreprise}_{label_n}_vs_{label_n1}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    except Exception as e:
        st.warning(f"Export Excel non disponible : {e}")
