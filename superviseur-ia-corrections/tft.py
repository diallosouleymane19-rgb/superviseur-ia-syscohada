# -*- coding: utf-8 -*-
"""
Module TFT PCG France — Méthode indirecte — SMD Consulting
Tableau de Flux de Trésorerie conforme ANC/CRC 99-02
Horizon 1 à 3 exercices comparatifs
"""
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO

# ─── Structure TFT méthode indirecte (ANC / CRC 99-02) ───────────────────────

TFT_STRUCTURE = {
    "I. FLUX DE TRÉSORERIE LIÉS À L'ACTIVITÉ": {
        "color": "#1E8449",
        "lignes": [
            ("Résultat net (bénéfice + / perte -)", "+"),
            ("Dotations aux amortissements et provisions (nettes de reprises)", "+"),
            ("Plus-values de cessions nettes d'impôts", "-"),
            ("Moins-values de cessions nettes d'impôts", "+"),
            ("Variation des stocks (augmentation -)", "±"),
            ("Variation des créances d'exploitation (augmentation -)", "±"),
            ("Variation des dettes d'exploitation (augmentation +)", "±"),
            ("Variation des autres créances (augmentation -)", "±"),
            ("Variation des autres dettes (augmentation +)", "±"),
            ("Impôts sur les sociétés payés", "-"),
            ("Dividendes reçus des participations", "+"),
        ]
    },
    "II. FLUX DE TRÉSORERIE LIÉS AUX OPÉRATIONS D'INVESTISSEMENT": {
        "color": "#2980B9",
        "lignes": [
            ("Acquisitions d'immobilisations incorporelles", "-"),
            ("Acquisitions d'immobilisations corporelles", "-"),
            ("Acquisitions d'immobilisations financières", "-"),
            ("Cessions d'immobilisations incorporelles", "+"),
            ("Cessions d'immobilisations corporelles", "+"),
            ("Cessions d'immobilisations financières", "+"),
            ("Variation des créances sur cessions d'actifs", "±"),
        ]
    },
    "III. FLUX DE TRÉSORERIE LIÉS AUX OPÉRATIONS DE FINANCEMENT": {
        "color": "#8E44AD",
        "lignes": [
            ("Augmentation de capital en numéraire", "+"),
            ("Remboursements de capital", "-"),
            ("Émission d'emprunts", "+"),
            ("Remboursements d'emprunts", "-"),
            ("Dividendes versés aux actionnaires", "-"),
            ("Variation des concours bancaires courants", "±"),
        ]
    }
}

SECTION_KEYS = list(TFT_STRUCTURE.keys())
K_OP  = SECTION_KEYS[0]
K_INV = SECTION_KEYS[1]
K_FIN = SECTION_KEYS[2]


# ─── Calcul ──────────────────────────────────────────────────────────────────

def _calculer_tft(data: dict, exercices: list) -> dict:
    resultats = {}
    for ex in exercices:
        flux_op  = sum(data.get(lib, {}).get(ex, 0) for lib, _ in TFT_STRUCTURE[K_OP]["lignes"])
        flux_inv = sum(data.get(lib, {}).get(ex, 0) for lib, _ in TFT_STRUCTURE[K_INV]["lignes"])
        flux_fin = sum(data.get(lib, {}).get(ex, 0) for lib, _ in TFT_STRUCTURE[K_FIN]["lignes"])
        var_nette = flux_op + flux_inv + flux_fin
        treso_ouv = data.get("Trésorerie à l'ouverture", {}).get(ex, 0)
        treso_clo = treso_ouv + var_nette
        resultats[ex] = {
            "Flux activité (I)": flux_op,
            "Flux investissement (II)": flux_inv,
            "Flux financement (III)": flux_fin,
            "Variation nette (I+II+III)": var_nette,
            "Trésorerie ouverture": treso_ouv,
            "Trésorerie clôture": treso_clo,
        }
    return resultats


@st.cache_data(show_spinner=False)
def _extraire_depuis_balance(fichier_bytes: bytes, nom: str) -> dict:
    """Extrait résultat net, dotations, variation stocks/créances/dettes depuis balance PCG."""
    try:
        if nom.endswith(".xlsx"):
            df = pd.read_excel(BytesIO(fichier_bytes))
        else:
            df = pd.read_csv(BytesIO(fichier_bytes), sep=None, engine="python")
        df.columns = [str(c).strip().lower() for c in df.columns]
        col_c = next((c for c in df.columns if any(k in c for k in ["compte", "cpte"])), None)
        col_s = next((c for c in df.columns if any(k in c for k in ["solde", "credit", "crédit"])), None)
        if not col_c or not col_s:
            return {}
        df[col_c] = df[col_c].astype(str).str.strip()

        def s(prefixes):
            m = df[col_c].str.startswith(tuple(prefixes), na=False)
            return abs(df.loc[m, col_s].apply(pd.to_numeric, errors="coerce").fillna(0).sum())

        return {
            "Résultat net (bénéfice + / perte -)": s(["120", "121"]) - s(["129"]),
            "Dotations aux amortissements et provisions (nettes de reprises)": s(["681", "682", "686", "687"]) - s(["781", "786", "787"]),
            "Variation des stocks (augmentation -)": -s(["3"]),
            "Variation des créances d'exploitation (augmentation -)": -s(["411", "413", "409"]),
            "Variation des dettes d'exploitation (augmentation +)": s(["401", "403", "421", "431", "437", "445"]),
            "Acquisitions d'immobilisations corporelles": -s(["21", "22", "23"]),
            "Acquisitions d'immobilisations incorporelles": -s(["20"]),
            "Émission d'emprunts": s(["164", "165", "166", "167"]),
            "Dividendes versés aux actionnaires": -s(["457"]),
        }
    except Exception:
        return {}


# ─── Graphique ───────────────────────────────────────────────────────────────

def _chart_tft(resultats: dict, exercices: list) -> go.Figure:
    labels = ["Activité", "Investissement", "Financement"]
    keys = ["Flux activité (I)", "Flux investissement (II)", "Flux financement (III)"]
    colors = ["#1E8449", "#2980B9", "#8E44AD"]

    fig = go.Figure()
    for label, key, color in zip(labels, keys, colors):
        fig.add_trace(go.Bar(
            name=label, x=exercices,
            y=[resultats[ex][key] for ex in exercices],
            marker_color=color,
            hovertemplate="%{y:+,.0f} €<extra>" + label + "</extra>"
        ))
    fig.add_trace(go.Scatter(
        name="Trésorerie clôture", x=exercices,
        y=[resultats[ex]["Trésorerie clôture"] for ex in exercices],
        mode="lines+markers",
        line=dict(color="#E74C3C", width=2.5),
        marker=dict(size=9),
        hovertemplate="%{y:,.0f} €<extra>Trésorerie clôture</extra>"
    ))
    fig.add_hline(y=0, line_dash="dash", line_color="#888", line_width=1)
    fig.update_layout(
        title="Tableau de Flux de Trésorerie — Méthode indirecte",
        barmode="group", height=420,
        margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", y=-0.15),
        yaxis_tickformat=",.0f",
    )
    return fig


# ─── Export Excel ────────────────────────────────────────────────────────────

def _export_excel_tft(data: dict, resultats: dict, exercices: list, entreprise: str) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rows = []
        for section, info in TFT_STRUCTURE.items():
            rows.append({"Libellé": section, **{ex: "" for ex in exercices}, "_type": "section"})
            for lib, signe in info["lignes"]:
                row = {"Libellé": f"  {lib} ({signe})", "_type": "ligne"}
                for ex in exercices:
                    row[ex] = data.get(lib, {}).get(ex, 0)
                rows.append(row)

        rows.append({"Libellé": "FLUX ACTIVITÉ (I)", **{ex: resultats[ex]["Flux activité (I)"] for ex in exercices}, "_type": "total"})
        rows.append({"Libellé": "FLUX INVESTISSEMENT (II)", **{ex: resultats[ex]["Flux investissement (II)"] for ex in exercices}, "_type": "total"})
        rows.append({"Libellé": "FLUX FINANCEMENT (III)", **{ex: resultats[ex]["Flux financement (III)"] for ex in exercices}, "_type": "total"})
        rows.append({"Libellé": "VARIATION NETTE (I+II+III)", **{ex: resultats[ex]["Variation nette (I+II+III)"] for ex in exercices}, "_type": "total"})
        rows.append({"Libellé": "Trésorerie ouverture", **{ex: resultats[ex]["Trésorerie ouverture"] for ex in exercices}, "_type": "ligne"})
        rows.append({"Libellé": "TRÉSORERIE CLÔTURE", **{ex: resultats[ex]["Trésorerie clôture"] for ex in exercices}, "_type": "total"})

        df_export = pd.DataFrame(rows).drop(columns=["_type"])
        df_export.to_excel(writer, sheet_name="TFT", index=False)

        ws = writer.sheets["TFT"]
        types = [r["_type"] for r in rows]
        section_fill = PatternFill("solid", fgColor="2C3E50")
        total_fill   = PatternFill("solid", fgColor="1E8449")

        for i, t in enumerate(types, start=2):
            if t == "section":
                for cell in ws[i]:
                    cell.fill = section_fill
                    cell.font = Font(color="FFFFFF", bold=True)
            elif t == "total":
                for cell in ws[i]:
                    cell.fill = total_fill
                    cell.font = Font(color="FFFFFF", bold=True)

        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 35

    return buf.getvalue()


# ─── Analyse IA ──────────────────────────────────────────────────────────────

def _analyser_ia(resultats: dict, exercices: list, entreprise: str) -> str:
    try:
        from utils.ai import appel_mistral, extraire_contenu_mistral
        lignes = []
        for ex in exercices:
            r = resultats[ex]
            lignes.append(
                f"  {ex} : Activité={r['Flux activité (I)']:+,.0f}€ | "
                f"Investissement={r['Flux investissement (II)']:+,.0f}€ | "
                f"Financement={r['Flux financement (III)']:+,.0f}€ | "
                f"Tréso clôture={r['Trésorerie clôture']:,.0f}€"
            )
        prompt = f"""Tu es expert-comptable PCG France. Analyse ce TFT (méthode indirecte) pour {entreprise} :

{chr(10).join(lignes)}

Fournis :
1. Diagnostic de la santé de trésorerie
2. Qualité des flux d'activité (autofinancement)
3. Politique d'investissement (croissance ou désinvestissement)
4. Structure de financement (endettement, fonds propres)
5. Risques de liquidité identifiés
6. Recommandations concrètes

Sois concis et professionnel."""
        result = appel_mistral(prompt, temperature=0.3)
        return extraire_contenu_mistral(result) or "Analyse indisponible."
    except Exception as e:
        return f"Analyse IA indisponible : {e}"


# ─── Page principale ─────────────────────────────────────────────────────────

def page_tft():
    st.title("💹 Tableau de Flux de Trésorerie")
    st.markdown("*Méthode indirecte — ANC/CRC 99-02 — PCG France*")
    st.divider()

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        entreprise = st.text_input("Entreprise", value="Mon Entreprise")
    with col2:
        annee_ref = st.number_input("Exercice de référence", value=2024,
                                     min_value=2000, max_value=2050, step=1)
    with col3:
        nb_ex = st.slider("Exercices comparatifs", 1, 3, 2)

    exercices = [str(annee_ref - i) for i in range(nb_ex - 1, -1, -1)]
    st.caption(f"Exercices : {' | '.join(exercices)}")
    st.divider()

    # ── Initialisation données
    if "tft_data" not in st.session_state:
        st.session_state.tft_data = {}
    data = st.session_state.tft_data

    # ── Import balance optionnel
    with st.expander("📂 Importer une balance PCG pour pré-remplir"):
        col_f, col_ex = st.columns([2, 1])
        with col_f:
            fichier = st.file_uploader("Balance (.xlsx ou .csv)", type=["xlsx", "csv"], key="balance_tft")
        with col_ex:
            ex_import = st.selectbox("Pour l'exercice", exercices, key="ex_import_tft")
        if fichier and st.button("Extraire", key="btn_extract_tft"):
            with st.spinner("Extraction..."):
                vals = _extraire_depuis_balance(fichier.read(), fichier.name)
            if vals:
                for lib, val in vals.items():
                    if lib not in data:
                        data[lib] = {}
                    data[lib][ex_import] = round(val, 2)
                st.session_state.tft_data = data
                st.success(f"{len(vals)} postes extraits pour {ex_import}")
            else:
                st.warning("Extraction impossible — saisissez manuellement.")

    st.divider()

    # ── Saisie par section
    for section, info in TFT_STRUCTURE.items():
        st.markdown(f"**{section}**")
        rows = []
        for lib, signe in info["lignes"]:
            row = {"Poste": f"{lib}  ({signe})"}
            for ex in exercices:
                row[ex] = data.get(lib, {}).get(ex, 0.0)
            rows.append((lib, row))

        df_section = pd.DataFrame([r for _, r in rows])
        edited = st.data_editor(
            df_section, use_container_width=True, hide_index=True,
            column_config={ex: st.column_config.NumberColumn(ex, format="%.0f €")
                           for ex in exercices},
            key=f"tft_{section[:15]}"
        )
        for i, (lib, _) in enumerate(rows):
            if lib not in data:
                data[lib] = {}
            for ex in exercices:
                data[lib][ex] = edited.iloc[i][ex]

    # Trésorerie ouverture
    st.markdown("**Trésorerie**")
    treso_rows = [{"Poste": "Trésorerie à l'ouverture"}]
    for ex in exercices:
        treso_rows[0][ex] = data.get("Trésorerie à l'ouverture", {}).get(ex, 0.0)
    df_treso = pd.DataFrame(treso_rows)
    edited_t = st.data_editor(df_treso, use_container_width=True, hide_index=True,
                               column_config={ex: st.column_config.NumberColumn(ex, format="%.0f €")
                                              for ex in exercices},
                               key="tft_treso")
    if "Trésorerie à l'ouverture" not in data:
        data["Trésorerie à l'ouverture"] = {}
    for ex in exercices:
        data["Trésorerie à l'ouverture"][ex] = edited_t.iloc[0][ex]
    st.session_state.tft_data = data

    st.divider()

    # ── Synthèse
    resultats = _calculer_tft(data, exercices)
    st.subheader("📊 Synthèse")
    cols = st.columns(len(exercices))
    for i, ex in enumerate(exercices):
        r = resultats[ex]
        with cols[i]:
            st.metric(f"Activité {ex}", f"{r['Flux activité (I)']:+,.0f} €")
            st.metric(f"Investissement {ex}", f"{r['Flux investissement (II)']:+,.0f} €")
            st.metric(f"Financement {ex}", f"{r['Flux financement (III)']:+,.0f} €")
            color = "normal" if r["Trésorerie clôture"] >= 0 else "inverse"
            st.metric(f"Tréso clôture {ex}", f"{r['Trésorerie clôture']:,.0f} €", delta_color=color)

    st.plotly_chart(_chart_tft(resultats, exercices), use_container_width=True)
    st.divider()

    col_ia, col_xl = st.columns(2)
    with col_ia:
        if st.button("🤖 Analyse IA", type="primary", use_container_width=True):
            with st.spinner("Analyse en cours..."):
                analyse = _analyser_ia(resultats, exercices, entreprise)
            st.markdown("### 🤖 Analyse IA")
            st.markdown(analyse)
    with col_xl:
        excel_bytes = _export_excel_tft(data, resultats, exercices, entreprise)
        st.download_button(
            "📥 Exporter Excel",
            data=excel_bytes,
            file_name=f"TFT_{entreprise}_{exercices[-1]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
