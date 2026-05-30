# -*- coding: utf-8 -*-
"""
Module Plan de Financement PCG France — SMD Consulting
Saisie manuelle + Import balance | Analyse IA Mistral | Export Excel
"""
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import datetime

# ─── Catégories PCG France ───────────────────────────────────────────────────

RESSOURCES = [
    "Capacité d'autofinancement (CAF)",
    "Cessions d'éléments d'actif",
    "Augmentation de capital",
    "Subventions d'investissement reçues",
    "Nouveaux emprunts LT/MT",
    "Autres ressources durables",
]

EMPLOIS = [
    "Acquisitions d'immobilisations incorporelles",
    "Acquisitions d'immobilisations corporelles",
    "Acquisitions d'immobilisations financières",
    "Remboursements d'emprunts",
    "Distribution de dividendes",
    "Variation du besoin en fonds de roulement (BFR)",
    "Autres emplois stables",
]

# ─── Helpers ─────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def _extraire_caf_bfr_pcg(fichier_bytes: bytes, nom_fichier: str) -> dict:
    """Extrait CAF et BFR depuis une balance PCG France."""
    try:
        if nom_fichier.endswith(".xlsx"):
            df = pd.read_excel(BytesIO(fichier_bytes))
        else:
            df = pd.read_csv(BytesIO(fichier_bytes), sep=None, engine="python")

        df.columns = [str(c).strip().lower() for c in df.columns]
        col_cpte = next((c for c in df.columns if any(k in c for k in ["compte", "cpte", "n°"])), None)
        col_sol = next((c for c in df.columns if any(k in c for k in ["solde", "credit", "crédit", "montant"])), None)
        col_deb = next((c for c in df.columns if "débit" in c or "debit" in c), None)

        if not col_cpte:
            return {}

        df[col_cpte] = df[col_cpte].astype(str).str.strip()

        def somme(prefixes):
            mask = df[col_cpte].str.startswith(tuple(prefixes), na=False)
            if col_sol:
                return abs(df.loc[mask, col_sol].apply(pd.to_numeric, errors="coerce").fillna(0).sum())
            return 0.0

        resultat_net = somme(["120", "121"]) - somme(["129"])
        dotations = somme(["681", "682", "686", "687"])
        reprises = somme(["781", "786", "787"])
        caf = resultat_net + dotations - reprises

        stocks = somme(["3"])
        creances = somme(["411", "409", "413"])
        dettes_ct = somme(["401", "403", "421", "431", "437", "441", "443", "444", "445", "447"])
        bfr = stocks + creances - dettes_ct

        return {
            "CAF estimée": max(caf, 0),
            "Variation BFR estimée": abs(bfr),
        }
    except Exception:
        return {}


def _chart_plan(df_r: pd.DataFrame, df_e: pd.DataFrame, annees: list) -> go.Figure:
    total_r = [df_r[a].sum() for a in annees]
    total_e = [df_e[a].sum() for a in annees]
    solde = [r - e for r, e in zip(total_r, total_e)]

    fig = go.Figure()
    fig.add_trace(go.Bar(name="Ressources", x=annees, y=total_r,
                         marker_color="#1E8449",
                         hovertemplate="%{y:,.0f} €<extra>Ressources</extra>"))
    fig.add_trace(go.Bar(name="Emplois", x=annees, y=total_e,
                         marker_color="#C0392B",
                         hovertemplate="%{y:,.0f} €<extra>Emplois</extra>"))
    fig.add_trace(go.Scatter(name="Solde", x=annees, y=solde,
                             mode="lines+markers",
                             line=dict(color="#2C3E50", width=2.5, dash="dot"),
                             marker=dict(size=9),
                             hovertemplate="%{y:,.0f} €<extra>Solde</extra>"))
    fig.add_hline(y=0, line_dash="dash", line_color="#888", line_width=1)
    fig.update_layout(
        title="Plan de financement — Ressources vs Emplois",
        barmode="group", height=420,
        margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", y=-0.15),
        yaxis_tickformat=",.0f",
    )
    return fig


def _export_excel(df_r: pd.DataFrame, df_e: pd.DataFrame, annees: list, entreprise: str) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        def _style(ws, header_color):
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=header_color)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 28

        # Ressources
        df_r_exp = df_r.copy().set_index("Ressource")
        df_r_exp.loc["TOTAL RESSOURCES"] = df_r_exp[annees].sum()
        df_r_exp.reset_index().to_excel(writer, sheet_name="Ressources", index=False)
        _style(writer.sheets["Ressources"], "1E8449")

        # Emplois
        df_e_exp = df_e.copy().set_index("Emploi")
        df_e_exp.loc["TOTAL EMPLOIS"] = df_e_exp[annees].sum()
        df_e_exp.reset_index().to_excel(writer, sheet_name="Emplois", index=False)
        _style(writer.sheets["Emplois"], "C0392B")

        # Synthèse
        synth = pd.DataFrame({
            "Année": annees,
            "Total Ressources (€)": [df_r[a].sum() for a in annees],
            "Total Emplois (€)": [df_e[a].sum() for a in annees],
            "Solde (€)": [df_r[a].sum() - df_e[a].sum() for a in annees],
        })
        synth.to_excel(writer, sheet_name="Synthèse", index=False)
        _style(writer.sheets["Synthèse"], "2C3E50")

    return buf.getvalue()


def _analyser_ia(df_r: pd.DataFrame, df_e: pd.DataFrame, annees: list, entreprise: str) -> str:
    try:
        from utils.ai import appel_mistral
        lignes = []
        for a in annees:
            r = df_r[a].sum()
            e = df_e[a].sum()
            lignes.append(f"  {a} : Ressources={r:,.0f}€ | Emplois={e:,.0f}€ | Solde={r-e:,.0f}€")
        resume = "\n".join(lignes)

        prompt = f"""Tu es expert-comptable PCG France. Analyse ce plan de financement pour {entreprise} :

{resume}

Fournis :
1. Diagnostic de l'équilibre financier (ressources/emplois)
2. Risques identifiés (sous-financement, endettement, BFR)
3. Points forts du plan
4. Recommandations concrètes (refinancement, optimisation BFR, fonds propres)
5. Conformité avec les bonnes pratiques PCG

Sois concis et professionnel."""

        result = appel_mistral(prompt, temperature=0.3)
        from utils.ai import extraire_contenu_mistral
        return extraire_contenu_mistral(result) or "Analyse indisponible."
    except Exception as e:
        return f"Analyse IA indisponible : {e}"


# ─── Page principale ─────────────────────────────────────────────────────────

def page_plan_financement():
    st.title("📐 Plan de Financement")
    st.markdown("*PCG France — Équilibre ressources / emplois sur 1 à 5 ans*")
    st.divider()

    # ── Paramètres
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        entreprise = st.text_input("Entreprise", value="Mon Entreprise")
    with col2:
        annee_debut = st.number_input("Année de départ", value=datetime.now().year,
                                       min_value=2000, max_value=2050, step=1)
    with col3:
        nb_annees = st.slider("Nombre d'années", 1, 5, 3)

    annees = [str(annee_debut + i) for i in range(nb_annees)]
    st.caption(f"Période : **{annees[0]}** → **{annees[-1]}**")
    st.divider()

    # ── Import balance (optionnel)
    prefill_r: dict = {}
    prefill_e: dict = {}

    with st.expander("📂 Importer une balance PCG pour pré-remplir CAF et BFR"):
        fichier = st.file_uploader("Balance (.xlsx ou .csv)", type=["xlsx", "csv"],
                                    key="balance_plan")
        if fichier:
            with st.spinner("Extraction en cours..."):
                vals = _extraire_caf_bfr_pcg(fichier.read(), fichier.name)
            if vals:
                prefill_r["Capacité d'autofinancement (CAF)"] = vals.get("CAF estimée", 0)
                prefill_e["Variation du besoin en fonds de roulement (BFR)"] = vals.get("Variation BFR estimée", 0)
                caf_v = prefill_r["Capacité d'autofinancement (CAF)"]
                bfr_v = prefill_e["Variation du besoin en fonds de roulement (BFR)"]
                st.success(f"CAF estimée : {caf_v:,.0f} € | BFR estimé : {bfr_v:,.0f} €")
            else:
                st.warning("Extraction impossible — saisissez les valeurs manuellement.")

    st.divider()

    # ── Saisie Ressources
    st.subheader("📥 Ressources")
    r_data = {"Ressource": RESSOURCES}
    for a in annees:
        r_data[a] = [prefill_r.get(lib, 0.0) for lib in RESSOURCES]
    df_r = st.data_editor(
        pd.DataFrame(r_data),
        use_container_width=True,
        hide_index=True,
        column_config={a: st.column_config.NumberColumn(a, format="%.0f €", min_value=0)
                       for a in annees},
        key="editor_ressources",
    )

    st.divider()

    # ── Saisie Emplois
    st.subheader("📤 Emplois")
    e_data = {"Emploi": EMPLOIS}
    for a in annees:
        e_data[a] = [prefill_e.get(lib, 0.0) for lib in EMPLOIS]
    df_e = st.data_editor(
        pd.DataFrame(e_data),
        use_container_width=True,
        hide_index=True,
        column_config={a: st.column_config.NumberColumn(a, format="%.0f €", min_value=0)
                       for a in annees},
        key="editor_emplois",
    )

    st.divider()

    # ── Synthèse automatique
    st.subheader("📊 Synthèse")
    cols = st.columns(len(annees))
    for i, a in enumerate(annees):
        total_r = df_r[a].sum()
        total_e = df_e[a].sum()
        solde = total_r - total_e
        with cols[i]:
            st.metric(f"Ressources {a}", f"{total_r:,.0f} €")
            st.metric(f"Emplois {a}", f"{total_e:,.0f} €")
            delta_color = "normal" if solde >= 0 else "inverse"
            st.metric(f"Solde {a}", f"{solde:,.0f} €",
                      delta=f"{'Excédent' if solde >= 0 else 'Déficit'}",
                      delta_color=delta_color)

    st.plotly_chart(_chart_plan(df_r, df_e, annees), use_container_width=True)
    st.divider()

    # ── Actions
    col_ia, col_xl = st.columns(2)

    with col_ia:
        if st.button("🤖 Analyse IA du plan", type="primary", use_container_width=True):
            with st.spinner("Analyse en cours..."):
                analyse = _analyser_ia(df_r, df_e, annees, entreprise)
            st.markdown("### 🤖 Analyse IA")
            st.markdown(analyse)

    with col_xl:
        excel_bytes = _export_excel(df_r, df_e, annees, entreprise)
        st.download_button(
            "📥 Exporter Excel",
            data=excel_bytes,
            file_name=f"Plan_Financement_{entreprise}_{annees[0]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
