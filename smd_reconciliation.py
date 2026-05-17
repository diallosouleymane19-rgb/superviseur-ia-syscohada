# -*- coding: utf-8 -*-
"""
Module Rapprochement Bancaire - SMD Consulting
Comparaison Releve Bancaire vs Grand Livre Banque (SYSCOHADA)
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import date, timedelta

# =============================================================================
# CONSTANTES
# =============================================================================

TOLERANCE_JOURS = 5   # ecart de date acceptable pour matching automatique
TOLERANCE_MONTANT = 1  # ecart montant acceptable (arrondis)


# =============================================================================
# UTILITAIRES LECTURE FICHIER
# =============================================================================

def _lire_fichier(fichier) -> pd.DataFrame:
    """Lit un fichier Excel ou CSV et retourne un DataFrame."""
    try:
        nom = fichier.name.lower()
        if nom.endswith(".xlsx") or nom.endswith(".xls"):
            df = pd.read_excel(fichier)
        else:
            try:
                df = pd.read_csv(fichier, encoding="utf-8", sep=None, engine="python")
            except Exception:
                fichier.seek(0)
                df = pd.read_csv(fichier, encoding="latin-1", sep=None, engine="python")
        fichier.seek(0)
        return df, None
    except Exception as e:
        return None, str(e)


def _detecter_cols(df: pd.DataFrame) -> dict:
    """Detection automatique des colonnes Date / Libelle / Debit / Credit / Montant."""
    cols = {str(c).lower().strip(): c for c in df.columns}
    m = {}

    for k in ["date", "date operation", "date valeur", "date mvt", "date ecriture", "dt"]:
        if k in cols:
            m["date"] = cols[k]
            break

    for k in ["libelle", "libelle operation", "description", "motif", "nature", "intitule", "label"]:
        if k in cols:
            m["libelle"] = cols[k]
            break

    for k in ["debit", "debit (fcfa)", "montant debit", "sorties", "charges"]:
        if k in cols:
            m["debit"] = cols[k]
            break

    for k in ["credit", "credit (fcfa)", "montant credit", "entrees", "produits"]:
        if k in cols:
            m["credit"] = cols[k]
            break

    for k in ["montant", "solde", "amount", "valeur", "montant net"]:
        if k in cols:
            m["montant"] = cols[k]
            break

    for k in ["reference", "ref", "numero", "piece", "numero piece", "n piece"]:
        if k in cols:
            m["reference"] = cols[k]
            break

    return m


# =============================================================================
# NORMALISATION DU DATAFRAME
# =============================================================================

def _normaliser(df: pd.DataFrame, mapping: dict, source: str) -> pd.DataFrame:
    """
    Normalise un DataFrame brut en colonnes standard :
    Date, Libelle, Montant (positif = entree, negatif = sortie), Reference, Source
    """
    result = pd.DataFrame()

    # Date
    if "date" in mapping:
        result["Date"] = pd.to_datetime(df[mapping["date"]], dayfirst=True, errors="coerce")
    else:
        result["Date"] = pd.NaT

    # Libelle
    if "libelle" in mapping:
        result["Libelle"] = df[mapping["libelle"]].astype(str)
    else:
        result["Libelle"] = ""

    # Reference
    if "reference" in mapping:
        result["Reference"] = df[mapping["reference"]].astype(str)
    else:
        result["Reference"] = ""

    # Montant unifie
    if "montant" in mapping:
        result["Montant"] = pd.to_numeric(df[mapping["montant"]], errors="coerce").fillna(0)
    elif "debit" in mapping and "credit" in mapping:
        deb = pd.to_numeric(df[mapping["debit"]], errors="coerce").fillna(0)
        cre = pd.to_numeric(df[mapping["credit"]], errors="coerce").fillna(0)
        result["Montant"] = cre - deb   # convention : entrees positives, sorties negatives
    elif "debit" in mapping:
        result["Montant"] = -pd.to_numeric(df[mapping["debit"]], errors="coerce").fillna(0)
    elif "credit" in mapping:
        result["Montant"] = pd.to_numeric(df[mapping["credit"]], errors="coerce").fillna(0)
    else:
        result["Montant"] = 0.0

    result["Source"] = source
    result["_idx"] = range(len(result))

    # Supprimer lignes sans date ni montant
    result = result.dropna(subset=["Date"]).reset_index(drop=True)
    result["_idx"] = range(len(result))

    return result


# =============================================================================
# ALGORITHME DE RAPPROCHEMENT
# =============================================================================

def _rapprocher(df_banque: pd.DataFrame, df_gl: pd.DataFrame,
                tolerance_jours: int = TOLERANCE_JOURS) -> dict:
    """
    Matching entre releve bancaire et grand livre.
    Critere : meme montant (+-1 FCFA) + date proche (+-N jours).
    Retourne : rapproches, non_rapproches_banque, non_rapproches_gl
    """
    banque = df_banque.copy().reset_index(drop=True)
    gl     = df_gl.copy().reset_index(drop=True)

    matched_banque = set()
    matched_gl     = set()
    paires = []

    for i, row_b in banque.iterrows():
        if pd.isna(row_b["Date"]):
            continue
        mnt_b = row_b["Montant"]
        date_b = row_b["Date"]

        for j, row_g in gl.iterrows():
            if j in matched_gl:
                continue
            if pd.isna(row_g["Date"]):
                continue

            mnt_g = row_g["Montant"]
            date_g = row_g["Date"]

            ecart_mnt  = abs(mnt_b - mnt_g)
            ecart_date = abs((date_b - date_g).days)

            if ecart_mnt <= TOLERANCE_MONTANT and ecart_date <= tolerance_jours:
                paires.append({
                    "Date Banque":   date_b.strftime("%d/%m/%Y"),
                    "Date GL":       date_g.strftime("%d/%m/%Y"),
                    "Libelle Banque": row_b["Libelle"],
                    "Libelle GL":    row_g["Libelle"],
                    "Montant":       mnt_b,
                    "Ref Banque":    row_b["Reference"],
                    "Ref GL":        row_g["Reference"],
                    "Ecart Date (j)": ecart_date,
                })
                matched_banque.add(i)
                matched_gl.add(j)
                break

    non_rap_banque = banque[~banque.index.isin(matched_banque)].copy()
    non_rap_gl     = gl[~gl.index.isin(matched_gl)].copy()

    df_paires = pd.DataFrame(paires)

    return {
        "rapproches":       df_paires,
        "ecarts_banque":    non_rap_banque,
        "ecarts_gl":        non_rap_gl,
        "nb_rapproches":    len(paires),
        "nb_ecarts_banque": len(non_rap_banque),
        "nb_ecarts_gl":     len(non_rap_gl),
    }


# =============================================================================
# GRAPHIQUE SYNTHESE
# =============================================================================

def _chart_synthese(solde_banque: float, solde_gl: float,
                    total_ecarts_banque: float, total_ecarts_gl: float) -> go.Figure:
    """Graphique waterfall : GL → ajustements → Banque."""
    labels = [
        "Solde Grand Livre",
        "Ecarts GL non rapproches",
        "Ecarts Banque non rapproches",
        "Solde Releve Bancaire"
    ]
    values = [
        solde_gl,
        -total_ecarts_gl,
        total_ecarts_banque,
        0
    ]
    measures = ["absolute", "relative", "relative", "total"]

    fig = go.Figure(go.Waterfall(
        name="Rapprochement",
        orientation="v",
        measure=measures,
        x=labels,
        y=values,
        textposition="outside",
        text=[f"{v:,.0f}" for v in values],
        connector={"line": {"color": "#888"}},
        increasing={"marker": {"color": "#1E8449"}},
        decreasing={"marker": {"color": "#C0392B"}},
        totals={"marker": {"color": "#1f77b4"}},
    ))
    fig.update_layout(
        title="Passerelle Grand Livre → Relevé Bancaire",
        height=380,
        margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        yaxis_title="Montant (FCFA)",
    )
    return fig


def _chart_statut(nb_rap: int, nb_b: int, nb_gl: int) -> go.Figure:
    """Camembert des operations par statut."""
    labels = ["Rapprochees", "Ecarts Banque", "Ecarts GL"]
    values = [nb_rap, nb_b, nb_gl]
    colors = ["#1E8449", "#E67E22", "#C0392B"]

    fig = go.Figure(go.Pie(
        labels=labels,
        values=values,
        hole=0.45,
        marker_colors=colors,
        textinfo="label+percent",
    ))
    fig.update_layout(
        title="Repartition des operations",
        height=320,
        margin=dict(l=10, r=10, t=50, b=10),
        paper_bgcolor="rgba(0,0,0,0)",
        showlegend=False,
    )
    return fig


# =============================================================================
# EXPORT EXCEL
# =============================================================================

def _export_rapprochement_excel(result: dict, solde_banque: float,
                                 solde_gl: float, date_arrete: date,
                                 devise: str) -> BytesIO:
    """Exporte le rapprochement bancaire en Excel multi-onglets."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ── Onglet 1 : Synthese ──────────────────────────────────────
        ecart_global = solde_banque - solde_gl
        nb_rap = result["nb_rapproches"]
        nb_b   = result["nb_ecarts_banque"]
        nb_gl  = result["nb_ecarts_gl"]
        total_ecarts_b  = result["ecarts_banque"]["Montant"].sum() if not result["ecarts_banque"].empty else 0
        total_ecarts_gl = result["ecarts_gl"]["Montant"].sum()     if not result["ecarts_gl"].empty else 0

        synth = pd.DataFrame([
            ["Date d'arrete", date_arrete.strftime("%d/%m/%Y")],
            ["Solde Grand Livre (GL)", f"{solde_gl:,.0f} {devise}"],
            ["Solde Releve Bancaire",  f"{solde_banque:,.0f} {devise}"],
            ["Ecart Global",           f"{ecart_global:,.0f} {devise}"],
            ["", ""],
            ["Operations rapprochees", nb_rap],
            ["Ecarts cote Banque",     nb_b],
            ["Ecarts cote GL",         nb_gl],
            ["", ""],
            ["Total ecarts Banque",    f"{total_ecarts_b:,.0f} {devise}"],
            ["Total ecarts GL",        f"{total_ecarts_gl:,.0f} {devise}"],
        ], columns=["Indicateur", "Valeur"])
        synth.to_excel(writer, index=False, sheet_name="Synthese")

        ws = writer.sheets["Synthese"]
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 30
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # ── Onglet 2 : Operations Rapprochees ───────────────────────
        if not result["rapproches"].empty:
            result["rapproches"].to_excel(writer, index=False, sheet_name="Rapprochees")
            ws2 = writer.sheets["Rapprochees"]
            fill_ok = PatternFill("solid", fgColor="D4EDDA")
            for row in ws2.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = fill_ok
            for cell in ws2[1]:
                cell.fill = PatternFill("solid", fgColor="1E8449")
                cell.font = Font(color="FFFFFF", bold=True)
            for col in ws2.columns:
                ws2.column_dimensions[get_column_letter(col[0].column)].width = 22

        # ── Onglet 3 : Ecarts Banque ─────────────────────────────────
        cols_display = ["Date", "Libelle", "Montant", "Reference"]
        if not result["ecarts_banque"].empty:
            df_eb = result["ecarts_banque"][[c for c in cols_display if c in result["ecarts_banque"].columns]]
            df_eb.to_excel(writer, index=False, sheet_name="Ecarts Banque")
            ws3 = writer.sheets["Ecarts Banque"]
            fill_b = PatternFill("solid", fgColor="FFF3CD")
            for row in ws3.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = fill_b
            for cell in ws3[1]:
                cell.fill = PatternFill("solid", fgColor="E67E22")
                cell.font = Font(color="FFFFFF", bold=True)
            for col in ws3.columns:
                ws3.column_dimensions[get_column_letter(col[0].column)].width = 28

        # ── Onglet 4 : Ecarts GL ─────────────────────────────────────
        if not result["ecarts_gl"].empty:
            df_eg = result["ecarts_gl"][[c for c in cols_display if c in result["ecarts_gl"].columns]]
            df_eg.to_excel(writer, index=False, sheet_name="Ecarts GL")
            ws4 = writer.sheets["Ecarts GL"]
            fill_gl = PatternFill("solid", fgColor="F5C6CB")
            for row in ws4.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = fill_gl
            for cell in ws4[1]:
                cell.fill = PatternFill("solid", fgColor="C0392B")
                cell.font = Font(color="FFFFFF", bold=True)
            for col in ws4.columns:
                ws4.column_dimensions[get_column_letter(col[0].column)].width = 28

    buf.seek(0)
    return buf


# =============================================================================
# WIDGETS MAPPING COLONNES
# =============================================================================

def _widget_mapping(df: pd.DataFrame, mapping: dict, prefix: str) -> dict:
    """Affiche les selectboxes de mapping colonnes pour un fichier."""
    detection_ok = ("date" in mapping) and (("montant" in mapping) or ("debit" in mapping) or ("credit" in mapping))

    if not detection_ok:
        st.warning("Colonnes non detectees. Selectionnez-les manuellement ci-dessous.")

    with st.expander("Correspondance des colonnes", expanded=not detection_ok):
        st.markdown("**Colonnes disponibles :** " + ", ".join([f"`{c}`" for c in df.columns]))
        st.markdown("---")
        all_cols = ["-- Non disponible --"] + list(df.columns)

        col_date = st.selectbox(
            f"Colonne Date *",
            all_cols,
            index=all_cols.index(mapping["date"]) if "date" in mapping and mapping["date"] in all_cols else 0,
            key=f"{prefix}_date"
        )
        col_lib = st.selectbox(
            f"Colonne Libelle",
            all_cols,
            index=all_cols.index(mapping["libelle"]) if "libelle" in mapping and mapping["libelle"] in all_cols else 0,
            key=f"{prefix}_lib"
        )

        mode = st.radio(
            "Format du montant",
            ["Colonne unique Montant", "Deux colonnes Debit / Credit"],
            index=1 if ("debit" in mapping or "credit" in mapping) else 0,
            key=f"{prefix}_mode",
            horizontal=True
        )

        if mode == "Colonne unique Montant":
            col_mnt = st.selectbox(
                "Colonne Montant *",
                all_cols,
                index=all_cols.index(mapping["montant"]) if "montant" in mapping and mapping["montant"] in all_cols else 0,
                key=f"{prefix}_mnt"
            )
            col_deb, col_cre = None, None
        else:
            col_deb = st.selectbox(
                "Colonne Debit (sorties) *",
                all_cols,
                index=all_cols.index(mapping["debit"]) if "debit" in mapping and mapping["debit"] in all_cols else 0,
                key=f"{prefix}_deb"
            )
            col_cre = st.selectbox(
                "Colonne Credit (entrees) *",
                all_cols,
                index=all_cols.index(mapping["credit"]) if "credit" in mapping and mapping["credit"] in all_cols else 0,
                key=f"{prefix}_cre"
            )
            col_mnt = None

        col_ref = st.selectbox(
            "Colonne Reference (optionnel)",
            all_cols,
            index=all_cols.index(mapping["reference"]) if "reference" in mapping and mapping["reference"] in all_cols else 0,
            key=f"{prefix}_ref"
        )

    final = {}
    if col_date != "-- Non disponible --":
        final["date"] = col_date
    if col_lib != "-- Non disponible --":
        final["libelle"] = col_lib
    if col_mnt and col_mnt != "-- Non disponible --":
        final["montant"] = col_mnt
    if col_deb and col_deb != "-- Non disponible --":
        final["debit"] = col_deb
    if col_cre and col_cre != "-- Non disponible --":
        final["credit"] = col_cre
    if col_ref != "-- Non disponible --":
        final["reference"] = col_ref

    return final


# =============================================================================
# PAGE PRINCIPALE
# =============================================================================

def page_rapprochement_bancaire():
    st.title("🏦 Rapprochement Bancaire")
    st.markdown("*Comparaison Relevé Bancaire vs Grand Livre Compte Banque — SYSCOHADA*")
    st.divider()

    # ── Parametres ───────────────────────────────────────────────────
    col_p1, col_p2, col_p3 = st.columns(3)
    with col_p1:
        date_arrete = st.date_input("Date d'arrete", value=date.today())
    with col_p2:
        devise = st.selectbox("Devise", ["FCFA", "XOF", "XAF", "EUR", "USD"], index=0)
    with col_p3:
        tolerance = st.number_input(
            "Tolerance date matching (jours)",
            min_value=0, max_value=15, value=TOLERANCE_JOURS,
            help="Ecart de date maximal accepte pour rapprocher deux operations"
        )

    st.divider()

    # ── Upload des deux fichiers ──────────────────────────────────────
    col_f1, col_f2 = st.columns(2)

    with col_f1:
        st.subheader("🏦 Relevé Bancaire")
        st.caption("Export de votre banque (Excel ou CSV)")
        fichier_banque = st.file_uploader(
            "Importer le releve bancaire",
            type=["xlsx", "xls", "csv"],
            key="upload_banque"
        )

    with col_f2:
        st.subheader("📒 Grand Livre Banque")
        st.caption("Compte 521 ou equivalent SYSCOHADA")
        fichier_gl = st.file_uploader(
            "Importer le grand livre",
            type=["xlsx", "xls", "csv"],
            key="upload_gl"
        )

    if not fichier_banque or not fichier_gl:
        st.info("Importez les deux fichiers pour demarrer le rapprochement.")

        # Exemple de format attendu
        with st.expander("Format attendu des fichiers"):
            ex_col1, ex_col2 = st.columns(2)
            with ex_col1:
                st.markdown("**Releve Bancaire** *(minimum)*")
                st.dataframe(pd.DataFrame({
                    "Date":    ["01/05/2026", "03/05/2026", "10/05/2026"],
                    "Libelle": ["VIR SOTRACOM", "FRAIS TENUE COMPTE", "REMISE CHEQUE"],
                    "Debit":   [0, 15000, 0],
                    "Credit":  [4850000, 0, 2300000],
                }), use_container_width=True, hide_index=True)
            with ex_col2:
                st.markdown("**Grand Livre Banque** *(minimum)*")
                st.dataframe(pd.DataFrame({
                    "Date":    ["01/05/2026", "03/05/2026", "10/05/2026"],
                    "Libelle": ["Reglt client SOTRACOM", "Frais bancaires", "Remise cheque MBAYE"],
                    "Debit":   [0, 15000, 0],
                    "Credit":  [4850000, 0, 2300000],
                }), use_container_width=True, hide_index=True)
        return

    # ── Lecture des fichiers ──────────────────────────────────────────
    df_b_raw, err_b = _lire_fichier(fichier_banque)
    df_g_raw, err_g = _lire_fichier(fichier_gl)

    if err_b:
        st.error("Erreur lecture releve bancaire : " + err_b)
        return
    if err_g:
        st.error("Erreur lecture grand livre : " + err_g)
        return

    # ── Apercu ───────────────────────────────────────────────────────
    col_a1, col_a2 = st.columns(2)
    with col_a1:
        with st.expander(f"Apercu Releve Bancaire ({len(df_b_raw)} lignes)"):
            st.dataframe(df_b_raw.head(5), use_container_width=True, hide_index=True)
    with col_a2:
        with st.expander(f"Apercu Grand Livre ({len(df_g_raw)} lignes)"):
            st.dataframe(df_g_raw.head(5), use_container_width=True, hide_index=True)

    # ── Mapping colonnes ─────────────────────────────────────────────
    st.subheader("1️⃣ Correspondance des colonnes")
    col_m1, col_m2 = st.columns(2)

    with col_m1:
        st.markdown("**Relevé Bancaire**")
        map_b = _detecter_cols(df_b_raw)
        map_b = _widget_mapping(df_b_raw, map_b, "banque")

    with col_m2:
        st.markdown("**Grand Livre Banque**")
        map_g = _detecter_cols(df_g_raw)
        map_g = _widget_mapping(df_g_raw, map_g, "gl")

    # ── Soldes initiaux ──────────────────────────────────────────────
    st.divider()
    st.subheader("2️⃣ Soldes a la date d'arrete")
    col_s1, col_s2 = st.columns(2)
    with col_s1:
        solde_banque = st.number_input(
            "Solde Releve Bancaire au " + date_arrete.strftime("%d/%m/%Y"),
            value=0.0, step=1000.0, format="%.0f"
        )
    with col_s2:
        solde_gl = st.number_input(
            "Solde Grand Livre (cpte 521) au " + date_arrete.strftime("%d/%m/%Y"),
            value=0.0, step=1000.0, format="%.0f"
        )

    # ── Lancement ────────────────────────────────────────────────────
    st.divider()
    if st.button("🔍 Lancer le rapprochement", type="primary", use_container_width=True):

        # Validation mapping
        has_date_b = "date" in map_b
        has_mnt_b  = ("montant" in map_b) or ("debit" in map_b) or ("credit" in map_b)
        has_date_g = "date" in map_g
        has_mnt_g  = ("montant" in map_g) or ("debit" in map_g) or ("credit" in map_g)

        if not (has_date_b and has_mnt_b):
            st.error("Releve Bancaire : les colonnes Date et Montant/Debit/Credit sont obligatoires.")
            return
        if not (has_date_g and has_mnt_g):
            st.error("Grand Livre : les colonnes Date et Montant/Debit/Credit sont obligatoires.")
            return

        with st.spinner("Rapprochement en cours..."):
            df_banque = _normaliser(df_b_raw, map_b, "Banque")
            df_gl_n   = _normaliser(df_g_raw, map_g, "GL")

            if df_banque.empty:
                st.error("Aucune ligne valide dans le releve bancaire. Verifiez le format.")
                return
            if df_gl_n.empty:
                st.error("Aucune ligne valide dans le grand livre. Verifiez le format.")
                return

            result = _rapprocher(df_banque, df_gl_n, tolerance_jours=int(tolerance))

        # ── KPIs ─────────────────────────────────────────────────────
        ecart_global = solde_banque - solde_gl
        pct_rapproche = (result["nb_rapproches"] / max(len(df_banque), 1)) * 100

        st.divider()
        st.subheader("3️⃣ Resultats du rapprochement")

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Solde Banque",    f"{solde_banque:,.0f}", devise)
        k2.metric("Solde GL",        f"{solde_gl:,.0f}",     devise)
        k3.metric("Ecart Global",    f"{ecart_global:,.0f}", devise,
                  delta_color="off" if ecart_global == 0 else "inverse")
        k4.metric("Operations rapprochees", result["nb_rapproches"],
                  delta=f"{pct_rapproche:.0f}%")
        k5.metric("Operations en ecart",
                  result["nb_ecarts_banque"] + result["nb_ecarts_gl"],
                  delta_color="inverse")

        # Statut global
        if ecart_global == 0 and result["nb_ecarts_banque"] == 0 and result["nb_ecarts_gl"] == 0:
            st.success("✅ Rapprochement PARFAIT — Soldes identiques, aucun ecart.")
        elif ecart_global == 0:
            st.warning(f"⚠️ Soldes equilibres mais {result['nb_ecarts_banque'] + result['nb_ecarts_gl']} operation(s) non rapprochee(s).")
        else:
            st.error(f"🔴 Ecart de {ecart_global:,.0f} {devise} entre la banque et le GL.")

        # ── Graphiques ───────────────────────────────────────────────
        st.divider()
        gc1, gc2 = st.columns([3, 2])
        total_ecarts_b  = result["ecarts_banque"]["Montant"].sum() if not result["ecarts_banque"].empty else 0
        total_ecarts_gl = result["ecarts_gl"]["Montant"].sum()     if not result["ecarts_gl"].empty else 0

        with gc1:
            st.plotly_chart(_chart_synthese(solde_banque, solde_gl, total_ecarts_b, total_ecarts_gl),
                            use_container_width=True)
        with gc2:
            st.plotly_chart(_chart_statut(result["nb_rapproches"],
                                          result["nb_ecarts_banque"],
                                          result["nb_ecarts_gl"]),
                            use_container_width=True)

        # ── Operations rapprochees ───────────────────────────────────
        st.divider()
        st.subheader(f"✅ Operations rapprochees ({result['nb_rapproches']})")
        if result["rapproches"].empty:
            st.info("Aucune operation rapprochee. Verifiez la tolerance de date ou le format des montants.")
        else:
            st.dataframe(
                result["rapproches"].style.applymap(
                    lambda v: "color: #1E8449; font-weight:bold" if isinstance(v, (int, float)) and v > 0 else "",
                    subset=["Montant"]
                ),
                use_container_width=True,
                hide_index=True
            )

        # ── Ecarts Banque ────────────────────────────────────────────
        st.divider()
        st.subheader(f"🟠 Ecarts cote Banque ({result['nb_ecarts_banque']}) — Operations au releve non comptabilisees")
        if result["ecarts_banque"].empty:
            st.success("Aucun ecart cote banque.")
        else:
            cols_show = ["Date", "Libelle", "Montant", "Reference"]
            df_show_b = result["ecarts_banque"][[c for c in cols_show if c in result["ecarts_banque"].columns]]
            df_show_b = df_show_b.copy()
            if "Date" in df_show_b.columns:
                df_show_b["Date"] = df_show_b["Date"].dt.strftime("%d/%m/%Y")
            st.dataframe(df_show_b, use_container_width=True, hide_index=True)
            st.caption(f"Total ecarts banque : **{total_ecarts_b:,.0f} {devise}**")

        # ── Ecarts GL ────────────────────────────────────────────────
        st.divider()
        st.subheader(f"🔴 Ecarts cote GL ({result['nb_ecarts_gl']}) — Ecritures GL sans contrepartie bancaire")
        if result["ecarts_gl"].empty:
            st.success("Aucun ecart cote GL.")
        else:
            cols_show = ["Date", "Libelle", "Montant", "Reference"]
            df_show_g = result["ecarts_gl"][[c for c in cols_show if c in result["ecarts_gl"].columns]]
            df_show_g = df_show_g.copy()
            if "Date" in df_show_g.columns:
                df_show_g["Date"] = df_show_g["Date"].dt.strftime("%d/%m/%Y")
            st.dataframe(df_show_g, use_container_width=True, hide_index=True)
            st.caption(f"Total ecarts GL : **{total_ecarts_gl:,.0f} {devise}**")

        # ── Export ───────────────────────────────────────────────────
        st.divider()
        excel_buf = _export_rapprochement_excel(
            result, solde_banque, solde_gl, date_arrete, devise
        )
        st.download_button(
            label="📥 Telecharger le rapprochement Excel",
            data=excel_buf,
            file_name=f"Rapprochement_Bancaire_{date_arrete.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
