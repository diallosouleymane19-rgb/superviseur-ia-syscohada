# -*- coding: utf-8 -*-
"""
Module TFT SYSCOHADA — SMD Consulting
Tableau de Flux de Trésorerie — Format SYSCOHADA révisé 2017
Horizon 1 à 3 exercices comparatifs
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO

# ─── Structure TFT SYSCOHADA révisé 2017 ────────────────────────────────────

TFT_STRUCTURE = {
    "I. FLUX DE TRÉSORERIE LIÉS AUX ACTIVITÉS OPÉRATIONNELLES": {
        "color": "#1E8449",
        "lignes": [
            ("Résultat net de l'exercice", "+"),
            ("Dotations aux amortissements et provisions", "+"),
            ("Reprises sur amortissements et provisions", "-"),
            ("Plus-values de cessions (nettes d'impôts)", "-"),
            ("Moins-values de cessions", "+"),
            ("Variation des stocks (augmentation -)", "±"),
            ("Variation des créances clients (augmentation -)", "±"),
            ("Variation des dettes fournisseurs (augmentation +)", "±"),
            ("Variation des autres créances d'exploitation (augmentation -)", "±"),
            ("Variation des autres dettes d'exploitation (augmentation +)", "±"),
            ("Impôts sur les bénéfices payés", "-"),
        ]
    },
    "II. FLUX DE TRÉSORERIE LIÉS AUX ACTIVITÉS D'INVESTISSEMENT": {
        "color": "#2980B9",
        "lignes": [
            ("Acquisitions d'immobilisations incorporelles", "-"),
            ("Acquisitions d'immobilisations corporelles", "-"),
            ("Acquisitions d'immobilisations financières", "-"),
            ("Cessions d'immobilisations incorporelles", "+"),
            ("Cessions d'immobilisations corporelles", "+"),
            ("Cessions d'immobilisations financières", "+"),
            ("Variation des autres actifs non courants", "±"),
        ]
    },
    "III. FLUX DE TRÉSORERIE LIÉS AUX ACTIVITÉS DE FINANCEMENT": {
        "color": "#8E44AD",
        "lignes": [
            ("Augmentation de capital (en numéraire)", "+"),
            ("Subventions d'investissement reçues", "+"),
            ("Remboursements de capital", "-"),
            ("Emprunts à long et moyen terme souscrits", "+"),
            ("Remboursements d'emprunts LMT", "-"),
            ("Dividendes versés aux actionnaires", "-"),
            ("Variation des concours bancaires courants", "±"),
        ]
    }
}

TOTAUX = [
    "FLUX NET DE TRÉSORERIE — ACTIVITÉS OPÉRATIONNELLES (I)",
    "FLUX NET DE TRÉSORERIE — ACTIVITÉS D'INVESTISSEMENT (II)",
    "FLUX NET DE TRÉSORERIE — ACTIVITÉS DE FINANCEMENT (III)",
    "VARIATION NETTE DE TRÉSORERIE (I + II + III)",
    "Trésorerie à l'ouverture de l'exercice",
    "TRÉSORERIE À LA CLÔTURE DE L'EXERCICE",
]


# ─── Helpers ────────────────────────────────────────────────────────────────

def _calculer_tft(data, exercices):
    """Calcule les totaux de chaque section et la trésorerie finale."""
    resultats = {}
    for ex in exercices:
        flux_op  = sum(data.get(lib, {}).get(ex, 0) for lib, _ in TFT_STRUCTURE["I. FLUX DE TRÉSORERIE LIÉS AUX ACTIVITÉS OPÉRATIONNELLES"]["lignes"])
        flux_inv = sum(data.get(lib, {}).get(ex, 0) for lib, _ in TFT_STRUCTURE["II. FLUX DE TRÉSORERIE LIÉS AUX ACTIVITÉS D'INVESTISSEMENT"]["lignes"])
        flux_fin = sum(data.get(lib, {}).get(ex, 0) for lib, _ in TFT_STRUCTURE["III. FLUX DE TRÉSORERIE LIÉS AUX ACTIVITÉS DE FINANCEMENT"]["lignes"])
        var_nette = flux_op + flux_inv + flux_fin
        treso_ouv = data.get("Trésorerie à l'ouverture de l'exercice", {}).get(ex, 0)
        treso_clo = treso_ouv + var_nette

        resultats[ex] = {
            "FLUX NET DE TRÉSORERIE — ACTIVITÉS OPÉRATIONNELLES (I)": flux_op,
            "FLUX NET DE TRÉSORERIE — ACTIVITÉS D'INVESTISSEMENT (II)": flux_inv,
            "FLUX NET DE TRÉSORERIE — ACTIVITÉS DE FINANCEMENT (III)": flux_fin,
            "VARIATION NETTE DE TRÉSORERIE (I + II + III)": var_nette,
            "Trésorerie à l'ouverture de l'exercice": treso_ouv,
            "TRÉSORERIE À LA CLÔTURE DE L'EXERCICE": treso_clo,
        }
    return resultats


def _chart_tft(resultats, exercices, devise):
    sections = [
        "FLUX NET DE TRÉSORERIE — ACTIVITÉS OPÉRATIONNELLES (I)",
        "FLUX NET DE TRÉSORERIE — ACTIVITÉS D'INVESTISSEMENT (II)",
        "FLUX NET DE TRÉSORERIE — ACTIVITÉS DE FINANCEMENT (III)",
    ]
    labels_short = ["Opérationnel", "Investissement", "Financement"]
    colors = ["#1E8449", "#2980B9", "#8E44AD"]

    fig = go.Figure()
    for sec, label, color in zip(sections, labels_short, colors):
        fig.add_trace(go.Bar(
            name=label,
            x=exercices,
            y=[resultats[ex][sec] for ex in exercices],
            marker_color=color,
            hovertemplate=f"<b>%{{x}}</b><br>{label} : %{{y:,.0f}} {devise}<extra></extra>"
        ))

    # Ligne trésorerie clôture
    fig.add_trace(go.Scatter(
        name="Trésorerie clôture",
        x=exercices,
        y=[resultats[ex]["TRÉSORERIE À LA CLÔTURE DE L'EXERCICE"] for ex in exercices],
        mode="lines+markers",
        line=dict(color="#E67E22", width=2.5),
        marker=dict(size=9),
        yaxis="y2",
        hovertemplate="Trésorerie clôture : %{y:,.0f} " + devise + "<extra></extra>"
    ))

    fig.update_layout(
        title="TFT SYSCOHADA — Flux par activité",
        barmode="group", height=420,
        margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", y=-0.18),
        yaxis=dict(title=f"Flux ({devise})", tickformat=",.0f"),
        yaxis2=dict(title="Trésorerie clôture", overlaying="y",
                    side="right", tickformat=",.0f"),
        hovermode="x unified",
    )
    return fig


def _export_excel_tft(data, resultats, exercices, entreprise, devise):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rows = []
        for section, info in TFT_STRUCTURE.items():
            rows.append({"Libellé": section, **{ex: "" for ex in exercices}, "_type": "header"})
            for lib, signe in info["lignes"]:
                row = {"Libellé": f"  {lib} ({signe})"}
                for ex in exercices:
                    row[ex] = data.get(lib, {}).get(ex, 0)
                row["_type"] = "detail"
                rows.append(row)
            # Total section
            key = f"FLUX NET DE TRÉSORERIE — ACTIVITÉS {section.split('—')[1].strip().split('(')[0].strip()} ({section[-2]})"
            matching = [k for k in TOTAUX if section[-2] + ")" in k]
            if matching:
                row = {"Libellé": matching[0]}
                for ex in exercices:
                    row[ex] = resultats[ex].get(matching[0], 0)
                row["_type"] = "total"
                rows.append(row)

        # Trésorerie
        for lab in ["VARIATION NETTE DE TRÉSORERIE (I + II + III)",
                    "Trésorerie à l'ouverture de l'exercice",
                    "TRÉSORERIE À LA CLÔTURE DE L'EXERCICE"]:
            row = {"Libellé": lab}
            for ex in exercices:
                row[ex] = resultats[ex].get(lab, 0) if lab != "Trésorerie à l'ouverture de l'exercice" else data.get(lab, {}).get(ex, 0)
            row["_type"] = "total" if lab != "Trésorerie à l'ouverture de l'exercice" else "detail"
            rows.append(row)

        df_exp = pd.DataFrame(rows)
        types  = df_exp.pop("_type")
        df_exp.to_excel(writer, index=False, sheet_name="TFT SYSCOHADA")
        ws = writer.sheets["TFT SYSCOHADA"]

        header_fill  = PatternFill("solid", fgColor="1F4E79")
        section_fill = PatternFill("solid", fgColor="D6EAF8")
        total_fill   = PatternFill("solid", fgColor="D5F5E3")
        neg_fill     = PatternFill("solid", fgColor="FADBD8")

        for i, (_, row_type) in enumerate(zip(df_exp.itertuples(), types), start=2):
            for cell in ws[i]:
                if row_type == "header":
                    cell.fill = section_fill
                    cell.font = Font(bold=True)
                elif row_type == "total":
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.fill = neg_fill

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 55
        for j in range(2, len(exercices) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 20

    buf.seek(0)
    return buf


def _analyser_tft_ia(resultats, exercices, devise, entreprise):
    from utils.ai import appel_mistral

    lignes = []
    _k_inv  = "FLUX NET DE TRÉSORERIE — ACTIVITÉS D'INVESTISSEMENT (II)"
    _k_clot = "TRÉSORERIE À LA CLÔTURE DE L'EXERCICE"
    for ex in exercices:
        r = resultats[ex]
        _op  = r["FLUX NET DE TRÉSORERIE — ACTIVITÉS OPÉRATIONNELLES (I)"]
        _inv = r[_k_inv]
        _fin = r["FLUX NET DE TRÉSORERIE — ACTIVITÉS DE FINANCEMENT (III)"]
        _var = r["VARIATION NETTE DE TRÉSORERIE (I + II + III)"]
        _clo = r[_k_clot]
        lignes.append(
            f"  {ex} : Op={_op:+,.0f} | Inv={_inv:+,.0f} | "
            f"Fin={_fin:+,.0f} | Var={_var:+,.0f} | "
            f"Tréso clôture={_clo:,.0f} {devise}"
        )

    prompt = f"""Tu es un expert-comptable SYSCOHADA analysant un Tableau de Flux de Trésorerie.

Entreprise : {entreprise or 'Non précisée'}
Exercice(s) : {', '.join(exercices)}
Devise : {devise}

Données TFT :
{chr(10).join(lignes)}

Analyse ce TFT selon les normes SYSCOHADA révisées 2017 :

## 1. Qualité des flux opérationnels
L'entreprise génère-t-elle suffisamment de cash par son activité principale ?

## 2. Politique d'investissement
Analyse des flux d'investissement : croissance, maintenance, désinvestissement.

## 3. Structure de financement
Les choix de financement sont-ils adaptés à la situation ?

## 4. Évolution de la trésorerie
Tendance et viabilité de la position de trésorerie.

## 5. Points de vigilance SYSCOHADA
Risques spécifiques selon les normes OHADA.

## 6. Recommandations
Actions prioritaires pour améliorer la gestion de trésorerie.
"""
    return appel_mistral(prompt)


# ─── Page principale ─────────────────────────────────────────────────────────

def page_tft():
    st.title("💹 TFT SYSCOHADA")
    st.markdown("*Tableau de Flux de Trésorerie — Norme SYSCOHADA révisée 2017*")
    st.divider()

    # ── Paramètres ───────────────────────────────────────────────────────
    col1, col2, col3 = st.columns(3)
    with col1:
        annee_n = st.number_input("Exercice N (le plus récent)", value=2025,
                                   min_value=2000, max_value=2050, step=1)
    with col2:
        nb_ex = st.selectbox("Nombre d'exercices", [1, 2, 3], index=1,
                              help="1=N seul | 2=N et N-1 | 3=N, N-1 et N-2")
    with col3:
        devise = st.selectbox("Devise", ["FCFA","XOF","XAF","EUR","USD"])

    entreprise = st.text_input("Entreprise", placeholder="Nom de l'entreprise")
    exercices  = [str(int(annee_n) - i) for i in range(nb_ex - 1, -1, -1)]
    st.caption(f"Exercices analysés : **{' | '.join(exercices)}**")
    st.divider()

    # ── Saisie des données ────────────────────────────────────────────────
    if "tft_data" not in st.session_state:
        st.session_state.tft_data = {}

    data = st.session_state.tft_data

    for section, info in TFT_STRUCTURE.items():
        st.markdown(f"#### {section}")
        cols_h = st.columns([3] + [1]*len(exercices))
        cols_h[0].markdown("**Libellé**")
        for i, ex in enumerate(exercices):
            cols_h[i+1].markdown(f"**{ex}**")

        for lib, signe in info["lignes"]:
            if lib not in data:
                data[lib] = {}
            cols_row = st.columns([3] + [1]*len(exercices))
            cols_row[0].markdown(f"<small>{lib} <span style='color:gray'>({signe})</span></small>",
                                  unsafe_allow_html=True)
            for i, ex in enumerate(exercices):
                val = cols_row[i+1].number_input(
                    "", value=float(data[lib].get(ex, 0.0)),
                    step=100_000.0, format="%.0f",
                    key=f"tft_{lib}_{ex}",
                    label_visibility="collapsed"
                )
                data[lib][ex] = val

    # Trésorerie ouverture
    st.divider()
    st.markdown("#### 💰 Trésorerie")
    lib_treso = "Trésorerie à l'ouverture de l'exercice"
    if lib_treso not in data:
        data[lib_treso] = {}
    cols_treso = st.columns([3] + [1]*len(exercices))
    cols_treso[0].markdown(f"**{lib_treso}**")
    for i, ex in enumerate(exercices):
        val = cols_treso[i+1].number_input(
            "", value=float(data[lib_treso].get(ex, 0.0)),
            step=100_000.0, format="%.0f",
            key=f"tft_treso_{ex}",
            label_visibility="collapsed"
        )
        data[lib_treso][ex] = val

    st.session_state.tft_data = data
    st.divider()

    if st.button("💹 Générer le TFT SYSCOHADA", type="primary", use_container_width=True):
        resultats = _calculer_tft(data, exercices)
        st.session_state["tft_resultats"] = resultats
        st.session_state["tft_exercices"] = exercices

    if st.session_state.get("tft_resultats"):
        resultats = st.session_state["tft_resultats"]
        exercices = st.session_state["tft_exercices"]

        # ── KPIs ──────────────────────────────────────────────────────────
        st.subheader("📊 Synthèse")
        cols_k = st.columns(len(exercices))
        for i, ex in enumerate(exercices):
            r = resultats[ex]
            var = r["VARIATION NETTE DE TRÉSORERIE (I + II + III)"]
            treso = r["TRÉSORERIE À LA CLÔTURE DE L'EXERCICE"]
            cols_k[i].metric(
                f"Trésorerie clôture {ex}",
                f"{treso:,.0f} {devise}",
                delta=f"{var:+,.0f}",
                delta_color="normal" if var >= 0 else "inverse"
            )

        # Alertes
        for ex in exercices:
            r = resultats[ex]
            op = r["FLUX NET DE TRÉSORERIE — ACTIVITÉS OPÉRATIONNELLES (I)"]
            if op < 0:
                st.error(f"🚨 {ex} : Flux opérationnels négatifs ({op:,.0f} {devise}) — l'activité détruit de la trésorerie !")
            if r["TRÉSORERIE À LA CLÔTURE DE L'EXERCICE"] < 0:
                st.warning(f"⚠️ {ex} : Trésorerie clôture négative — risque de cessation de paiements.")

        st.divider()

        # ── Tableau TFT complet ────────────────────────────────────────────
        st.subheader("📋 TFT Complet")
        rows_display = []
        for section, info in TFT_STRUCTURE.items():
            rows_display.append({"Libellé": f"**{section}**",
                                  **{ex: "" for ex in exercices}})
            for lib, signe in info["lignes"]:
                row = {"Libellé": f"  {lib}"}
                for ex in exercices:
                    val = data.get(lib, {}).get(ex, 0)
                    row[ex] = f"{val:,.0f}"
                rows_display.append(row)

            matching = [k for k in resultats[exercices[0]] if section[-2]+")" in k]
            if matching:
                row = {"Libellé": f"**{matching[0]}**"}
                for ex in exercices:
                    v = resultats[ex][matching[0]]
                    row[ex] = f"**{v:+,.0f}**"
                rows_display.append(row)

        for lab in ["VARIATION NETTE DE TRÉSORERIE (I + II + III)",
                    "Trésorerie à l'ouverture de l'exercice",
                    "TRÉSORERIE À LA CLÔTURE DE L'EXERCICE"]:
            row = {"Libellé": f"**{lab}**"}
            for ex in exercices:
                if lab == "Trésorerie à l'ouverture de l'exercice":
                    v = data.get(lab, {}).get(ex, 0)
                else:
                    v = resultats[ex].get(lab, 0)
                row[ex] = f"**{v:+,.0f}**"
            rows_display.append(row)

        st.dataframe(pd.DataFrame(rows_display).set_index("Libellé"),
                     use_container_width=True)

        # ── Graphique ─────────────────────────────────────────────────────
        st.plotly_chart(_chart_tft(resultats, exercices, devise), use_container_width=True)

        # ── Analyse IA ────────────────────────────────────────────────────
        st.divider()
        if st.button("🤖 Analyse IA du TFT", use_container_width=True):
            with st.spinner("Analyse IA SYSCOHADA en cours..."):
                analyse = _analyser_tft_ia(resultats, exercices, devise, entreprise)
                st.session_state["tft_analyse_ia"] = analyse

        if st.session_state.get("tft_analyse_ia"):
            st.subheader("🤖 Analyse IA SYSCOHADA")
            st.markdown(st.session_state["tft_analyse_ia"])

        # ── Export ────────────────────────────────────────────────────────
        st.divider()
        excel_buf = _export_excel_tft(data, resultats, exercices, entreprise, devise)
        st.download_button(
            "📊 Télécharger TFT Excel",
            excel_buf,
            f"TFT_SYSCOHADA_{entreprise or 'SMD'}_{exercices[-1]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
