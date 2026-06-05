# -*- coding: utf-8 -*-
"""
Liasse Fiscale Officielle — DGID Sénégal
Format SYSCOHADA révisé 2017 — États Annuels (SAES)
Mapping comptes SYSCOHADA → lignes officielles DGID
RevisionPro SYSCOHADA — SMD Global Consulting LLC © 2026
"""

import pandas as pd
import io

# =============================================================================
# HELPERS DE CALCUL
# =============================================================================

def _solde(df: pd.DataFrame, prefixes: list, sens: str = 'D') -> float:
    """
    Agrège les soldes pour une liste de préfixes de numéros de comptes.
    sens='D' → solde débiteur net (actif / charges)
    sens='C' → solde créditeur net (passif / produits)
    """
    if not prefixes:
        return 0.0
    pattern = '|'.join(f'^{p}' for p in prefixes)
    mask = df['compte'].astype(str).str.match(pattern, na=False)
    subset = df[mask]
    if subset.empty:
        return 0.0
    total_d = pd.to_numeric(subset['debit'],  errors='coerce').fillna(0).sum()
    total_c = pd.to_numeric(subset['credit'], errors='coerce').fillna(0).sum()
    return float(max(total_d - total_c, 0.0) if sens == 'D' else max(total_c - total_d, 0.0))


def _brut(df: pd.DataFrame, prefixes: list) -> float:
    """Solde brut débiteur (valeur d'origine des immobilisations)"""
    return _solde(df, prefixes, 'D')


def _amort(df: pd.DataFrame, prefixes_amort: list) -> float:
    """Amortissements / dépréciations (solde créditeur comptes 28x, 29x, etc.)"""
    return _solde(df, prefixes_amort, 'C')


def _net(df: pd.DataFrame, prefixes_brut: list, prefixes_amort: list = None) -> tuple:
    """Retourne (brut, amort, net)"""
    b = _brut(df, prefixes_brut)
    a = _amort(df, prefixes_amort) if prefixes_amort else 0.0
    return b, a, max(b - a, 0.0)


def _fmt(val):
    """Formate un nombre en entier avec séparateur de milliers"""
    if isinstance(val, (int, float)):
        if val == 0:
            return "-"
        return f"{val:,.0f}".replace(",", " ")
    return val


# =============================================================================
# TABLEAU 1 — BILAN ACTIF
# =============================================================================

def generer_bilan_actif(df: pd.DataFrame) -> pd.DataFrame:
    """
    Bilan Actif SYSCOHADA — format DGID Sénégal
    Colonnes : REF | POSTE | NOTE | BRUT | AMORT/DEPREC | NET N
    """
    lignes = []

    def L(ref, poste, pfx_brut, pfx_amort=None, note=""):
        b, a, n = _net(df, pfx_brut, pfx_amort)
        lignes.append(dict(REF=ref, POSTE=poste, NOTE=note,
                           BRUT=b, AMORT=a, NET=n, _b=b, _a=a, _n=n))

    def TITRE(label):
        lignes.append(dict(REF="", POSTE=f"▶ {label}", NOTE="",
                           BRUT="", AMORT="", NET="", _b=0, _a=0, _n=0))

    def TOTAL(ref, poste, keys):
        b = sum(r['_b'] for r in lignes if r.get('REF') in keys)
        a = sum(r['_a'] for r in lignes if r.get('REF') in keys)
        n = sum(r['_n'] for r in lignes if r.get('REF') in keys)
        lignes.append(dict(REF=ref, POSTE=poste, NOTE="",
                           BRUT=b, AMORT=a, NET=n, _b=b, _a=a, _n=n))

    # ── ACTIF IMMOBILISÉ ─────────────────────────────────────────────
    TITRE("ACTIF IMMOBILISÉ")
    L('AA', 'Charges immobilisées',                        ['20'],         ['280'],         '2')
    L('AB', 'Frais de recherche et développement',         ['211'],        ['2811'],        '2')
    L('AC', 'Brevets, licences, logiciels et droits',      ['212', '213'], ['2812', '2813'],'2')
    L('AD', 'Fonds commercial',                            ['215'],        ['2815'],        '2')
    L('AE', 'Autres immobilisations incorporelles',        ['218'],        ['2818'],        '2')
    L('AF', 'Terrains',                                    ['22'],         ['282'],         '3')
    L('AG', 'Bâtiments',                                   ['231', '232'], ['2831', '2832'],'3')
    L('AH', 'Aménagements, agencements, installations',    ['233', '234'], ['2833', '2834'],'3')
    L('AI', 'Matériel, mobilier et actifs biologiques',    ['24'],         ['284'],         '3')
    L('AJ', 'Matériel de transport',                       ['245'],        ['2845'],        '3')
    L('AK', 'Avances et acomptes/immobilisations',         ['251', '252'], [],              '3')
    L('AL', 'Titres de participation',                     ['261', '262'], ['2961', '2962'],'4')
    L('AM', 'Autres immobilisations financières',          ['27'],         ['297'],         '4')
    TOTAL('AZ', 'TOTAL ACTIF IMMOBILISÉ (AA→AM)',
          ['AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM'])

    # ── ACTIF CIRCULANT ──────────────────────────────────────────────
    TITRE("ACTIF CIRCULANT")
    L('BA', 'Actif circulant HAO',                         ['45'],                 ['495'], '5')
    L('BB', 'Stocks de marchandises',                      ['31'],                 ['391'], '6')
    L('BC', 'Stocks de matières premières et fournitures', ['32'],                 ['392'], '6')
    L('BD', 'En-cours de production',                      ['33', '34'],           ['393', '394'], '6')
    L('BE', 'Stocks de produits fabriqués',                ['35'],                 ['395'], '6')
    L('BF', 'Stocks de produits agricoles et sylvicoles',  ['36'],                 ['396'], '6')
    L('BG', 'Fournisseurs, avances versées',               ['4091', '4092'],       [],      '7')
    L('BH', 'Clients et comptes rattachés',                ['41'],                 ['491'], '7')
    L('BI', 'Autres créances',                             ['42','43','44','46','47','48'], ['499'], '7')
    TOTAL('BZ', 'TOTAL ACTIF CIRCULANT (BA→BI)',
          ['BA','BB','BC','BD','BE','BF','BG','BH','BI'])

    # ── TRÉSORERIE ACTIF ─────────────────────────────────────────────
    TITRE("TRÉSORERIE - ACTIF")
    L('CA', 'Titres de placement',                         ['50'],             ['590'])
    L('CB', 'Valeurs à encaisser',                         ['51'],             [])
    L('CC', 'Banques, chèques postaux, caisse',            ['52','53','54','57'], [])
    TOTAL('CZ', 'TOTAL TRÉSORERIE - ACTIF (CA+CB+CC)', ['CA','CB','CC'])

    # ── ÉCART DE CONVERSION ACTIF ────────────────────────────────────
    L('DA', 'Écart de conversion actif',                   ['476'],            [])

    # ── TOTAL GÉNÉRAL ACTIF ──────────────────────────────────────────
    TOTAL('DZ', 'TOTAL GÉNÉRAL ACTIF (AZ+BZ+CZ+DA)', ['AZ','BZ','CZ','DA'])

    # Mise en forme
    result = []
    for r in lignes:
        is_titre = r.get('REF') == "" and r['POSTE'].startswith('▶')
        is_total = r.get('REF') in ['AZ','BZ','CZ','DZ']
        result.append({
            'REF':         r['REF'],
            'POSTE':       r['POSTE'],
            'NOTE':        r.get('NOTE', ''),
            'BRUT':        _fmt(r['BRUT']) if not is_titre else '',
            'AMORT/DEPREC':_fmt(r['AMORT']) if not is_titre else '',
            'NET N':       _fmt(r['NET']) if not is_titre else '',
            '_total':      is_total,
        })

    return pd.DataFrame(result, columns=['REF','POSTE','NOTE','BRUT','AMORT/DEPREC','NET N'])


# =============================================================================
# TABLEAU 2 — BILAN PASSIF
# =============================================================================

def generer_bilan_passif(df: pd.DataFrame) -> pd.DataFrame:
    """
    Bilan Passif SYSCOHADA — format DGID Sénégal
    Colonnes : REF | POSTE | NOTE | MONTANT N
    """
    lignes = []

    def L(ref, poste, pfx, sens='C', note=""):
        val = _solde(df, pfx, sens)
        lignes.append(dict(REF=ref, POSTE=poste, NOTE=note, VAL=val))

    def TITRE(label):
        lignes.append(dict(REF="", POSTE=f"▶ {label}", NOTE="", VAL=None))

    def TOTAL(ref, poste, keys):
        val = sum(r['VAL'] for r in lignes if r.get('REF') in keys and r['VAL'] is not None)
        lignes.append(dict(REF=ref, POSTE=poste, NOTE="", VAL=val))

    # ── CAPITAUX PROPRES ─────────────────────────────────────────────
    TITRE("CAPITAUX PROPRES ET RESSOURCES ASSIMILÉES")
    L('CA', 'Capital',                                     ['101','102','103','104'], 'C', '8')
    L('CB', 'Apporteurs, capital souscrit non appelé',     ['109'],                  'D', '8')
    L('CC', 'Primes liées au capital social',              ['105','106'],            'C', '8')
    L('CD', 'Écarts de réévaluation',                      ['1061'],                 'C')
    L('CE', 'Réserves indisponibles',                      ['111','112','113'],      'C', '9')
    L('CF', 'Réserves libres',                             ['118'],                  'C', '9')
    L('CG', 'Report à nouveau (+ ou -)',                   ['12'],                   'C')
    L('CH', 'Résultat net de l\'exercice (bénéfice +)',    ['13'],                   'C', '13')
    L('CI', 'Subventions d\'investissement',               ['14'],                   'C', '10')
    L('CJ', 'Provisions réglementées',                     ['15'],                   'C', '11')
    TOTAL('CZ', 'TOTAL CAPITAUX PROPRES (CA-CB+CC+CD+CE+CF+CG+CH+CI+CJ)',
          ['CA','CB','CC','CD','CE','CF','CG','CH','CI','CJ'])

    # ── DETTES FINANCIÈRES ────────────────────────────────────────────
    TITRE("DETTES FINANCIÈRES ET RESSOURCES ASSIMILÉES")
    L('DA', 'Emprunts',                                    ['161','162','163'],      'C', '12')
    L('DB', 'Dettes de location-acquisition',              ['164'],                  'C', '12')
    L('DC', 'Provisions financières pour risques/charges', ['19'],                   'C')
    L('DD', 'Fournisseurs d\'investissement, avances reçues',['481','482'],          'C')
    L('DE', 'Autres dettes financières',                   ['165','166','167','168'],'C')
    TOTAL('DZ', 'TOTAL DETTES FINANCIÈRES (DA+DB+DC+DD+DE)',
          ['DA','DB','DC','DD','DE'])

    TOTAL('EZ', 'TOTAL RESSOURCES DURABLES (CZ+DZ)', ['CZ','DZ'])

    # ── PASSIF CIRCULANT ──────────────────────────────────────────────
    TITRE("PASSIF CIRCULANT")
    L('FA', 'Passif circulant HAO',                        ['45'],                   'C')
    L('FB', 'Clients, avances reçues',                     ['4191','4192'],          'C')
    L('FC', 'Fournisseurs d\'exploitation',                ['401','402','408'],      'C', '14')
    L('FD', 'Dettes fiscales',                             ['441','442','443','444'],'C', '15')
    L('FE', 'Dettes sociales',                             ['421','422','423','424','425','426','427','428'], 'C', '15')
    L('FF', 'Autres dettes et provisions court terme',     ['46','47','48','49'],    'C')
    TOTAL('FZ', 'TOTAL PASSIF CIRCULANT (FA+FB+FC+FD+FE+FF)',
          ['FA','FB','FC','FD','FE','FF'])

    # ── TRÉSORERIE PASSIF ─────────────────────────────────────────────
    TITRE("TRÉSORERIE - PASSIF")
    L('GA', 'Banques, crédits d\'escompte',                ['561','562'],            'C')
    L('GB', 'Banques, crédits de trésorerie et d\'escompte',['563','564','565'],     'C')
    TOTAL('GZ', 'TOTAL TRÉSORERIE - PASSIF (GA+GB)', ['GA','GB'])

    # ── ÉCART DE CONVERSION PASSIF ────────────────────────────────────
    L('HA', 'Écart de conversion passif',                  ['477'],                  'C')

    # ── TOTAL GÉNÉRAL PASSIF ──────────────────────────────────────────
    TOTAL('HZ', 'TOTAL GÉNÉRAL PASSIF (EZ+FZ+GZ+HA)', ['EZ','FZ','GZ','HA'])

    # Mise en forme
    result = []
    for r in lignes:
        result.append({
            'REF':        r['REF'],
            'POSTE':      r['POSTE'],
            'NOTE':       r.get('NOTE', ''),
            'MONTANT N':  _fmt(r['VAL']) if r['VAL'] is not None else '',
        })

    return pd.DataFrame(result, columns=['REF','POSTE','NOTE','MONTANT N'])


# =============================================================================
# TABLEAU 3 — COMPTE DE RÉSULTAT
# =============================================================================

def generer_compte_resultat_dgid(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compte de Résultat SYSCOHADA — format DGID Sénégal
    Colonnes : REF | POSTE | NOTE | MONTANT N
    """
    lignes = []

    def P(ref, poste, pfx, note=""):
        """Produit (crédit)"""
        val = _solde(df, pfx, 'C')
        lignes.append(dict(REF=ref, POSTE=poste, NOTE=note, VAL=val))
        return val

    def C(ref, poste, pfx, note=""):
        """Charge (débit) — stockée en négatif"""
        val = _solde(df, pfx, 'D')
        lignes.append(dict(REF=ref, POSTE=poste, NOTE=note, VAL=-val))
        return val

    def TITRE(label):
        lignes.append(dict(REF="", POSTE=f"▶ {label}", NOTE="", VAL=None))

    def SIG(ref, poste, keys):
        val = sum(r['VAL'] for r in lignes if r.get('REF') in keys and r['VAL'] is not None)
        lignes.append(dict(REF=ref, POSTE=poste, NOTE="", VAL=val))
        return val

    # ── ACTIVITÉ D'EXPLOITATION ───────────────────────────────────────
    TITRE("ACTIVITÉ D'EXPLOITATION")

    v_march  = P('TA', 'Ventes de marchandises',                         ['701'],              '16')
    a_march  = C('RA', '(-) Achats de marchandises',                     ['601'])
    vs_march = _solde(df, ['6031'], 'D') - _solde(df, ['6031'], 'C')   # variation stocks march
    lignes.append(dict(REF='RB', POSTE='(-) Variation de stocks de marchandises', NOTE='', VAL=-vs_march))

    marge = v_march - a_march + vs_march   # attention RB est déjà signé
    marge = v_march - _solde(df, ['601'], 'D') - vs_march
    lignes[-1]['VAL'] = -vs_march
    # recalcul propre
    marge = (_solde(df, ['701'], 'C')
             - _solde(df, ['601'], 'D')
             + (_solde(df, ['6031'], 'C') - _solde(df, ['6031'], 'D')))

    lignes.append(dict(REF='XA', POSTE='MARGE COMMERCIALE (TA-RA-RB)', NOTE='', VAL=marge))

    P('TB', 'Ventes de produits fabriqués',                              ['702'],              '16')
    P('TC', 'Travaux, services vendus',                                  ['703','704','705'],  '16')
    P('TD', 'Produits accessoires',                                      ['706','707','708'],  '16')

    ca = (_solde(df, ['701','702','703','704','705','706','707','708'], 'C'))
    lignes.append(dict(REF='XB', POSTE="CHIFFRE D'AFFAIRES (TA+TB+TC+TD)", NOTE='', VAL=ca))

    # Production stockée : 73 crédit - débit
    prod_stock = _solde(df, ['73'], 'C') - _solde(df, ['73'], 'D')
    lignes.append(dict(REF='TE', POSTE='Production stockée (+) ou déstockée (-)',  NOTE='', VAL=prod_stock))
    P('TF', 'Production immobilisée',                                    ['72'])
    P('TG', "Subventions d'exploitation",                                ['71'])
    P('TH', 'Autres produits',                                           ['75'],               '17')
    P('TI', "Transferts de charges d'exploitation",                      ['791'])

    C('RC', '(-) Achats de matières premières et fournitures liées',    ['602'])
    vs_mp = _solde(df, ['6032','6033'], 'D') - _solde(df, ['6032','6033'], 'C')
    lignes.append(dict(REF='RD', POSTE='(-) Variation de stocks de matières premières', NOTE='', VAL=-vs_mp))
    C('RE', '(-) Autres achats',                                         ['604','605','606','607','608'])
    C('RF', '(-) Transports',                                            ['61'])
    C('RG', '(-) Services extérieurs',                                   ['62','63'])
    C('RH', '(-) Impôts et taxes',                                       ['64'])
    C('RI', '(-) Autres charges',                                        ['65'],               '18')

    SIG('XC', 'VALEUR AJOUTÉE (XB+TE+TF+TG+TH+TI-RC-RD-RE-RF-RG-RH-RI)',
        ['XB','TE','TF','TG','TH','TI','RC','RD','RE','RF','RG','RH','RI'])

    C('RJ', '(-) Charges de personnel',                                  ['66'],               '19')
    SIG('XD', "EXCÉDENT BRUT D'EXPLOITATION (XC-RJ)", ['XC','RJ'])

    P('TJ', "Reprises d'amortissements, provisions et dépréciations",   ['798','799'])
    P('TK', 'Transferts de charges',                                     ['792','793','794','795','796','797'])
    C('RK', '(-) Dotations aux amortissements, provisions et dépréciations',
      ['681','682','683','684','685','686','687','688'],                                        '20')
    SIG('XE', "RÉSULTAT D'EXPLOITATION (XD+TJ+TK-RK)", ['XD','TJ','TK','RK'])

    # ── ACTIVITÉ FINANCIÈRE ───────────────────────────────────────────
    TITRE("ACTIVITÉ FINANCIÈRE")
    P('TL', 'Revenus financiers et assimilés',                           ['77'],               '21')
    P('TM', 'Reprises de provisions et dépréciations financières',       ['797'])
    P('TN', 'Transferts de charges financières',                         ['796'])
    C('RL', '(-) Frais financiers et charges assimilées',                ['67'],               '21')
    C('RM', '(-) Dotations aux provisions et dépréciations financières', ['697'])
    SIG('XF', 'RÉSULTAT FINANCIER (TL+TM+TN-RL-RM)', ['TL','TM','TN','RL','RM'])
    SIG('XG', "RÉSULTAT DES ACTIVITÉS ORDINAIRES (XE+XF)", ['XE','XF'])

    # ── HORS ACTIVITÉS ORDINAIRES ─────────────────────────────────────
    TITRE("HORS ACTIVITÉS ORDINAIRES (HAO)")
    P('TO', "Produits des cessions d'immobilisations",                   ['82'],               '22')
    P('TP', 'Autres Produits HAO',                                       ['84','86','88'],     '22')
    P('TQ', 'Reprises HAO',                                              ['856','866'])
    C('RN', "(-) Valeurs comptables des cessions d'immobilisations",     ['81'],               '22')
    C('RO', '(-) Autres Charges HAO',                                    ['83','85','87'],     '22')
    C('RP', '(-) Dotations HAO',                                         ['851','861'])
    SIG('XH', 'RÉSULTAT HAO (TO+TP+TQ-RN-RO-RP)', ['TO','TP','TQ','RN','RO','RP'])

    # ── RÉSULTAT NET ──────────────────────────────────────────────────
    C('RQ', '(-) Participation des travailleurs',                        ['664'])
    C('RS', '(-) Impôts sur le résultat',                                ['891','892'],        '23')
    SIG('XI', 'RÉSULTAT NET (XG+XH-RQ-RS)', ['XG','XH','RQ','RS'])

    # Mise en forme
    result = []
    for r in lignes:
        result.append({
            'REF':       r['REF'],
            'POSTE':     r['POSTE'],
            'NOTE':      r.get('NOTE', ''),
            'MONTANT N': _fmt(r['VAL']) if r['VAL'] is not None else '',
        })

    return pd.DataFrame(result, columns=['REF','POSTE','NOTE','MONTANT N'])


# =============================================================================
# EXPORT EXCEL — 3 ONGLETS
# =============================================================================

def export_liasse_excel(df_balance: pd.DataFrame, entreprise: str = "", exercice: str = "") -> bytes:
    """
    Génère un fichier Excel avec 3 onglets :
    - Bilan Actif
    - Bilan Passif
    - Compte de Résultat
    Retourne les bytes du fichier .xlsx
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Couleurs
    BLEU_DGID   = "1F4E79"   # En-têtes
    BLEU_CLAIR  = "D6E4F0"   # Totaux
    VERT_CLAIR  = "E8F5E9"   # SIG (soldes intermédiaires)
    GRIS        = "F5F5F5"   # Titres de section
    BLANC       = "FFFFFF"

    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row, cols, fill_color=BLEU_DGID, font_color="FFFFFF"):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(bold=True, color=font_color, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    def style_titre_section(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=GRIS)
            cell.font = Font(bold=True, color="1F4E79", size=9)
            cell.border = border

    def style_total(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=BLEU_CLAIR)
            cell.font = Font(bold=True, size=9)
            cell.border = border

    def style_sig(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=VERT_CLAIR)
            cell.font = Font(bold=True, size=9)
            cell.border = border

    def write_df_to_sheet(ws, df_data, col_widths, start_row=4, totaux_refs=None, sig_refs=None):
        totaux_refs = totaux_refs or []
        sig_refs    = sig_refs or []
        # En-têtes colonnes
        for ci, col in enumerate(df_data.columns, 1):
            ws.cell(row=start_row, column=ci).value = col
        style_header(ws, start_row, len(df_data.columns))
        # Données
        for ri, row_data in enumerate(df_data.itertuples(index=False), start_row + 1):
            ref = row_data[0]
            is_titre   = ref == "" and str(row_data[1]).startswith('▶')
            is_total   = ref in totaux_refs
            is_sig     = ref in sig_refs
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci)
                cell.value = str(val) if val else ""
                cell.font = Font(size=9)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            if is_titre:
                style_titre_section(ws, ri, len(df_data.columns))
            elif is_total:
                style_total(ws, ri, len(df_data.columns))
            elif is_sig:
                style_sig(ws, ri, len(df_data.columns))
        # Largeurs colonnes
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[start_row].height = 30

    # ── En-tête commune ──────────────────────────────────────────────
    def en_tete(ws, titre):
        ws.merge_cells('A1:F1')
        ws['A1'] = f"ÉTATS FINANCIERS ANNUELS — SYSCOHADA — DGID SÉNÉGAL"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill("solid", fgColor=BLEU_DGID)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22
        ws.merge_cells('A2:F2')
        ws['A2'] = f"{titre}  |  Entreprise : {entreprise or 'N/A'}  |  Exercice : {exercice or 'N/A'}"
        ws['A2'].font = Font(bold=True, size=10, color="1F4E79")
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

    # ── ONGLET 1 : BILAN ACTIF ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Bilan Actif"
    en_tete(ws1, "BILAN — ACTIF")
    df_ba = generer_bilan_actif(df_balance)
    write_df_to_sheet(ws1, df_ba,
                      col_widths=[5, 52, 5, 16, 16, 16],
                      totaux_refs=['AZ', 'BZ', 'CZ', 'DZ'],
                      sig_refs=[])

    # ── ONGLET 2 : BILAN PASSIF ──────────────────────────────────────
    ws2 = wb.create_sheet("Bilan Passif")
    en_tete(ws2, "BILAN — PASSIF")
    df_bp = generer_bilan_passif(df_balance)
    write_df_to_sheet(ws2, df_bp,
                      col_widths=[5, 60, 5, 18],
                      totaux_refs=['CZ', 'DZ', 'EZ', 'FZ', 'GZ', 'HZ'],
                      sig_refs=[])

    # ── ONGLET 3 : COMPTE DE RÉSULTAT ────────────────────────────────
    ws3 = wb.create_sheet("Compte de Résultat")
    en_tete(ws3, "COMPTE DE RÉSULTAT")
    df_cr = generer_compte_resultat_dgid(df_balance)
    write_df_to_sheet(ws3, df_cr,
                      col_widths=[5, 62, 5, 18],
                      totaux_refs=['XA','XB','XC','XD','XE','XF','XG','XH','XI'],
                      sig_refs=['XA','XB','XC','XD','XE','XF','XG','XH','XI'])

    # ── ONGLET 4 : TAFIRE ───────────────────────────────────────────────
    ws4 = wb.create_sheet("TAFIRE")
    en_tete(ws4, "TAFIRE — TABLEAU DE FINANCEMENT")
    df_taf = generer_tafire(df_balance)
    write_df_to_sheet(ws4, df_taf,
                      col_widths=[6, 72, 14, 18],
                      totaux_refs=["ZC","ZR","ZE","FRG","M4","M8","BFR","TN","EQ","C5"],
                      sig_refs=["FRG","BFR","TN","EQ"])

    # ── ONGLETS 5-6-7 : NOTES ANNEXES ────────────────────────────────────
    notes = generer_notes_annexes(df_balance)

    ws5 = wb.create_sheet("A-Immobilisations")
    en_tete(ws5, "NOTE A — ÉTAT DES IMMOBILISATIONS")
    write_df_to_sheet(ws5, notes["immo"],
                      col_widths=[5, 45, 18, 18, 18],
                      totaux_refs=["AZ"], sig_refs=[])

    ws6 = wb.create_sheet("B-Amort & Provisions")
    en_tete(ws6, "NOTE B — DOTATIONS  |  NOTE C — PROVISIONS")
    # Dotations
    ws6.cell(row=4, column=1).value = "── Tableau B : Dotations aux amortissements (Exercice N) ──"
    ws6.cell(row=4, column=1).font = Font(bold=True, color="1F4E79", size=10)
    write_df_to_sheet(ws6, notes["amort"],
                      col_widths=[8, 52, 20],
                      start_row=5,
                      totaux_refs=["TOT"], sig_refs=[])
    next_row = 5 + len(notes["amort"]) + 3
    ws6.cell(row=next_row, column=1).value = "── Tableau C : Provisions et dépréciations ──"
    ws6.cell(row=next_row, column=1).font = Font(bold=True, color="1F4E79", size=10)
    write_df_to_sheet(ws6, notes["provisions"],
                      col_widths=[8, 52, 20],
                      start_row=next_row + 1,
                      totaux_refs=["TOT"], sig_refs=[])

    ws7 = wb.create_sheet("D-Créances & E-Dettes")
    en_tete(ws7, "NOTE D — CRÉANCES  |  NOTE E — DETTES")
    ws7.cell(row=4, column=1).value = "── Tableau D : État des créances ──"
    ws7.cell(row=4, column=1).font = Font(bold=True, color="1F4E79", size=10)
    write_df_to_sheet(ws7, notes["creances"],
                      col_widths=[8, 44, 16, 14, 16],
                      start_row=5,
                      totaux_refs=["TOT"], sig_refs=[])
    next_row2 = 5 + len(notes["creances"]) + 3
    ws7.cell(row=next_row2, column=1).value = "── Tableau E : État des dettes ──"
    ws7.cell(row=next_row2, column=1).font = Font(bold=True, color="1F4E79", size=10)
    write_df_to_sheet(ws7, notes["dettes"],
                      col_widths=[10, 50, 20],
                      start_row=next_row2 + 1,
                      totaux_refs=["TOT"], sig_refs=[])

    # Sauvegarde en mémoire
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()



# =============================================================================
# TABLEAU 4 — TAFIRE (Tableau de Financement des Ressources et Emplois)
# =============================================================================

def generer_tafire(df: pd.DataFrame) -> pd.DataFrame:
    """
    TAFIRE SYSCOHADA — Tableau de Financement des Ressources et Emplois.
    Calculé depuis une balance unique (exercice N).
    Les variations BFR requièrent les données N-1 pour être complètes ;
    le module affiche les POSITIONS N en attendant.
    """
    lignes = []

    def L(ref, poste, val, note="", _total=False):
        lignes.append({"REF": ref, "POSTE": poste, "NOTE": note, "VAL": val, "_total": _total})

    def TITRE(label):
        lignes.append({"REF": "", "POSTE": f"▶ {label}", "NOTE": "", "VAL": None, "_total": False})

    def TOT(ref, poste, val, note=""):
        L(ref, poste, val, note, _total=True)

    # ── CALCUL DE LA CAF (méthode additive) ─────────────────────────────
    res_net = _solde(df, ["13"], "C") - _solde(df, ["13"], "D")

    dotations = (_solde(df, ["681","682","683","684","685","686","687","688"], "D")
                 + _solde(df, ["851","861"], "D"))

    reprises = (_solde(df, ["798","799"], "C")
                + _solde(df, ["797"], "C")
                + _solde(df, ["856","866"], "C"))

    vnc_cessions  = _solde(df, ["81"], "D")
    prix_cession  = _solde(df, ["82"], "C")
    pv_nette      = prix_cession - vnc_cessions   # + = plus-value

    caf = res_net + dotations - reprises - pv_nette
    dividendes    = _solde(df, ["4641","4642"], "D") + _solde(df, ["129"], "D")
    autofinancement = caf - dividendes

    # ── TABLEAU I : RESSOURCES STABLES ───────────────────────────────────
    TITRE("TABLEAU I — RESSOURCES STABLES DE L'EXERCICE")

    L("ZA", "Capacité d'Autofinancement (CAF/MBA)", caf, "A")
    L("ZB", "(-) Dividendes distribués",             dividendes)
    TOT("ZC", "AUTOFINANCEMENT (ZA - ZB)",           autofinancement)

    cessions_immo   = prix_cession
    aug_capital     = _solde(df, ["4611","4612"], "C")
    subv_invest     = _solde(df, ["141","142","143"], "C")
    emprunts_nv     = _solde(df, ["161","162","163"], "C")
    autres_df_c     = _solde(df, ["164","165","166","167","168"], "C")

    L("RF", "Cessions/réductions d'actif immobilisé",    cessions_immo)
    L("RG", "Augmentation de capital par apports nouveaux", aug_capital)
    L("RH", "Subventions d'investissement obtenues",      subv_invest)
    L("RI", "Emprunts à MLT nouveaux",                    emprunts_nv)
    L("RJ", "Autres dettes financières LT",               autres_df_c)

    tot_ressources = (autofinancement + cessions_immo + aug_capital
                      + subv_invest + emprunts_nv + autres_df_c)
    TOT("ZR", "TOTAL RESSOURCES STABLES (ZC+RF+RG+RH+RI+RJ)", tot_ressources)

    # ── TABLEAU I : EMPLOIS STABLES ──────────────────────────────────────
    TITRE("EMPLOIS STABLES DE L'EXERCICE")

    acqui_immo    = _solde(df, ["21","22","23","24","25","26","27"], "D")
    charges_immo  = _solde(df, ["20"], "D")
    remb_df       = _solde(df, ["161","162","163","164","165","166","167","168"], "D")

    L("EF", "Acquisitions d'éléments d'actif immobilisé", acqui_immo)
    L("EG", "Charges immobilisées de l'exercice",          charges_immo)
    L("EH", "Remboursements de dettes financières MLT",     remb_df)
    L("EI", "Emplois HAO nets (VNC des immob. cédées)",     vnc_cessions)

    tot_emplois = acqui_immo + charges_immo + remb_df + vnc_cessions
    TOT("ZE", "TOTAL EMPLOIS STABLES (EF+EG+EH+EI)", tot_emplois)

    frng = tot_ressources - tot_emplois
    TOT("FRG",
        "VARIATION DU FRNG (ZR - ZE)   [ + = Ressource nette  /  - = Emploi net ]",
        frng, "+ si ressources > emplois")

    # ── TABLEAU II : POSITIONS BFR ET TRÉSORERIE (Exercice N) ────────────
    TITRE("TABLEAU II — BFR ET TRÉSORERIE — POSITIONS EXERCICE N")

    stocks_nets      = (_solde(df, ["31","32","33","34","35","36"], "D")
                        - _solde(df, ["391","392","393","394","395","396"], "C"))
    clients_nets     = _solde(df, ["41"], "D") - _solde(df, ["491"], "C")
    autres_creances  = (_solde(df, ["42","43","44","46","47","48"], "D")
                        - _solde(df, ["499"], "C"))
    bfr_actif        = stocks_nets + clients_nets + autres_creances

    dettes_fourn     = _solde(df, ["401","402","408"], "C")
    dettes_fisc      = _solde(df, ["441","442","443","444"], "C")
    dettes_soc       = _solde(df, ["421","422","423","424","425","426","427","428"], "C")
    autres_det_circ  = _solde(df, ["46","47","48"], "C")
    bfr_passif       = dettes_fourn + dettes_fisc + dettes_soc + autres_det_circ
    bfr_net          = bfr_actif - bfr_passif

    treso_actif      = (_solde(df, ["50","51","52","53","54","57"], "D")
                        - _solde(df, ["590"], "C"))
    treso_passif     = _solde(df, ["561","562","563","564","565"], "C")
    treso_nette      = treso_actif - treso_passif

    L("M1", "Stocks nets",                     stocks_nets)
    L("M2", "Créances clients nettes",         clients_nets)
    L("M3", "Autres créances nettes",          autres_creances)
    TOT("M4", "TOTAL ACTIFS CIRCULANTS D'EXPLOITATION", bfr_actif)

    L("M5", "(-) Dettes fournisseurs",         dettes_fourn)
    L("M6", "(-) Dettes fiscales",             dettes_fisc)
    L("M7", "(-) Dettes sociales",             dettes_soc)
    TOT("M8", "TOTAL PASSIFS CIRCULANTS D'EXPLOITATION", bfr_passif)

    TOT("BFR", "BESOIN EN FONDS DE ROULEMENT (M4 - M8)", bfr_net)

    L("T1", "Trésorerie Actif  (50, 51, 52, 57)", treso_actif)
    L("T2", "(-) Trésorerie Passif  (56x)",       treso_passif)
    TOT("TN", "TRÉSORERIE NETTE  (T1 - T2)",       treso_nette)

    TOT("EQ",
        "ÉQUILIBRE  :  FRNG - BFR = Trésorerie  [Vérification automatique]",
        frng - bfr_net)

    # ── MÉMO CAF ─────────────────────────────────────────────────────────
    TITRE("MÉMO — DÉTAIL DE LA CAF")
    L("C1", "Résultat net",                             res_net)
    L("C2", "(+) Dotations amortissements & provisions", dotations)
    L("C3", "(-) Reprises amortissements & provisions", reprises)
    L("C4", "(-) Plus-value nette de cession",          pv_nette)
    TOT("C5", "CAF (C1+C2-C3-C4)",                     caf)

    # Mise en forme
    result = []
    for r in lignes:
        result.append({
            "REF":        r["REF"],
            "POSTE":      r["POSTE"],
            "NOTE":       r.get("NOTE", ""),
            "MONTANT N":  _fmt(r["VAL"]) if r.get("VAL") is not None else "",
        })
    return pd.DataFrame(result, columns=["REF", "POSTE", "NOTE", "MONTANT N"])


# =============================================================================
# TABLEAUX ANNEXES — Notes A à E
# =============================================================================

def generer_notes_annexes(df: pd.DataFrame) -> dict:
    """
    Génère les tableaux annexes obligatoires SYSCOHADA.
    Retourne un dict de DataFrames :
      'immo'       : Tableau A — État des immobilisations
      'amort'      : Tableau B — Dotations aux amortissements (exercice N)
      'provisions' : Tableau C — État des provisions et dépréciations
      'creances'   : Tableau D — État des créances
      'dettes'     : Tableau E — État des dettes
    """

    # ── TABLEAU A : ÉTAT DES IMMOBILISATIONS ─────────────────────────────
    immo_config = [
        ("AA", "Charges immobilisées",                   ["20"],         ["280"]),
        ("AB", "Frais de R&D",                           ["211"],        ["2811"]),
        ("AC", "Brevets, licences, logiciels",           ["212","213"],  ["2812","2813"]),
        ("AD", "Fonds commercial",                       ["215"],        ["2815"]),
        ("AE", "Autres immob. incorporelles",            ["218"],        ["2818"]),
        ("AF", "Terrains",                               ["22"],         ["282"]),
        ("AG", "Bâtiments",                              ["231","232"],  ["2831","2832"]),
        ("AH", "Aménagements, agencements, installations",["233","234"], ["2833","2834"]),
        ("AI", "Matériel, mobilier et actifs biologiques",["24"],        ["284"]),
        ("AJ", "Matériel de transport",                  ["245"],        ["2845"]),
        ("AK", "Avances et acomptes/immobilisations",    ["251","252"],  []),
        ("AL", "Titres de participation",                ["261","262"],  ["2961","2962"]),
        ("AM", "Autres immob. financières",              ["27"],         ["297"]),
    ]
    rows_immo, tot_b, tot_a, tot_n = [], 0, 0, 0
    for ref, lib, pfx_b, pfx_a in immo_config:
        b, a, n = _net(df, pfx_b, pfx_a)
        rows_immo.append({"REF": ref, "CATÉGORIE": lib,
                          "BRUT N": _fmt(b), "AMORT CUMULÉS N": _fmt(a), "NET N": _fmt(n)})
        tot_b += b; tot_a += a; tot_n += n
    rows_immo.append({"REF": "AZ", "CATÉGORIE": "TOTAL ACTIF IMMOBILISÉ",
                      "BRUT N": _fmt(tot_b), "AMORT CUMULÉS N": _fmt(tot_a), "NET N": _fmt(tot_n)})
    df_immo = pd.DataFrame(rows_immo, columns=["REF","CATÉGORIE","BRUT N","AMORT CUMULÉS N","NET N"])

    # ── TABLEAU B : DOTATIONS AUX AMORTISSEMENTS (Exercice N) ────────────
    amort_config = [
        ("681", "Dotations immob. incorporelles",         ["681"]),
        ("682", "Dotations immob. corporelles",           ["682"]),
        ("683", "Dotations immob. financières",           ["683"]),
        ("684", "Dotations actif circulant (stocks)",     ["684"]),
        ("685", "Dotations provisions financières",       ["685"]),
        ("686", "Dotations risques CT",                   ["686"]),
        ("687", "Dotations charges à répartir",           ["687"]),
        ("688", "Dotations autres amortissements",        ["688"]),
        ("851", "Dotations HAO",                          ["851","861"]),
    ]
    rows_amort, tot_dot = [], 0
    for compte, lib, pfx in amort_config:
        val = _solde(df, pfx, "D")
        rows_amort.append({"COMPTE": compte, "NATURE": lib, "DOTATION N": _fmt(val)})
        tot_dot += val
    rows_amort.append({"COMPTE": "TOT", "NATURE": "TOTAL DOTATIONS EXERCICE N", "DOTATION N": _fmt(tot_dot)})
    df_amort = pd.DataFrame(rows_amort, columns=["COMPTE","NATURE","DOTATION N"])

    # ── TABLEAU C : PROVISIONS ET DÉPRÉCIATIONS ───────────────────────────
    prov_config = [
        ("15",  "Provisions réglementées (15)",                ["15"],  "C"),
        ("19",  "Provisions financières pour risques (19)",    ["19"],  "C"),
        ("391-396", "Dépréciations des stocks (39x)",         ["391","392","393","394","395","396"], "C"),
        ("491", "Dépréciations créances clients (491)",        ["491"], "C"),
        ("499", "Dépréciations autres créances (499)",         ["499"], "C"),
        ("590", "Dépréciations titres de placement (590)",     ["590"], "C"),
        ("296-297", "Dépréciations immob. financières (29x)", ["296","297","2961","2962"], "C"),
    ]
    rows_prov, tot_prov = [], 0
    for ref, lib, pfx, sens in prov_config:
        val = _solde(df, pfx, sens)
        rows_prov.append({"COMPTE": ref, "NATURE": lib, "SOLDE N": _fmt(val)})
        tot_prov += val
    rows_prov.append({"COMPTE": "TOT", "NATURE": "TOTAL PROVISIONS ET DÉPRÉCIATIONS", "SOLDE N": _fmt(tot_prov)})
    df_prov = pd.DataFrame(rows_prov, columns=["COMPTE","NATURE","SOLDE N"])

    # ── TABLEAU D : ÉTAT DES CRÉANCES ─────────────────────────────────────
    creances_config = [
        ("41x", "Clients et comptes rattachés",         ["41"],         ["491"]),
        ("42x", "Personnel — créances",                 ["42"],         []),
        ("44x", "État et organismes sociaux — créances",["44"],         []),
        ("45x", "Groupe / associés — créances",         ["45"],         ["495"]),
        ("46x", "Débiteurs divers",                     ["46"],         []),
        ("471", "Comptes de régularisation actif",      ["471","476"],  []),
        ("50x", "Valeurs mob. de placement",            ["50","51"],    ["590"]),
        ("52-57","Banques, caisse",                     ["52","53","54","57"], []),
    ]
    rows_cr, tot_cr = [], 0
    for compte, lib, pfx_d, pfx_dep in creances_config:
        brut = _solde(df, pfx_d, "D")
        dep  = _solde(df, pfx_dep, "C") if pfx_dep else 0
        net  = max(brut - dep, 0)
        rows_cr.append({"COMPTE": compte, "NATURE": lib,
                        "BRUT": _fmt(brut), "DÉPRÉC.": _fmt(dep), "NET": _fmt(net)})
        tot_cr += net
    rows_cr.append({"COMPTE": "TOT", "NATURE": "TOTAL CRÉANCES",
                    "BRUT": "", "DÉPRÉC.": "", "NET": _fmt(tot_cr)})
    df_creances = pd.DataFrame(rows_cr, columns=["COMPTE","NATURE","BRUT","DÉPRÉC.","NET"])

    # ── TABLEAU E : ÉTAT DES DETTES ───────────────────────────────────────
    dettes_config = [
        ("161-168", "Emprunts et dettes financières MLT",         ["161","162","163","164","165","166","167","168"]),
        ("401-408", "Fournisseurs et comptes rattachés",          ["401","402","408"]),
        ("421-428", "Dettes envers le personnel",                 ["421","422","423","424","425","426","427","428"]),
        ("441-448", "Dettes fiscales",                            ["441","442","443","444","445"]),
        ("451-458", "Dettes groupe et associés",                  ["451","452","453","454"]),
        ("461-468", "Créditeurs divers",                          ["461","462","463","464","465"]),
        ("481-482", "Fournisseurs d'investissement",             ["481","482"]),
        ("56x",     "Concours bancaires courants (56x)",          ["561","562","563","564","565"]),
    ]
    rows_det, tot_det = [], 0
    for compte, lib, pfx in dettes_config:
        val = _solde(df, pfx, "C")
        rows_det.append({"COMPTE": compte, "NATURE": lib, "SOLDE N": _fmt(val)})
        tot_det += val
    rows_det.append({"COMPTE": "TOT", "NATURE": "TOTAL DETTES", "SOLDE N": _fmt(tot_det)})
    df_dettes = pd.DataFrame(rows_det, columns=["COMPTE","NATURE","SOLDE N"])

    return {
        "immo":       df_immo,
        "amort":      df_amort,
        "provisions": df_prov,
        "creances":   df_creances,
        "dettes":     df_dettes,
    }

# =============================================================================
# POINT D'ENTRÉE PRINCIPAL
# =============================================================================

def generer_liasse_sn(df_balance: pd.DataFrame, entreprise: str = "", exercice: str = "") -> dict:
    """
    Génère tous les états financiers DGID Sénégal depuis une balance SYSCOHADA.

    Retourne un dict :
    {
      'bilan_actif':       DataFrame,
      'bilan_passif':      DataFrame,
      'compte_resultat':   DataFrame,
      'tafire':            DataFrame,
      'notes_annexes':     dict  (immo, amort, provisions, creances, dettes),
      'excel_bytes':       bytes  (7 onglets),
      'totaux': {
          'total_actif', 'total_passif', 'resultat_net', 'ca',
          'caf', 'bfr', 'treso_nette'
      }
    }
    """
    df_ba  = generer_bilan_actif(df_balance)
    df_bp  = generer_bilan_passif(df_balance)
    df_cr  = generer_compte_resultat_dgid(df_balance)
    df_taf = generer_tafire(df_balance)
    notes  = generer_notes_annexes(df_balance)

    def _get_val(df, ref):
        row = df[df['REF'] == ref]
        if row.empty:
            return 0.0
        val_str = str(row.iloc[0, -1]).replace(' ', '').replace(',', '').replace('-', '0')
        try:
            return float(val_str)
        except Exception:
            return 0.0

    totaux = {
        'total_actif':  _get_val(df_ba,  'DZ'),
        'total_passif': _get_val(df_bp,  'HZ'),
        'resultat_net': _get_val(df_cr,  'XI'),
        'ca':           _get_val(df_cr,  'XB'),
        'caf':          _get_val(df_taf, 'ZA'),
        'bfr':          _get_val(df_taf, 'BFR'),
        'treso_nette':  _get_val(df_taf, 'TN'),
    }

    excel_bytes = export_liasse_excel(df_balance, entreprise, exercice)

    return {
        'bilan_actif':     df_ba,
        'bilan_passif':    df_bp,
        'compte_resultat': df_cr,
        'tafire':          df_taf,
        'notes_annexes':   notes,
        'excel_bytes':     excel_bytes,
        'totaux':          totaux,
    }


def page_liasse_officielle():
    import streamlit as st
    import pandas as pd
    from datetime import datetime
    from utils.syscohada_helpers import (
        selectionner_entreprise, charger_balance_avec_ui, valider_structure_balance,
        sauvegarder_si_entreprise, telecharger_html, telecharger_word,
        get_info_pays, get_code_pays,
    )
    from data.plan_comptable_syscohada import PLAN_COMPTABLE, FISCALITE_UEMOA, rechercher_comptes
    from utils.ai import appel_mistral
    from utils.rendu_financier import afficher_rapport
    info_pays = get_info_pays()
    code_pays = get_code_pays()
    st.title("📄 Liasse Fiscale Officielle — DGID Sénégal")
    st.markdown("*États Financiers Annuels (SAES) — Format SYSCOHADA révisé 2017*")
    st.divider()

    # Vérification pays
    if code_pays != 'SN':
        st.warning("⚠ Ce module est actuellement disponible pour le **Sénégal** uniquement. "
                   "Sélectionnez 🇸🇳 Sénégal dans la barre latérale.")
        st.stop()

    ent_id, ent_nom, exercice = selectionner_entreprise("lfo")

    # Init session_state
    for k, v in [('lfo_resultat', None), ('lfo_nom_fichier', None)]:
        if k not in st.session_state:
            st.session_state[k] = v

    st.info("""
    📌 **Ce module génère les 3 états officiels SAES reconnus par la DGID Sénégal :**
    - **Tableau 1** — Bilan Actif (REF AA → DZ)
    - **Tableau 2** — Bilan Passif (REF CA → HZ)
    - **Tableau 3** — Compte de Résultat avec SIG (REF TA/RA → XI)

    ✅ Mapping direct depuis la balance SYSCOHADA — aucune saisie manuelle.
    """)

    fichier = st.file_uploader("📎 Importer la balance SYSCOHADA (Excel ou CSV)",
                               type=["xlsx", "csv"], key="lfo_uploader")
    if fichier:
        if st.session_state.get('lfo_nom_fichier') != fichier.name:
            st.session_state.lfo_resultat = None
            st.session_state.lfo_nom_fichier = fichier.name

        try:
            df_propre, infos_col = charger_balance_avec_ui(fichier, "lfo")

            if df_propre is not None:
                with st.expander("👀 Aperçu de la balance chargée"):
                    st.dataframe(df_propre.head(20), use_container_width=True)
                st.success(f"✅ {len(df_propre)} lignes exploitables — balance prête.")

                if st.button("📄 Générer la Liasse Fiscale Officielle DGID",
                             type="primary", use_container_width=True):
                    with st.spinner("Génération des états financiers SAES en cours..."):
                        logger.info(f"Génération Liasse DGID SN pour {ent_nom or 'N/A'}")
                        liasse = generer_liasse_sn(df_propre, ent_nom, exercice)
                        st.session_state.lfo_resultat = liasse

                if isinstance(st.session_state.lfo_resultat, dict):
                    liasse = st.session_state.lfo_resultat
                    totaux = liasse['totaux']

                    # ── Indicateurs clés ────────────────────────────────
                    st.divider()
                    st.subheader("📊 Indicateurs Clés")
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("Total Actif",    f"{totaux['total_actif']:,.0f} FCFA")
                    col2.metric("Total Passif",   f"{totaux['total_passif']:,.0f} FCFA")
                    col3.metric("Chiffre d'affaires", f"{totaux['ca']:,.0f} FCFA")
                    col4.metric("Résultat Net",   f"{totaux['resultat_net']:,.0f} FCFA")
                    col5, col6, col7, _ = st.columns(4)
                    col5.metric("CAF",            f"{totaux.get('caf', 0):,.0f} FCFA")
                    col6.metric("BFR",            f"{totaux.get('bfr', 0):,.0f} FCFA")
                    col7.metric("Trésorerie Nette", f"{totaux.get('treso_nette', 0):,.0f} FCFA")

                    equilibre = abs(totaux['total_actif'] - totaux['total_passif'])
                    if totaux['total_actif'] > 0 and equilibre / max(totaux['total_actif'], 1) < 0.01:
                        st.success("✅ Bilan équilibré — Actif = Passif")
                    elif totaux['total_actif'] > 0:
                        st.warning(f"⚠ Écart Actif/Passif : {equilibre:,.0f} FCFA — Vérifiez la balance.")

                    # ── Tableau 1 : Bilan Actif ──────────────────────────
                    st.divider()
                    st.subheader("📋 Tableau 1 — Bilan Actif")
                    st.dataframe(
                        liasse['bilan_actif'].style.apply(
                            lambda row: ['background-color: #D6E4F0; font-weight: bold'
                                         if row['REF'] in ['AZ','BZ','CZ','DZ'] else ''
                                         for _ in row], axis=1
                        ),
                        use_container_width=True, hide_index=True
                    )

                    # ── Tableau 2 : Bilan Passif ─────────────────────────
                    st.divider()
                    st.subheader("📋 Tableau 2 — Bilan Passif")
                    st.dataframe(
                        liasse['bilan_passif'].style.apply(
                            lambda row: ['background-color: #D6E4F0; font-weight: bold'
                                         if row['REF'] in ['CZ','DZ','EZ','FZ','GZ','HZ'] else ''
                                         for _ in row], axis=1
                        ),
                        use_container_width=True, hide_index=True
                    )

                    # ── Tableau 3 : Compte de Résultat ───────────────────
                    st.divider()
                    st.subheader("📋 Tableau 3 — Compte de Résultat")
                    sig_refs = ['XA','XB','XC','XD','XE','XF','XG','XH','XI']
                    st.dataframe(
                        liasse['compte_resultat'].style.apply(
                            lambda row: ['background-color: #E8F5E9; font-weight: bold'
                                         if row['REF'] in sig_refs else ''
                                         for _ in row], axis=1
                        ),
                        use_container_width=True, hide_index=True
                    )

                    # ── Tableau 4 : TAFIRE ───────────────────────────────
                    st.divider()
                    st.subheader("📋 Tableau 4 — TAFIRE (Financement des Ressources et Emplois)")
                    st.caption("ℹ Les variations BFR (lignes M1→M8) montrent les positions de l'exercice N. "
                               "Pour les variations N/N-1, importez également la balance N-1.")
                    tafire_refs = ["ZC","ZR","ZE","FRG","M4","M8","BFR","TN","EQ","C5"]
                    st.dataframe(
                        liasse['tafire'].style.apply(
                            lambda row: ['background-color: #D6E4F0; font-weight: bold'
                                         if row['REF'] in tafire_refs else ''
                                         for _ in row], axis=1
                        ),
                        use_container_width=True, hide_index=True
                    )

                    # ── Notes Annexes ─────────────────────────────────────
                    st.divider()
                    st.subheader("📎 Notes Annexes — Tableaux Obligatoires SYSCOHADA")
                    notes = liasse.get('notes_annexes', {})
                    if notes:
                        tab_a, tab_b, tab_c, tab_d, tab_e = st.tabs([
                            "A — Immobilisations",
                            "B — Amortissements",
                            "C — Provisions",
                            "D — Créances",
                            "E — Dettes"
                        ])
                        with tab_a:
                            st.caption("État des immobilisations au bilan (Valeur Brute / Amortissements / Valeur Nette)")
                            st.dataframe(
                                notes['immo'].style.apply(
                                    lambda row: ['background-color: #D6E4F0; font-weight: bold'
                                                 if row['REF'] == 'AZ' else '' for _ in row], axis=1
                                ), use_container_width=True, hide_index=True)
                        with tab_b:
                            st.caption("Dotations aux amortissements de l'exercice N (comptes 681-688, 851, 861)")
                            st.dataframe(
                                notes['amort'].style.apply(
                                    lambda row: ['background-color: #D6E4F0; font-weight: bold'
                                                 if row['COMPTE'] == 'TOT' else '' for _ in row], axis=1
                                ), use_container_width=True, hide_index=True)
                        with tab_c:
                            st.caption("État des provisions et dépréciations (comptes 15, 19, 39x, 49x, 59x, 29x)")
                            st.dataframe(
                                notes['provisions'].style.apply(
                                    lambda row: ['background-color: #D6E4F0; font-weight: bold'
                                                 if row['COMPTE'] == 'TOT' else '' for _ in row], axis=1
                                ), use_container_width=True, hide_index=True)
                        with tab_d:
                            st.caption("État des créances par nature (Brut / Dépréciations / Net)")
                            st.dataframe(
                                notes['creances'].style.apply(
                                    lambda row: ['background-color: #D6E4F0; font-weight: bold'
                                                 if row['COMPTE'] == 'TOT' else '' for _ in row], axis=1
                                ), use_container_width=True, hide_index=True)
                        with tab_e:
                            st.caption("État des dettes par nature (emprunts, fournisseurs, fiscal, social…)")
                            st.dataframe(
                                notes['dettes'].style.apply(
                                    lambda row: ['background-color: #D6E4F0; font-weight: bold'
                                                 if row['COMPTE'] == 'TOT' else '' for _ in row], axis=1
                                ), use_container_width=True, hide_index=True)

                    # ── Téléchargement Excel ──────────────────────────────
                    st.divider()
                    st.subheader("📥 Télécharger la Liasse Officielle")
                    col1, col2 = st.columns(2)
                    with col1:
                        nom_fichier = f"Liasse_DGID_SN_{ent_nom or 'Entreprise'}_{exercice or '2024'}.xlsx"
                        st.download_button(
                            label="📊 Télécharger Excel DGID (Bilan + CR + TAFIRE + Notes)",
                            data=liasse['excel_bytes'],
                            file_name=nom_fichier,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
                    with col2:
                        if st.button("💾 Sauvegarder dans dossier entreprise",
                                     use_container_width=True):
                            sauvegarder_si_autorise(
                                ent_id, "📄 Liasse DGID", f"Liasse DGID {exercice}",
                                f"Liasse DGID générée — CA: {totaux['ca']:,.0f} | RN: {totaux['resultat_net']:,.0f}",
                                info_pays['nom'], exercice
                            )

        except Exception as e:
            st.error(f"❌ Erreur : {e}")
            logger.error(f"Erreur Liasse DGID SN : {e}")

    # =============================================================================
    # PAGE : LIASSE FISCALE IA
    # =============================================================================
    # ── Helpers SYSCOHADA ──────────────────────────────────────────────────
    from utils.syscohada_helpers import (
        selectionner_entreprise, charger_balance_avec_ui, valider_structure_balance,
        sauvegarder_si_entreprise, telecharger_html, telecharger_word,
        get_info_pays, get_code_pays,
    )
    from data.plan_comptable_syscohada import PLAN_COMPTABLE, FISCALITE_UEMOA, rechercher_comptes
    from utils.ai import appel_mistral
    from utils.rendu_financier import afficher_rapport
    info_pays  = get_info_pays()
    code_pays  = get_code_pays()
